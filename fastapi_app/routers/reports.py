import csv
import io
import re
from datetime import date as date_type, datetime, timedelta, timezone

from fastapi import APIRouter, Depends, HTTPException, Query, Response, status
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from sqlalchemy import and_
from sqlalchemy.orm import Session

from auth import get_current_user
from database import get_db
from models import DailyReport, Department, User
from schemas import LeaveIn, ReportIn, ReportListOut, ReportOut

router = APIRouter(prefix="/api/reports", tags=["reports"])

# Excel sheet-name rules: max 31 chars, no  : \ / ? * [ ]
_INVALID_SHEET_CHARS = set(r":\/?*[]")


def _sanitize_sheet_name(name: str) -> str:
    cleaned = "".join("_" if c in _INVALID_SHEET_CHARS else c for c in (name or "Sheet"))
    cleaned = cleaned.strip() or "Sheet"
    return cleaned[:31]


def _unique_sheet_name(wb: Workbook, base: str) -> str:
    name = _sanitize_sheet_name(base)
    existing = set(wb.sheetnames)
    if name not in existing:
        return name
    # Append (2), (3), ... while keeping total <= 31 chars.
    n = 2
    while True:
        suffix = f" ({n})"
        candidate = name[: 31 - len(suffix)] + suffix
        if candidate not in existing:
            return candidate
        n += 1


def _summarise_field(values_with_dates):
    """Combine multiple per-day cells into ONE summary cell.

    `values_with_dates` is a list of `(date, raw_value)` tuples for the same
    employee + field across the date range.

    Heuristic:
      - Drop empty values and leave markers.
      - If every remaining value parses as a number (allowing ₹, commas,
        whitespace), return the numeric sum — int if it lands on a whole
        number, otherwise rounded to 2 dp.
      - Otherwise, concatenate `date: value` lines so HR can still see what
        was filled when.
    """
    pairs: list[tuple] = []
    for d, v in values_with_dates:
        s = ("" if v is None else str(v)).strip()
        if not s or s.lower() == "on leave":
            continue
        pairs.append((d, s))
    if not pairs:
        return ""

    nums: list[float] = []
    is_numeric = True
    for _, s in pairs:
        cleaned = s.replace("₹", "").replace("$", "").replace(",", "").replace(" ", "")
        try:
            nums.append(float(cleaned))
        except ValueError:
            is_numeric = False
            break

    if is_numeric:
        total = sum(nums)
        return int(total) if total == int(total) else round(total, 2)

    return "\n".join(
        f"{(d.isoformat() if d else '-')}: {s}" for d, s in pairs
    )


@router.get("", response_model=ReportListOut)
def list_reports(
    employee: int | None = Query(None, description="Filter by employee id"),
    department: str | None = Query(None, description="Filter by department slug"),
    start: date_type | None = Query(None, description="Inclusive start date"),
    end: date_type | None = Query(None, description="Inclusive end date"),
    limit: int = Query(1000, ge=1, le=5000, description="Page size"),
    offset: int = Query(0, ge=0, description="Rows to skip"),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    q = db.query(DailyReport)
    if employee is not None:
        q = q.filter(DailyReport.user_id == employee)
    if department:
        dept = db.query(Department).filter(Department.slug == department).first()
        if not dept:
            return {"items": [], "total": 0, "limit": limit, "offset": offset}
        q = q.join(User, DailyReport.user_id == User.id).filter(User.department_id == dept.id)
    if start:
        q = q.filter(DailyReport.date >= start)
    if end:
        q = q.filter(DailyReport.date <= end)

    total = q.count()
    items = (
        q.order_by(DailyReport.date.desc(), DailyReport.user_id)
        .offset(offset)
        .limit(limit)
        .all()
    )
    return {"items": items, "total": total, "limit": limit, "offset": offset}


# ---------- Styled XLSX export ----------
#
# Match the look of the manual "All Department Employee Summary Report" file:
# 16-pt title row, grey subtitle, navy-blue header band with white text, light
# zebra fill on data, blue badge for the department column, bold employee
# names, em-dashes in place of empty values.  See _build_styled_sheet for the
# per-sheet layout.

_DASH = "—"

# Colour palette mirrors the on-screen "summary table" component so the
# downloaded XLSX is visually identical to the dashboard view.
_TITLE_FONT = Font(bold=True, size=16, color="1B5E8B")          # blue title
_SUBTITLE_FONT = Font(size=10, color="666666", italic=True)
_HEADER_FONT = Font(bold=True, size=10, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="1A2A3A")           # navy header
_NAME_FONT = Font(size=10, color="1A1A1A")                       # employee names
_BADGE_FONT = Font(bold=True, size=9, color="1B5E8B")
_BADGE_FILL = PatternFill("solid", fgColor="D6E8F5")
_BODY_FONT = Font(size=10, color="1A1A1A")
_MUTED_FONT = Font(size=10, color="888888")
_ROW_FILL = PatternFill("solid", fgColor="FFFFFF")              # white rows
_TOTAL_FONT = Font(bold=True, size=10, color="1B5E8B")          # navy bold totals
_TOTAL_FILL = PatternFill("solid", fgColor="EAF1F8")            # subtle blue total bar
_LEFT_TOP_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
# Thin light-grey border for table cells — gives the layout clear gridlines
# so HR can read rows/columns at a glance.
_CELL_BORDER = Border(
    left=Side(style="thin", color="C9D2DC"),
    right=Side(style="thin", color="C9D2DC"),
    top=Side(style="thin", color="C9D2DC"),
    bottom=Side(style="thin", color="C9D2DC"),
)


# Field-label classifiers: which column to treat as "Meetings" and which as
# "Revenue" in the new compact report layout (mirrors the manual HR report).
_MEETING_LABEL_RE = re.compile(r"\b(meeting|call|visit|enquir)", re.I)
_REVENUE_LABEL_RE = re.compile(
    r"\b(revenue|amount|invoice|sales|rupees|payment|earning|₹)", re.I
)
# Broad numeric-label hint — any column whose label reads like a counter or
# amount.  Used by the per-dept detail sheet to decide which columns to sum.
_NUMERIC_LABEL_RE = re.compile(
    r"\b(no\.?|number|count|total|amount|revenue|calls?|meetings?|enquiries|"
    r"leads?|companies|visits?|orders?|hours?|sum|rate|₹|rs\.?|inr|amt|pis?|"
    r"picked|closed|lost|following|invoice|sent|shared|received|made|type|works?)\b",
    re.I,
)
# Number token: handles thousand-separator commas, optional decimals, and
# detects trailing ordinal suffix so "25th" / "1st" can be skipped.
_NUMBER_TOKEN_RE = re.compile(
    r"(\d{1,3}(?:,\d{3})+(?:\.\d+)?|\d+(?:\.\d+)?)(?:\s*(st|nd|rd|th)\b)?",
    re.I,
)


def _extract_numbers_text(s):
    """Pull every number out of a free-text cell.  Handles thousand-separator
    commas, skips year-looking integers (1900-2100), skips ordinal dates
    ("25th", "1st").  So "Monthly Sales- 1,287.32" → [1287.32], and
    "4, 2, 25th May 2026, 3" → [4, 2, 3].
    """
    if s is None:
        return []
    out = []
    for m in _NUMBER_TOKEN_RE.finditer(str(s)):
        if m.group(2):
            continue  # ordinal suffix — skip
        try:
            n = float(m.group(1).replace(",", ""))
        except ValueError:
            continue
        if n.is_integer() and 1900 <= int(n) <= 2100:
            continue  # year-like — skip
        out.append(n)
    return out


# Department display order: Sales first, Inside Sales second, then every
# other department alphabetically.  Applied to both the Summary sheet and
# the per-department sheet creation order so HR sees the same priority in
# both surfaces.
_DEPT_PRIORITY = {"sales": 0, "insideSales": 1}


def _dept_sort_key(group):
    d = group["dept"]
    slug = getattr(d, "slug", "") if d else ""
    name = (d.name if d else "No Department").lower()
    return (_DEPT_PRIORITY.get(slug, 99), name)


def _condense_summary(text, *, max_chars: int = 200) -> str:
    """Shrink a long, multi-paragraph block into a short, readable line.

    Produces clean prose with NO truncation markers — no "…", no
    "(+N more)".  Each surviving phrase is whole; if a phrase doesn't fit
    the budget, it's dropped entirely so the cell stays clean.

    Strategy (deterministic, no LLM):
      1. Split on newlines / periods / semicolons / bullets so each
         "thought" becomes its own fragment.
      2. Strip ISO date prefixes (`2026-05-04:`) and leading list numbers
         (`1. `, `2) `) so fragments compare cleanly.
      3. Dedupe by **keyword overlap** — two fragments are considered
         the same when ≥ 50% of their content words match.  Catches near
         duplicates like "Visit Kamala site" and "visit kamala site again".
      4. For each surviving fragment, keep only the first 5–6 words so it
         reads like a tight headline ("Module installation at Mavikalan").
      5. Pack the headlines into `max_chars`, comma-joined.  Stop cleanly
         when the next phrase would overflow — no trailing marker.
    """
    if not text:
        return ""
    raw = re.split(r"[\n.;•]+|(?<=\s)[-*]\s+", str(text))
    seen_keywords: list[frozenset[str]] = []
    headlines: list[str] = []
    for frag in raw:
        s = frag.strip()
        if not s:
            continue
        s = re.sub(r"^\d{4}-\d{2}-\d{2}\s*[:\-]?\s*", "", s)
        s = re.sub(r"^\d+[.\)]\s*", "", s)
        s = s.strip(" ,")
        if not s:
            continue
        # Keyword set = words ≥ 3 chars, lowercased — used for dedupe.
        kw = frozenset(re.findall(r"[a-z]{3,}", s.lower()))
        if not kw:
            continue
        is_dup = False
        for prev in seen_keywords:
            shared = len(kw & prev)
            if shared and shared >= max(2, min(len(kw), len(prev)) // 2):
                is_dup = True
                break
        if is_dup:
            continue
        seen_keywords.append(kw)
        # Keep only the first 6 words so each headline is tight.
        words = s.split()
        head = " ".join(words[:6]) if len(words) > 6 else s
        headlines.append(head)
    return _compact_join(headlines, max_chars=max_chars)


def _compact_join(values, *, max_chars: int = 120) -> str:
    """Pack whole phrases (comma-joined) into `max_chars`, stopping
    cleanly when the next phrase wouldn't fit.  NO "…", NO "(+N more)".
    """
    if not values:
        return ""
    out = ""
    for v in values:
        s = str(v or "").strip()
        if not s:
            continue
        candidate = s if not out else f"{out}, {s}"
        if len(candidate) > max_chars:
            # Stop — but if we haven't taken anything yet, take one whole
            # phrase even if it slightly exceeds the budget (so the cell
            # isn't empty).
            if not out:
                return s
            break
        out = candidate
    return out


def _format_revenue(n):
    """Render a revenue value in lakhs — `₹X.XX L`.

    1 lakh = 100,000 rupees.  Two decimals are always shown for sub-lakh
    precision, with thousands separators on the lakh portion.
    """
    if n is None:
        return _DASH
    try:
        amount = float(n)
    except (TypeError, ValueError):
        return _DASH
    lakhs = amount / 100000.0
    return f"₹{lakhs:,.2f} L"


def _summarise_user_reports(reports):
    """Condense one employee's set of daily reports into a single summary row:

      (full_name, role, key_activities, meetings_total, revenue_total, n_reports)

    `key_activities` is a multi-line "Field Label: distinct values…" string
    that consolidates every text/free-form field for the date range.  Fields
    whose label looks like a meeting/call counter feed the `meetings_total`
    bucket; fields that look like revenue feed `revenue_total`.  Both
    numbers are `None` when no values were filled.
    """
    user = reports[0].user
    full_name = ""
    role = ""
    if user:
        full_name = (
            f"{(user.first_name or '').strip()} {(user.last_name or '').strip()}".strip()
            or user.username
        )
        role = (user.title or "").strip()

    dept = user.department if user else None
    fields = list(dept.report_fields) if dept and dept.report_fields else []

    meetings_total = None
    revenue_total = None
    # Pool every non-empty text value from every text/non-counter field into
    # one big buffer, then run a single `_condense_summary` pass over it.
    # Result: Key Activities reads as ONE unified summary of what the
    # employee actually did — no "Field Name: …" prefixes.  Meeting and
    # revenue columns are still routed into their own numeric buckets
    # (Meetings / Revenue) so they don't show up twice.
    pooled_text_parts: list[str] = []

    for f in fields:
        label = (f.get("label") or f.get("key") or "").strip()
        key = f.get("key") or ""
        non_empty = []
        for r in reports:
            v = (r.data or {}).get(key, "")
            s = ("" if v is None else str(v)).strip()
            if not s or s.lower() == "on leave":
                continue
            non_empty.append((r.date, s))
        if not non_empty:
            continue

        if _MEETING_LABEL_RE.search(label) and meetings_total is None:
            total = sum(n for _, v in non_empty for n in _extract_numbers_text(v))
            meetings_total = total if total > 0 else None
            continue
        if _REVENUE_LABEL_RE.search(label) and revenue_total is None:
            total = sum(n for _, v in non_empty for n in _extract_numbers_text(v))
            revenue_total = total if total > 0 else None
            continue

        # Pure-numeric counter fields contribute their numbers to the pool as
        # a short phrase, but don't bloat the narrative with raw digits.
        # Text/narrative fields are added verbatim.  `_condense_summary`
        # below dedupes near-duplicates by keyword overlap so we can safely
        # dump everything in here.
        for _, v in non_empty:
            pooled_text_parts.append(v)

    big_block = "\n".join(pooled_text_parts)
    key_activities = _condense_summary(big_block, max_chars=220)
    return full_name, role, key_activities, meetings_total, revenue_total, len(reports)


@router.get("/export.xlsx")
def export_xlsx(
    employee: int | None = Query(None, description="Filter by employee id"),
    department: str | None = Query(None, description="Filter by department slug"),
    start: date_type | None = Query(None, description="Inclusive start date"),
    end: date_type | None = Query(None, description="Inclusive end date"),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """Styled Excel export — Summary sheet first, one sheet per department.

    Matches HR's manual report layout: title row, subtitle, navy header band,
    bold employee names, department badge, light row fill, em-dashes for
    empty cells.  Numeric report fields are summed across the date range;
    text fields are concatenated as `date: value` lines.  Each employee
    appears on exactly one row per sheet.
    """
    q = db.query(DailyReport).join(User, DailyReport.user_id == User.id)
    if employee is not None:
        q = q.filter(DailyReport.user_id == employee)
    if department:
        dept = db.query(Department).filter(Department.slug == department).first()
        if not dept:
            return _empty_xlsx("daily-reports.xlsx")
        q = q.filter(User.department_id == dept.id)
    if start:
        q = q.filter(DailyReport.date >= start)
    if end:
        q = q.filter(DailyReport.date <= end)

    rows = (
        q.order_by(DailyReport.date.desc(), DailyReport.user_id)
        .limit(50000)
        .all()
    )

    # Group by (department, user).  Skip leave-day rows from the rollups so
    # totals don't get contaminated by "On Leave" placeholders.
    by_dept: dict[int | None, dict] = {}
    for r in rows:
        if (r.data or {}).get("__leave__") == "1":
            continue
        u = r.user
        d = u.department if u else None
        key = d.id if d else None
        if key not in by_dept:
            by_dept[key] = {"dept": d, "users": {}}
        by_dept[key]["users"].setdefault(u.id if u else 0, []).append(r)

    wb = Workbook()
    wb.remove(wb.active)

    range_label = _format_range_label(start, end)
    today_label = date_type.today().strftime("%d %b %Y")
    total_reports = sum(
        len(rs) for grp in by_dept.values() for rs in grp["users"].values()
    )
    total_employees = sum(len(grp["users"]) for grp in by_dept.values())
    dept_count = len([g for g in by_dept.values() if g["dept"] is not None])

    # Three-sheet layout — matches HR's manual Numbers report:
    #   1. Sales — Detail        (per-field columns, only Sales employees)
    #   2. Inside Sales — Detail (per-field columns, only Inside Sales)
    #   3. Detailed Summary      (all OTHER departments, one row per
    #                             employee with consolidated Key Activities)
    sales_group = None
    sales_head_group = None
    sales_service_group = None
    inside_sales_group = None
    service_group = None
    project_group = None
    marketing_group = None
    production_group = None
    logistics_group = None
    design_group = None
    procurement_group = None
    reception_group = None
    hr_group = None
    om_group = None          # Operations & Maintenance
    finance_group = None
    rd_bess_group = None     # R&D BESS
    web_dev_group = None
    other_groups = []
    for group in by_dept.values():
        slug = getattr(group["dept"], "slug", "") if group["dept"] else ""
        name = (getattr(group["dept"], "name", "") if group["dept"] else "") or ""
        slug_l = (slug or "").lower()
        name_l = name.lower()
        # "Sales Head" / "Sales Enquiry" share the same detail tab — they're
        # the same team, just renamed.  Keep both old + new detections so
        # existing DB rows with either slug continue to route correctly.
        is_sales_head = (
            slug_l in (
                "saleshead", "sales_head", "sales-head",
                "salesenquiry", "sales_enquiry", "sales-enquiry",
            )
            or ("sales" in name_l and "head" in name_l)
            or ("sales" in name_l and "enquir" in name_l)
        )
        is_reception = (
            slug_l in ("reception", "frontoffice", "front_office", "front-office")
            or "reception" in name_l
            or "front office" in name_l
        )
        is_hr = (
            slug_l in ("hr", "humanresources", "human_resources", "human-resources")
            or name_l in ("hr", "human resources")
            or "human resources" in name_l
        )
        # O&M (Operations & Maintenance) — slug may use any of these forms.
        is_om = (
            slug_l in (
                "om", "oandm", "o_and_m", "o-m", "operations",
                "operationsmaintenance", "operations_maintenance",
            )
            or name_l in ("o&m", "o & m", "operations & maintenance",
                          "operations and maintenance")
            or ("operation" in name_l and "maintenance" in name_l)
        )
        is_finance = (
            slug_l in ("finance", "accounts", "accounting", "fin")
            or name_l in ("finance", "accounts", "accounting")
        )
        # R&D BESS — Research & Development for Battery Energy Storage Systems.
        is_rd_bess = (
            slug_l in (
                "rdbess", "rd_bess", "rd-bess", "r_and_d_bess",
                "rndbess", "researchbess",
            )
            or ("r&d" in name_l and "bess" in name_l)
            or ("r & d" in name_l and "bess" in name_l)
            or ("research" in name_l and "bess" in name_l)
            or ("rd" in name_l and "bess" in name_l)
        )
        is_web_dev = (
            slug_l in (
                "webdev", "web_dev", "web-dev", "webdevelopment",
                "web_development", "web-development",
            )
            or name_l in ("web dev", "web development", "webdev")
            or ("web" in name_l and "dev" in name_l)
        )
        if slug == "sales":
            sales_group = group
        elif is_sales_head:
            sales_head_group = group
        elif slug == "salesService" or "sales service" in name_l:
            sales_service_group = group
        elif slug == "insideSales":
            inside_sales_group = group
        elif slug == "service":
            service_group = group
        elif slug == "project":
            project_group = group
        elif slug == "marketing":
            marketing_group = group
        elif slug == "production":
            production_group = group
        elif slug == "logistics":
            logistics_group = group
        elif slug == "design":
            design_group = group
        elif slug == "procurement":
            procurement_group = group
        elif is_reception:
            reception_group = group
        elif is_hr:
            hr_group = group
        elif is_om:
            om_group = group
        elif is_finance:
            finance_group = group
        elif is_rd_bess:
            rd_bess_group = group
        elif is_web_dev:
            web_dev_group = group
        else:
            other_groups.append(group)

    # All Employees roster — first tab so HR sees the headcount snapshot
    # before any per-department breakdown.
    ws = wb.create_sheet(title=_unique_sheet_name(wb, "All Employees"))
    _build_all_employees_sheet(ws, db, range_label=range_label, today_label=today_label)

    if sales_group:
        ws = wb.create_sheet(title=_unique_sheet_name(wb, "Sales — Detail"))
        _build_dept_detail_sheet(ws, sales_group, range_label=range_label)

    if sales_head_group:
        # Tab title follows the department's actual name in the DB — so
        # renaming "Sales Head" → "Sales Enquiry" automatically updates here.
        sh_name = (
            getattr(sales_head_group["dept"], "name", "")
            if sales_head_group.get("dept") else ""
        ) or "Sales Enquiry"
        ws = wb.create_sheet(title=_unique_sheet_name(wb, f"{sh_name} — Detail"))
        _build_dept_detail_sheet(ws, sales_head_group, range_label=range_label)

    if sales_service_group:
        ws = wb.create_sheet(title=_unique_sheet_name(wb, "Sales Service — Detail"))
        _build_sales_service_detail_sheet(
            ws, sales_service_group, range_label=range_label,
        )

    if inside_sales_group:
        ws = wb.create_sheet(title=_unique_sheet_name(wb, "Inside Sales — Detail"))
        # Drop only "Other Works" — Data Called Type / Mail Sent / WhatsApp Sent
        # are kept on the Excel tab even though they're hidden on screen.
        _build_dept_detail_sheet(
            ws, inside_sales_group, range_label=range_label,
            exclude_keys=("otherWorks",),
        )

    if service_group:
        ws = wb.create_sheet(title=_unique_sheet_name(wb, "Service — Detail"))
        _build_service_detail_sheet(ws, service_group, range_label=range_label)

    if project_group:
        ws = wb.create_sheet(title=_unique_sheet_name(wb, "Project — Detail"))
        _build_project_detail_sheet(ws, project_group, range_label=range_label)

    if marketing_group:
        ws = wb.create_sheet(title=_unique_sheet_name(wb, "Marketing — Detail"))
        _build_marketing_detail_sheet(ws, marketing_group, range_label=range_label)

    if production_group:
        ws = wb.create_sheet(title=_unique_sheet_name(wb, "Production — Detail"))
        _build_production_detail_sheet(ws, production_group, range_label=range_label)

    if logistics_group:
        ws = wb.create_sheet(title=_unique_sheet_name(wb, "Logistics — Detail"))
        _build_logistics_detail_sheet(ws, logistics_group, range_label=range_label)

    if design_group:
        ws = wb.create_sheet(title=_unique_sheet_name(wb, "Design — Detail"))
        _build_design_detail_sheet(ws, design_group, range_label=range_label)

    if procurement_group:
        ws = wb.create_sheet(title=_unique_sheet_name(wb, "Procurement — Detail"))
        _build_procurement_detail_sheet(ws, procurement_group, range_label=range_label)

    if reception_group:
        ws = wb.create_sheet(title=_unique_sheet_name(wb, "Reception — Detail"))
        _build_dept_detail_sheet(ws, reception_group, range_label=range_label)

    if hr_group:
        ws = wb.create_sheet(title=_unique_sheet_name(wb, "HR — Detail"))
        _build_hr_detail_sheet(ws, hr_group, range_label=range_label)

    if om_group:
        # Title follows the department's actual name in the DB so "O&M" /
        # "Operations & Maintenance" / any rename surfaces verbatim.
        om_name = (
            getattr(om_group["dept"], "name", "")
            if om_group.get("dept") else ""
        ) or "O&M"
        ws = wb.create_sheet(title=_unique_sheet_name(wb, f"{om_name} — Detail"))
        _build_dept_detail_sheet(ws, om_group, range_label=range_label)

    if finance_group:
        ws = wb.create_sheet(title=_unique_sheet_name(wb, "Finance — Detail"))
        _build_dept_detail_sheet(ws, finance_group, range_label=range_label)

    if rd_bess_group:
        rd_name = (
            getattr(rd_bess_group["dept"], "name", "")
            if rd_bess_group.get("dept") else ""
        ) or "R&D BESS"
        ws = wb.create_sheet(title=_unique_sheet_name(wb, f"{rd_name} — Detail"))
        _build_dept_detail_sheet(ws, rd_bess_group, range_label=range_label)

    if web_dev_group:
        wd_name = (
            getattr(web_dev_group["dept"], "name", "")
            if web_dev_group.get("dept") else ""
        ) or "Web Dev"
        ws = wb.create_sheet(title=_unique_sheet_name(wb, f"{wd_name} — Detail"))
        _build_dept_detail_sheet(ws, web_dev_group, range_label=range_label)

    if other_groups:
        ws = wb.create_sheet(title=_unique_sheet_name(wb, "Detailed Summary"))
        _build_combined_summary_sheet(
            ws, other_groups,
            range_label=range_label,
            today_label=today_label,
        )

    if not wb.sheetnames:
        ws = wb.create_sheet(title="No reports")
        ws["B2"] = "No reports match the current filters."
        ws["B2"].font = _SUBTITLE_FONT

    buf = io.BytesIO()
    wb.save(buf)
    payload = buf.getvalue()

    parts = ["daily-reports"]
    if start:
        parts.append(start.isoformat())
    if end:
        parts.append(end.isoformat())
    if not (start or end):
        parts.append(date_type.today().isoformat())
    filename = "_".join(parts) + ".xlsx"

    return Response(
        content=payload,
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Content-Length": str(len(payload)),
        },
    )


def _format_range_label(start, end) -> str:
    if start and end:
        if start == end:
            return start.strftime("%d %b %Y")
        return f"{start.strftime('%d %b %Y')} → {end.strftime('%d %b %Y')}"
    if start:
        return f"from {start.strftime('%d %b %Y')}"
    if end:
        return f"until {end.strftime('%d %b %Y')}"
    return "all time"


def _put(ws, row, col, value, font=None, fill=None, align=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font is not None:
        cell.font = font
    if fill is not None:
        cell.fill = fill
    if align is not None:
        cell.alignment = align
    return cell


def _merge_title(ws, row, last_col, value, font):
    ws.cell(row=row, column=2, value=value).font = font
    ws.cell(row=row, column=2).alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=last_col)


def _apply_table_borders(ws, *, header_row: int, last_row: int, first_col: int, last_col: int):
    """Paint a thin light-grey border on every cell in the data table region
    (header row + data rows) so the layout reads like a clean table."""
    if last_row < header_row:
        return
    for r in range(header_row, last_row + 1):
        for c in range(first_col, last_col + 1):
            ws.cell(row=r, column=c).border = _CELL_BORDER


def _build_all_employees_sheet(ws, db, *, range_label: str, today_label: str):
    """Roster snapshot of every active employee — matches the on-screen
    All Employees page (sorted by department, then name).  Columns:
      Organisation | S. No. | Name | Department | Reporting Manager |
      Date of Joining | This Week | This Month
    The This Week / This Month cells show "<filled> / <total>" workdays
    derived from the reports filed across the current calendar week /
    month respectively — same logic as the on-screen badges.
    """
    # Departments whose employees are skipped on the roster sheet.  Finance
    # was previously excluded but HR now wants them counted alongside every
    # other reporting team.  R&D stays excluded.
    NON_REPORTING_DEPTS = {"rd"}

    employees = (
        db.query(User)
        .filter(User.is_active.is_(True))
        .filter(User.role != "hr")
        .all()
    )

    # Date windows for the week / month counters.
    today = date_type.today()
    # Current calendar week (Mon → Fri).
    monday = today - timedelta(days=today.weekday())
    friday = monday + timedelta(days=4)
    # Current calendar month.
    month_first = today.replace(day=1)
    if today.month == 12:
        next_first = date_type(today.year + 1, 1, 1)
    else:
        next_first = date_type(today.year, today.month + 1, 1)
    month_last = next_first - timedelta(days=1)
    # Mon-Fri workdays from start-of-month through TODAY — this is the
    # denominator HR wants on the "This Month" cell ("days they SHOULD have
    # filed so far this month").  Full-month workday count is no longer
    # used; the elapsed-so-far figure makes the ratio meaningful before
    # month-end (e.g. "0 / 16" on 22 Jun instead of "0 / 22").
    month_workdays_elapsed = 0
    d = month_first
    while d <= today:
        if d.weekday() < 5:
            month_workdays_elapsed += 1
        d += timedelta(days=1)
    month_label = month_first.strftime("%b").upper()

    # Pull every report in the week + month windows in one go.  Include
    # `data` for the month query so we can detect leave rows (`__leave__=1`).
    week_reports = (
        db.query(DailyReport.user_id, DailyReport.date)
        .filter(DailyReport.date >= monday, DailyReport.date <= friday)
        .all()
    )
    month_reports = (
        db.query(DailyReport.user_id, DailyReport.date, DailyReport.data)
        .filter(DailyReport.date >= month_first, DailyReport.date <= month_last)
        .all()
    )

    # week_submitted_by_user[uid] = highest weekday index reached (Mon=1..Fri=5).
    # Leave rows count as "submitted" for the week badge (the employee did
    # account for the day, even if it was a leave entry).
    week_submitted_by_user: dict[int, int] = {}
    for uid, dt in week_reports:
        if dt.weekday() >= 5:
            continue
        idx = dt.weekday() + 1  # Mon=0→1 ... Fri=4→5
        if idx > week_submitted_by_user.get(uid, 0):
            week_submitted_by_user[uid] = idx

    # month_submitted_by_user[uid] = distinct weekdays filed in the month.
    # leave_count_by_user[uid]     = distinct DAYS the employee was on leave
    # (any day, including weekends — leaves are entered per real calendar day).
    month_dates_by_user: dict[int, set] = {}
    leave_dates_by_user: dict[int, set] = {}
    for uid, dt, data in month_reports:
        is_leave = bool(data) and str(data.get("__leave__", "")) == "1"
        if is_leave:
            leave_dates_by_user.setdefault(uid, set()).add(dt)
            # Leave rows still count toward "filed for this weekday" so the
            # employee isn't double-penalised on the missing list.
        if dt.weekday() >= 5:
            continue
        month_dates_by_user.setdefault(uid, set()).add(dt)
    month_count_by_user = {
        uid: len(dates) for uid, dates in month_dates_by_user.items()
    }
    leave_count_by_user = {
        uid: len(dates) for uid, dates in leave_dates_by_user.items()
    }

    def _emp_dept_name(e):
        return (e.department.name if e.department else "") or ""

    def _full_name(e):
        n = f"{(e.first_name or '').strip()} {(e.last_name or '').strip()}".strip()
        return n or e.username or "—"

    # Sort: NON-MISSING employees first (department A→Z, name A→Z), then
    # the missing/red rows in the same dept+name order at the BOTTOM.  HR
    # asked for the red rows to be visually separated so the filing list
    # stays on top and the laggards collect at the bottom.
    def sort_key(e):
        dn = _emp_dept_name(e)
        is_missing = month_count_by_user.get(e.id, 0) == 0
        return (
            1 if is_missing else 0,
            1 if not dn else 0,
            dn.lower(),
            _full_name(e).lower(),
        )
    employees.sort(key=sort_key)

    # Columns: B Organisation, C S.No, D Name, E Department, F Reporting Manager,
    # G Date of Joining, H This Week, I This Month, J Leaves  → last_col = 10
    last_col = 10

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 14   # Organisation
    ws.column_dimensions["C"].width = 6    # S. No.
    ws.column_dimensions["D"].width = 26   # Name
    ws.column_dimensions["E"].width = 20   # Department
    ws.column_dimensions["F"].width = 22   # Reporting Manager
    ws.column_dimensions["G"].width = 18   # Date of Joining
    ws.column_dimensions["H"].width = 14   # This Week
    ws.column_dimensions["I"].width = 14   # This Month
    ws.column_dimensions["J"].width = 12   # Leaves

    _merge_title(ws, 1, last_col, "All Employees — Roster", _TITLE_FONT)
    ws.row_dimensions[1].height = 28
    subtitle = (
        f"{len(employees)} active reporting employees  |  "
        f"Generated: {today_label}  |  Range: {range_label}"
    )
    _merge_title(ws, 2, last_col, subtitle, _SUBTITLE_FONT)

    headers = [
        "Organisation", "S. No.", "Name", "Department",
        "Reporting Manager", "Date of Joining",
        "This Week", f"This Month ({month_label})",
        f"Leaves ({month_label})",
    ]
    for i, h in enumerate(headers, start=2):
        _put(ws, 4, i, h, font=_HEADER_FONT, fill=_HEADER_FILL, align=_CENTER)
    ws.row_dimensions[4].height = 30

    body_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=False)
    name_font = Font(bold=True, size=10, color="1A1A1A")
    # Highlight scheme for employees who haven't filed any report in the
    # current calendar month — every cell on that row gets a rose-pink fill
    # and a bold dark-red name so it pops at a glance.
    miss_fill = PatternFill("solid", fgColor="FFE4E6")          # rose-100
    miss_name_font = Font(bold=True, size=10, color="9F1239")   # rose-800
    miss_body_font = Font(size=10, color="9F1239")              # rose-800

    row_idx = 5
    serial = 0
    # Track (row_idx, (is_missing, dept_name)) for every rendered row so we
    # can draw a thicker divider on the LAST row of each department group.
    rendered_keys: list[tuple[int, tuple[bool, str]]] = []
    for emp in employees:
        # Skip non-reporting depts (R&D, Finance) — matches the on-screen
        # "Total Employees" headline card scope.
        if emp.department and emp.department.slug in NON_REPORTING_DEPTS:
            continue
        serial += 1
        wk = week_submitted_by_user.get(emp.id, 0)
        mo = month_count_by_user.get(emp.id, 0)
        # "Not filing" = zero distinct days submitted this calendar month.
        is_missing = mo == 0
        rendered_keys.append((row_idx, (is_missing, _emp_dept_name(emp))))
        row_fill = miss_fill if is_missing else _ROW_FILL
        body_f = miss_body_font if is_missing else _BODY_FONT
        name_f = miss_name_font if is_missing else name_font
        _put(ws, row_idx, 2, emp.organisation or "—",
             font=body_f, fill=row_fill, align=body_left)
        _put(ws, row_idx, 3, serial,
             font=body_f, fill=row_fill, align=center_align)
        _put(ws, row_idx, 4, _full_name(emp),
             font=name_f, fill=row_fill, align=body_left)
        _put(ws, row_idx, 5, _emp_dept_name(emp) or _DASH,
             font=body_f if emp.department else _MUTED_FONT,
             fill=row_fill, align=body_left)
        _put(ws, row_idx, 6, emp.reporting_manager or "—",
             font=body_f, fill=row_fill, align=body_left)
        _put(ws, row_idx, 7,
             emp.date_of_joining.strftime("%a, %d %b %Y") if emp.date_of_joining else "—",
             font=body_f, fill=row_fill, align=center_align)
        _put(ws, row_idx, 8, f"{wk} / 5",
             font=body_f, fill=row_fill, align=center_align)
        # "This Month" denominator = workdays elapsed from 1st through today
        # (HR's ask) — so a non-filer reads "0 / 16" on 22 Jun, not "0 / 22".
        _put(ws, row_idx, 9, f"{mo} / {month_workdays_elapsed}",
             font=body_f, fill=row_fill, align=center_align)
        # Leaves (current calendar month) — count of distinct days the
        # employee filed a leave entry (`__leave__=1`).
        lv = leave_count_by_user.get(emp.id, 0)
        _put(ws, row_idx, 10, lv,
             font=body_f, fill=row_fill, align=center_align)
        row_idx += 1

    last_row = row_idx - 1 if serial > 0 else 4
    _apply_table_borders(
        ws, header_row=4, last_row=last_row, first_col=2, last_col=last_col,
    )

    # Draw a thicker divider at the bottom of every department group (and
    # also between the non-missing and missing sections, which falls out of
    # the same key-change check).  Cells keep their existing thin sides; we
    # only swap the bottom side for the thicker one.
    divider_rows: set[int] = set()
    for i in range(len(rendered_keys) - 1):
        if rendered_keys[i][1] != rendered_keys[i + 1][1]:
            divider_rows.add(rendered_keys[i][0])
    divider_side = Side(style="medium", color="6B7280")  # zinc-500
    for r in divider_rows:
        for c in range(2, last_col + 1):
            existing = ws.cell(row=r, column=c).border
            ws.cell(row=r, column=c).border = Border(
                left=existing.left,
                right=existing.right,
                top=existing.top,
                bottom=divider_side,
            )

    ws.freeze_panes = "B5"
    _set_print_landscape_fit(ws)


def _append_daily_reports_section(ws, group, *, start_row: int, dept_label: str | None = None) -> int:
    """Append a 'Daily Reports' block two rows below the summary table.

    Columns are: Date | Employee | one column per report_field of the
    department.  Each row is a single submitted daily report (raw values,
    no aggregation).  Returns the last row used so callers can chain.
    """
    d = group["dept"]
    fields = list(d.report_fields) if d and d.report_fields else []
    n_fields = len(fields)
    last_col = 3 + n_fields  # B Date, C Employee, D.. fields

    def _emp_name(u):
        if not u:
            return ""
        return (
            f"{(u.first_name or '').strip()} {(u.last_name or '').strip()}".strip()
            or (u.username or "")
        )

    all_reports = []
    for _, user_reports in group["users"].items():
        all_reports.extend(user_reports)
    # Group by employee (alphabetical), newest first within each block —
    # so all of one person's reports appear together instead of mixed by date.
    all_reports.sort(key=lambda r: (
        _emp_name(r.user).lower(),
        -(r.date.toordinal() if r.date else 0),
    ))

    def _widen(col_letter, min_width):
        cur = ws.column_dimensions[col_letter].width or 0
        if cur < min_width:
            ws.column_dimensions[col_letter].width = min_width
    _widen("B", 14)
    _widen("C", 22)
    for i in range(n_fields):
        _widen(get_column_letter(4 + i), 24)

    title_row = start_row + 2
    section_title = dept_label or "Daily Reports"
    section_title = f"{section_title} ({len(all_reports)} report{'' if len(all_reports) == 1 else 's'})"
    _merge_title(ws, title_row, last_col, section_title, _TITLE_FONT)
    ws.row_dimensions[title_row].height = 24

    header_row = title_row + 2
    _put(ws, header_row, 2, "Date",
         font=_HEADER_FONT, fill=_HEADER_FILL, align=_CENTER)
    _put(ws, header_row, 3, "Employee",
         font=_HEADER_FONT, fill=_HEADER_FILL, align=_CENTER)
    for i, f in enumerate(fields):
        label = f.get("label") or f.get("key") or ""
        _put(ws, header_row, 4 + i, label,
             font=_HEADER_FONT, fill=_HEADER_FILL, align=_CENTER)
    ws.row_dimensions[header_row].height = 26

    body_left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    num_align = Alignment(horizontal="center", vertical="center", wrap_text=False)

    row_idx = header_row + 1
    if not all_reports:
        _put(ws, row_idx, 2, "No reports in range.",
             font=_MUTED_FONT, fill=_ROW_FILL, align=body_left)
        ws.merge_cells(
            start_row=row_idx, start_column=2,
            end_row=row_idx, end_column=last_col,
        )
        last_row = row_idx
    else:
        # Show the employee name only on the first row of each block, and
        # merge that name cell down across the block so the grouping reads
        # at a glance.  Records the start row for every block so we can
        # merge after the rows are written.
        prev_name = None
        block_start = None
        block_name = None
        block_ranges: list[tuple[str, int, int]] = []
        bold_name_font = Font(bold=True, size=10, color="1A1A1A")
        name_align = Alignment(horizontal="left", vertical="top", wrap_text=True)

        for r in all_reports:
            name = _emp_name(r.user)
            date_str = r.date.strftime("%d %b %Y") if r.date else ""
            _put(ws, row_idx, 2, date_str,
                 font=_BODY_FONT, fill=_ROW_FILL, align=num_align)
            if name != prev_name:
                # New employee block — close the previous one, start fresh.
                if block_start is not None:
                    block_ranges.append((block_name, block_start, row_idx - 1))
                _put(ws, row_idx, 3, name,
                     font=bold_name_font, fill=_ROW_FILL, align=name_align)
                block_start = row_idx
                block_name = name
                prev_name = name
            else:
                # Continuation row — leave the name cell empty so the merge
                # span shows the name once at the top.
                _put(ws, row_idx, 3, "",
                     font=bold_name_font, fill=_ROW_FILL, align=name_align)
            for i, f in enumerate(fields):
                key = f.get("key") or ""
                v = (r.data or {}).get(key, "")
                s = "" if v is None else str(v).strip()
                if not s:
                    _put(ws, row_idx, 4 + i, _DASH,
                         font=_MUTED_FONT, fill=_ROW_FILL, align=_CENTER)
                else:
                    _put(ws, row_idx, 4 + i, s,
                         font=_BODY_FONT, fill=_ROW_FILL, align=body_left)
            row_idx += 1
        # Close the final block.
        if block_start is not None:
            block_ranges.append((block_name, block_start, row_idx - 1))
        last_row = row_idx - 1

        # Merge each employee's name cell vertically across their block.
        for _bname, start_r, end_r in block_ranges:
            if end_r > start_r:
                ws.merge_cells(
                    start_row=start_r, start_column=3,
                    end_row=end_r, end_column=3,
                )
                top_cell = ws.cell(row=start_r, column=3)
                top_cell.font = bold_name_font
                top_cell.alignment = Alignment(
                    horizontal="left", vertical="top", wrap_text=True,
                )

    _apply_table_borders(
        ws, header_row=header_row,
        last_row=last_row, first_col=2, last_col=last_col,
    )
    return last_row


def _set_print_landscape_fit(ws):
    """Configure the sheet so Excel's print/PDF lands on a single landscape
    page wide and lets rows flow to as many pages as needed."""
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)
    ws.print_options.horizontalCentered = True


def _build_summary_sheet(
    ws,
    by_dept,
    *,
    range_label: str,
    today_label: str,
    total_reports: int,
    total_employees: int,
    dept_count: int,
):
    """All-employees summary — 6 columns matching HR's manual report:

    A (margin) | B Employee Name | C Department | D Role | E Key Activities | F Meetings | G Revenue (₹)
    """
    last_col = 7
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 24   # Employee
    ws.column_dimensions["C"].width = 20   # Department
    ws.column_dimensions["D"].width = 22   # Role
    ws.column_dimensions["E"].width = 70   # Key Activities (wide narrative)
    ws.column_dimensions["F"].width = 12   # Meetings
    ws.column_dimensions["G"].width = 16   # Revenue

    _merge_title(
        ws, 1, last_col,
        f"All Department Employee Summary Report — {range_label}",
        _TITLE_FONT,
    )
    ws.row_dimensions[1].height = 30
    subtitle = (
        f"Ornate Solar  |  {total_reports} reports  |  {total_employees} employees  "
        f"|  {dept_count} departments  |  Generated: {today_label}"
    )
    _merge_title(ws, 2, last_col, subtitle, _SUBTITLE_FONT)

    headers = [
        "Employee Name",
        "Department",
        "Role / Designation",
        f"Key Activities ({range_label})",
        "Meetings",
        "Revenue (₹)",
    ]
    for i, h in enumerate(headers, start=2):
        _put(ws, 4, i, h, font=_HEADER_FONT, fill=_HEADER_FILL, align=_CENTER)
    ws.row_dimensions[4].height = 28

    sorted_groups = sorted(by_dept.values(), key=_dept_sort_key)

    row_idx = 5
    for group in sorted_groups:
        d = group["dept"]
        dept_name = d.name if d else "No Department"
        user_entries = []
        for _, user_reports in group["users"].items():
            sorted_reports = sorted(user_reports, key=lambda r: r.date or date_type.min)
            user_entries.append(_summarise_user_reports(sorted_reports))
        user_entries.sort(key=lambda t: t[0].lower())

        for idx, (name, role, key_acts, meetings, revenue, _count) in enumerate(user_entries):
            _put(ws, row_idx, 2, name, font=_NAME_FONT, fill=_ROW_FILL, align=_LEFT_TOP_WRAP)
            if idx == 0:
                _put(ws, row_idx, 3, dept_name, font=_BADGE_FONT, fill=_BADGE_FILL, align=_CENTER)
            else:
                _put(ws, row_idx, 3, "", fill=_ROW_FILL)
            _put(ws, row_idx, 4, role or _DASH,
                 font=_BODY_FONT if role else _MUTED_FONT,
                 fill=_ROW_FILL, align=_LEFT_TOP_WRAP)
            _put(ws, row_idx, 5, key_acts or _DASH,
                 font=_BODY_FONT if key_acts else _MUTED_FONT,
                 fill=_ROW_FILL, align=_LEFT_TOP_WRAP)
            if meetings is not None:
                disp = int(meetings) if float(meetings).is_integer() else round(meetings, 2)
                _put(ws, row_idx, 6, disp, font=_NAME_FONT, fill=_ROW_FILL, align=_CENTER)
            else:
                _put(ws, row_idx, 6, _DASH, font=_MUTED_FONT, fill=_ROW_FILL, align=_CENTER)
            if revenue is not None:
                _put(ws, row_idx, 7, _format_revenue(revenue),
                     font=Font(bold=True, size=10, color="155F00"),
                     fill=_ROW_FILL, align=_CENTER)
            else:
                _put(ws, row_idx, 7, _DASH, font=_MUTED_FONT, fill=_ROW_FILL, align=_CENTER)
            row_idx += 1

    ws.freeze_panes = "B5"
    _set_print_landscape_fit(ws)


def _build_department_sheet(ws, group, *, range_label: str):
    """Per-department sheet — 5 columns matching HR's manual report:

    A (margin) | B Employee | C Role | D Key Activities | E Meetings | F Revenue (₹)
    """
    d = group["dept"]
    dept_name = d.name if d else "No Department"
    last_col = 6

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 24   # Employee
    ws.column_dimensions["C"].width = 22   # Role
    ws.column_dimensions["D"].width = 80   # Key Activities (wide, narrative)
    ws.column_dimensions["E"].width = 12   # Meetings
    ws.column_dimensions["F"].width = 16   # Revenue

    _merge_title(
        ws, 1, last_col,
        f"{dept_name} — Employee Report ({range_label})",
        _TITLE_FONT,
    )
    ws.row_dimensions[1].height = 30
    n_users = len(group["users"])
    n_reports = sum(len(rs) for rs in group["users"].values())
    subtitle = (
        f"{n_users} {'employee' if n_users == 1 else 'employees'}  ·  "
        f"{n_reports} reports submitted"
    )
    _merge_title(ws, 2, last_col, subtitle, _SUBTITLE_FONT)

    headers = ["Employee", "Role", f"Key Activities ({range_label})", "Meetings", "Revenue (₹)"]
    for i, h in enumerate(headers, start=2):
        _put(ws, 4, i, h, font=_HEADER_FONT, fill=_HEADER_FILL, align=_CENTER)
    ws.row_dimensions[4].height = 28

    user_rows = []
    for _, user_reports in group["users"].items():
        sorted_reports = sorted(user_reports, key=lambda r: r.date or date_type.min)
        name, role, key_acts, meetings, revenue, count = _summarise_user_reports(sorted_reports)
        user_rows.append({
            "name": name, "role": role, "key_acts": key_acts,
            "meetings": meetings, "revenue": revenue, "count": count,
        })
    user_rows.sort(key=lambda r: r["name"].lower())

    row_idx = 5
    for row in user_rows:
        _put(ws, row_idx, 2, row["name"], font=_NAME_FONT, fill=_ROW_FILL, align=_LEFT_TOP_WRAP)
        _put(ws, row_idx, 3, row["role"] or _DASH,
             font=_BODY_FONT if row["role"] else _MUTED_FONT,
             fill=_ROW_FILL, align=_LEFT_TOP_WRAP)
        _put(ws, row_idx, 4, row["key_acts"] or _DASH,
             font=_BODY_FONT if row["key_acts"] else _MUTED_FONT,
             fill=_ROW_FILL, align=_LEFT_TOP_WRAP)
        if row["meetings"] is not None:
            disp = int(row["meetings"]) if float(row["meetings"]).is_integer() else round(row["meetings"], 2)
            _put(ws, row_idx, 5, disp, font=_NAME_FONT, fill=_ROW_FILL, align=_CENTER)
        else:
            _put(ws, row_idx, 5, _DASH, font=_MUTED_FONT, fill=_ROW_FILL, align=_CENTER)
        if row["revenue"] is not None:
            _put(ws, row_idx, 6, _format_revenue(row["revenue"]),
                 font=Font(bold=True, size=10, color="155F00"),
                 fill=_ROW_FILL, align=_CENTER)
        else:
            _put(ws, row_idx, 6, _DASH, font=_MUTED_FONT, fill=_ROW_FILL, align=_CENTER)
        row_idx += 1

    # Department total row.
    if user_rows:
        _put(ws, row_idx, 2, "Department Total",
             font=_TOTAL_FONT, fill=_TOTAL_FILL, align=_LEFT_TOP_WRAP)
        _put(ws, row_idx, 3,
             f"{len(user_rows)} {'employee' if len(user_rows) == 1 else 'employees'}",
             font=_TOTAL_FONT, fill=_TOTAL_FILL, align=_LEFT_TOP_WRAP)
        _put(ws, row_idx, 4, f"{n_reports} reports submitted",
             font=_TOTAL_FONT, fill=_TOTAL_FILL, align=_LEFT_TOP_WRAP)
        meetings_sum = sum(r["meetings"] or 0 for r in user_rows)
        if any(r["meetings"] is not None for r in user_rows):
            disp = int(meetings_sum) if float(meetings_sum).is_integer() else round(meetings_sum, 2)
            _put(ws, row_idx, 5, disp, font=_TOTAL_FONT, fill=_TOTAL_FILL, align=_CENTER)
        else:
            _put(ws, row_idx, 5, _DASH, font=_MUTED_FONT, fill=_TOTAL_FILL, align=_CENTER)
        revenue_sum = sum(r["revenue"] or 0 for r in user_rows)
        if any(r["revenue"] is not None for r in user_rows):
            _put(ws, row_idx, 6, _format_revenue(revenue_sum),
                 font=Font(bold=True, size=10, color="155F00"),
                 fill=_TOTAL_FILL, align=_CENTER)
        else:
            _put(ws, row_idx, 6, _DASH, font=_MUTED_FONT, fill=_TOTAL_FILL, align=_CENTER)

    ws.freeze_panes = "B5"
    _set_print_landscape_fit(ws)


def _build_dept_detail_sheet(ws, group, *, range_label: str, exclude_keys=()):
    """Per-field columns layout for a single department.  Used for Sales and
    Inside Sales detail tabs — Employee + one column per report_field.
    Numeric fields are summed; text fields are joined as distinct values.

    `exclude_keys` skips named fields entirely (used by Inside Sales to drop
    Data Called Type / Mail Sent / WhatsApp Sent / Other Works columns).
    """
    d = group["dept"]
    dept_name = d.name if d else "No Department"
    fields = list(d.report_fields) if d and d.report_fields else []
    if exclude_keys:
        skip = set(exclude_keys)
        fields = [f for f in fields if f.get("key") not in skip]
    n_fields = len(fields)
    last_col = 2 + n_fields  # A margin + B Employee + N field cols

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 22  # Employee
    for i in range(n_fields):
        ws.column_dimensions[get_column_letter(3 + i)].width = 22

    _merge_title(
        ws, 1, last_col,
        f"{dept_name} — summary table",
        _TITLE_FONT,
    )
    ws.row_dimensions[1].height = 26
    n_users = len(group["users"])
    n_reports = sum(len(rs) for rs in group["users"].values())
    subtitle = (
        f"{range_label}  |  "
        f"{n_users} {'employee' if n_users == 1 else 'employees'}  ·  "
        f"{n_reports} reports submitted"
    )
    _merge_title(ws, 2, last_col, subtitle, _SUBTITLE_FONT)

    # Header row.
    _put(ws, 4, 2, "Employee", font=_HEADER_FONT, fill=_HEADER_FILL, align=_CENTER)
    for i, f in enumerate(fields):
        label = f.get("label") or f.get("key") or ""
        _put(ws, 4, 3 + i, label,
             font=_HEADER_FONT, fill=_HEADER_FILL, align=_CENTER)
    ws.row_dimensions[4].height = 28

    # Build per-employee rows.
    user_rows = []
    for _, user_reports in group["users"].items():
        sorted_reports = sorted(user_reports, key=lambda r: r.date or date_type.min)
        u = sorted_reports[0].user
        full_name = ""
        if u:
            full_name = (
                f"{(u.first_name or '').strip()} {(u.last_name or '').strip()}".strip()
                or u.username
            )
        cells = {}
        for f in fields:
            key = f.get("key") or ""
            label = f.get("label") or key
            pairs = [(r.date, (r.data or {}).get(key, "")) for r in sorted_reports]
            non_empty = [
                (d_, str(v).strip()) for d_, v in pairs
                if v and str(v).strip() and str(v).strip().lower() != "on leave"
            ]
            if not non_empty:
                cells[key] = (_DASH, "muted")
                continue
            # Decide numeric vs text by label hint OR strict parse.
            if _NUMERIC_LABEL_RE.search(label) or _MEETING_LABEL_RE.search(label) or _REVENUE_LABEL_RE.search(label):
                total = sum(n for _, v in non_empty for n in _extract_numbers_text(v))
                is_money = bool(_REVENUE_LABEL_RE.search(label))
                if is_money:
                    cells[key] = (_format_revenue(total) if total else "₹0", "money" if total else "muted")
                else:
                    disp = int(total) if float(total).is_integer() else round(total, 2)
                    cells[key] = (disp, "num")
            else:
                # Strict-parse check.
                nums, all_num = [], True
                for _, v in non_empty:
                    cleaned = v.replace("₹", "").replace("$", "").replace(",", "").replace(" ", "")
                    try:
                        nums.append(float(cleaned))
                    except ValueError:
                        all_num = False
                        break
                if all_num and nums:
                    total = sum(nums)
                    disp = int(total) if float(total).is_integer() else round(total, 2)
                    cells[key] = (disp, "num")
                else:
                    # Text — compact join: distinct values, headline-only,
                    # trimmed to fit ~120 chars total with a trailing "…"
                    # when truncated.  No "(+N more)" markers.
                    seen, distinct = set(), []
                    for _, v in non_empty:
                        if v.lower() in seen:
                            continue
                        seen.add(v.lower())
                        distinct.append(v)
                    cells[key] = (_compact_join(distinct, max_chars=120), "body")
        user_rows.append({"name": full_name, "cells": cells})
    user_rows.sort(key=lambda r: r["name"].lower())

    # Matches the on-screen "Inside Sales — summary table" look: clean data
    # rows (no green tint on money), CENTER-aligned numbers, navy bold totals.
    num_font = Font(size=10, color="1A1A1A")
    money_data_font = Font(size=10, color="1A1A1A")
    total_navy_font = Font(bold=True, size=10, color="1B5E8B")
    num_align = Alignment(horizontal="center", vertical="center", wrap_text=False)
    body_left = Alignment(horizontal="left", vertical="top", wrap_text=True)

    row_idx = 5
    for row in user_rows:
        _put(ws, row_idx, 2, row["name"], font=_NAME_FONT, fill=_ROW_FILL, align=body_left)
        for i, f in enumerate(fields):
            v, kind = row["cells"].get(f.get("key", ""), (_DASH, "muted"))
            if kind == "money":
                _put(ws, row_idx, 3 + i, v, font=money_data_font, fill=_ROW_FILL, align=num_align)
            elif kind == "num":
                _put(ws, row_idx, 3 + i, v, font=num_font, fill=_ROW_FILL, align=num_align)
            elif kind == "body":
                _put(ws, row_idx, 3 + i, v, font=_BODY_FONT, fill=_ROW_FILL, align=body_left)
            else:
                _put(ws, row_idx, 3 + i, v, font=_MUTED_FONT, fill=_ROW_FILL, align=_CENTER)
        row_idx += 1

    # Total row — navy bold text on light-blue fill for every cell.
    if user_rows:
        _put(ws, row_idx, 2, "Total",
             font=total_navy_font, fill=_TOTAL_FILL, align=body_left)
        for i, f in enumerate(fields):
            cells = [r["cells"].get(f.get("key", ""), (_DASH, "muted")) for r in user_rows]
            kinds = {k for _, k in cells}
            if "money" in kinds:
                total = sum(
                    n for v, k in cells if k == "money"
                    for n in _extract_numbers_text(str(v).replace("₹", "").replace(",", ""))
                )
                _put(ws, row_idx, 3 + i, _format_revenue(total) if total else "₹0",
                     font=total_navy_font, fill=_TOTAL_FILL, align=num_align)
            elif "num" in kinds and "body" not in kinds:
                total = sum(v for v, k in cells if k == "num" and isinstance(v, (int, float)))
                disp = int(total) if float(total).is_integer() else round(total, 2)
                _put(ws, row_idx, 3 + i, disp,
                     font=total_navy_font, fill=_TOTAL_FILL, align=num_align)
            elif "body" in kinds:
                all_vals = set()
                for v, k in cells:
                    if k == "body" and isinstance(v, str):
                        for piece in v.split(","):
                            p = piece.strip()
                            if p:
                                all_vals.add(p.lower())
                _put(ws, row_idx, 3 + i, len(all_vals) if all_vals else _DASH,
                     font=total_navy_font if all_vals else _MUTED_FONT,
                     fill=_TOTAL_FILL, align=_CENTER)
            else:
                _put(ws, row_idx, 3 + i, _DASH,
                     font=_MUTED_FONT, fill=_TOTAL_FILL, align=_CENTER)

    summary_last_row = row_idx if user_rows else row_idx - 1
    _apply_table_borders(
        ws, header_row=4,
        last_row=summary_last_row,
        first_col=2, last_col=last_col,
    )
    _append_daily_reports_section(ws, group, start_row=summary_last_row)
    ws.freeze_panes = "C5"
    _set_print_landscape_fit(ws)


def _build_sales_service_detail_sheet(ws, group, *, range_label: str):
    """Sales Service department — 5 aggregated columns + subtotal.

    Counts are anchored to the **actual daily reports** the way HR reads
    them: each numeric column equals the number of distinct days the
    employee logged real content for that metric (i.e. one report submitted
    on date X that mentions a complaint = 1 complaint).  Names / sites /
    projects / clients are extracted **only from that field's text**, so
    nothing is invented from unrelated columns.  Field detection is
    label-based (case-insensitive), making it robust to whatever the
    actual DB field keys are.
    """
    last_col = 7  # A margin + B Employee + 4 data cols + Subtotal

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 22
    widths = [48, 48, 48, 18, 14]
    for i, w in enumerate(widths, start=3):
        ws.column_dimensions[get_column_letter(i)].width = w

    _merge_title(ws, 1, last_col, "Sales Service — summary table", _TITLE_FONT)
    ws.row_dimensions[1].height = 26
    n_users = len(group["users"])
    n_reports = sum(len(rs) for rs in group["users"].values())
    subtitle = (
        f"{range_label}  |  "
        f"{n_users} {'employee' if n_users == 1 else 'employees'}  ·  "
        f"{n_reports} reports submitted"
    )
    _merge_title(ws, 2, last_col, subtitle, _SUBTITLE_FONT)

    headers = [
        "Employee",
        "Complaints (with sites)",
        "Kusum Docs Submitted (with project)",
        "Clients for Loan (with name)",
        "Calls for Kusum Docs",
        "Subtotal",
    ]
    for i, h in enumerate(headers, start=2):
        _put(ws, 4, i, h, font=_HEADER_FONT, fill=_HEADER_FILL, align=_CENTER)
    ws.row_dimensions[4].height = 32

    num_font = Font(size=10, color="1A1A1A")
    right_align = Alignment(horizontal="center", vertical="center", wrap_text=False)
    body_left = Alignment(horizontal="left", vertical="top", wrap_text=True)

    _NULL_TOKENS = {"", "na", "n/a", "n.a.", "n.a", "-", "—", "nil", "none", "on leave"}
    def _has_content(v):
        s = ("" if v is None else str(v)).strip()
        return s.lower() not in _NULL_TOKENS

    # ---- Auto-detect which report_field holds which metric ----
    # We inspect each field's label + key (case-insensitive).  A field
    # contributes to a metric only if its OWN label / key clearly matches
    # that metric — no "spillover" from one column to another.  This is the
    # key to keeping the counts accurate.
    dept_fields = list(group["dept"].report_fields) if group["dept"] and group["dept"].report_fields else []

    def _match(field, *patterns):
        text = ((field.get("label") or "") + " " + (field.get("key") or "")).lower()
        return all(re.search(p, text) for p in patterns)

    complaint_keys: list[str] = []
    kusum_doc_keys: list[str] = []
    loan_keys: list[str] = []
    kusum_call_keys: list[str] = []
    for f in dept_fields:
        # Calls for Kusum Docs: requires BOTH 'kusum' and 'call' tokens.
        if _match(f, r"kusum", r"call"):
            kusum_call_keys.append(f["key"])
            continue
        # Kusum Docs Submitted: 'kusum' or 'doc submit' or 'document submit'.
        if (
            _match(f, r"kusum") and _match(f, r"doc|submit")
            or _match(f, r"document", r"submit")
            or _match(f, r"doc", r"submit")
        ):
            kusum_doc_keys.append(f["key"])
            continue
        # Loan / finance.
        if _match(f, r"\bloan\b") or _match(f, r"\bfinance\b") or _match(f, r"\bemi\b") or _match(f, r"\bfunding\b"):
            loan_keys.append(f["key"])
            continue
        # Complaints / issues / tickets.
        if _match(f, r"complain|complaint|issue|ticket"):
            complaint_keys.append(f["key"])

    # Generic phrase scanner — used ONLY for the Calls counter (a pure
    # number column).  Names / sites are NOT extracted via this.
    HEADER_NUM_RE = re.compile(r"([^\n.;]{1,120}?)\s*[–\-:]\s*(\d+)\b", re.I)

    # Site/city names — used only to label complaints with where they came
    # from.  The total complaint COUNT is `days_with_content(complaint_keys)`
    # so even if we mis-identify a site, the count stays accurate.
    SITE_NAMES = [
        "Noida", "Delhi", "Gurugram", "Gurgaon", "Ghaziabad", "Faridabad",
        "Mumbai", "Pune", "Nagpur", "Nashik", "Mandsaur", "Indore",
        "Lucknow", "Agra", "Kanpur", "Varanasi", "Moradabad", "Bulandshahr",
        "Meerut", "Mathura", "Aligarh", "Bareilly", "Jaipur", "Jodhpur",
        "Udaipur", "Kota", "Ajmer", "Bikaner", "Ahmedabad", "Surat",
        "Vadodara", "Rajkot", "Gandhinagar", "Nadiad", "Chennai",
        "Coimbatore", "Bengaluru", "Bangalore", "Hyderabad", "Kolkata",
        "Chandigarh", "Ludhiana", "Amritsar", "Patiala", "Mohali",
        "Kaithal", "Karnal", "Hisar", "Rohtak", "Panipat", "Sonipat",
        "Bhopal", "Gwalior", "Jabalpur", "Patna", "Ranchi", "Goa",
        "Mavikala", "Sihari", "Simbhalki", "Lilason", "Kapaseda",
    ]
    CAPACITY_SITE_RE = re.compile(
        r"\d+(?:\.\d+)?\s*(?:MWp|MW|kWp|kW)\b[\s\-]*([A-Z][a-zA-Z]+(?:\s+[A-Z][a-zA-Z]+)?)",
    )
    NAMEY_RE = re.compile(r"\b([A-Z][a-z]{2,}(?:\s+[A-Z][a-z]{2,}){0,2})\b")
    NAME_BLACKLIST = {
        "Note", "Status", "Date", "Loan", "Kusum", "Doc", "Docs", "Submit",
        "Submitted", "Bank", "Client", "Clients", "Customer", "Customers",
        "Project", "Projects", "Site", "Sites", "Solar", "Inverter", "Module",
        "Total", "Subtotal", "Call", "Calls", "Called", "Today", "Tomorrow",
        "Yesterday", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun",
        "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday",
        "Sunday", "Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug",
        "Sep", "Oct", "Nov", "Dec", "Done", "Pending", "Completed", "Yes",
        "No", "Mail", "Email", "Phone", "Address",
    }

    def _scan_sites(text):
        out: dict[str, int] = {}
        if not text:
            return out
        for name in SITE_NAMES:
            c = len(re.findall(rf"\b{re.escape(name)}\b", text, re.I))
            if c > 0:
                out[name] = out.get(name, 0) + c
        for m in CAPACITY_SITE_RE.finditer(text):
            cap = m.group(1).strip()
            if any(cap.lower() in n.lower() or n.lower() in cap.lower() for n in out):
                continue
            out[cap] = out.get(cap, 0) + 1
        return out

    def _scan_names(text):
        """Title-case multi-word names from `text`, blacklist-filtered."""
        out: dict[str, int] = {}
        if not text:
            return out
        for m in NAMEY_RE.finditer(text):
            name = m.group(1).strip()
            first = name.split()[0]
            if first in NAME_BLACKLIST:
                continue
            # Require at least two words to qualify as a person / project name.
            if " " not in name:
                continue
            out[name] = out.get(name, 0) + 1
        return out

    def _fmt_with_count(label_counts, day_count, top=6):
        """Render '<name>: <count>, …' followed by ' (Total: N)' where N is
        the number of days the underlying field had real content — that's
        the authoritative count HR will check against the daily reports."""
        if day_count <= 0 and not label_counts:
            return _DASH
        if label_counts:
            items = sorted(label_counts.items(), key=lambda x: -x[1])[:top]
            names = ", ".join(f"{n}: {c}" for n, c in items)
            return f"{names} (Total: {day_count})"
        return f"Total: {day_count}"

    def _sum_phrase_counts(text, keyword_pattern):
        if not text:
            return 0
        total = 0
        for m in HEADER_NUM_RE.finditer(text):
            header = m.group(1).strip()
            if keyword_pattern.search(header):
                try:
                    total += int(m.group(2))
                except ValueError:
                    pass
        return total

    user_rows = []
    for _, user_reports in group["users"].items():
        sorted_reports = sorted(user_reports, key=lambda r: r.date or date_type.min)
        u = sorted_reports[0].user
        full_name = ""
        if u:
            full_name = (
                f"{(u.first_name or '').strip()} {(u.last_name or '').strip()}".strip()
                or u.username
            )

        # Count days where the specific field has real content.  This is the
        # number HR can verify by looking at the daily reports table below.
        def days_with(keys):
            if not keys:
                return 0
            c = 0
            for r in sorted_reports:
                data = r.data or {}
                if any(_has_content(data.get(k, "")) for k in keys):
                    c += 1
            return c

        def pooled(keys):
            parts = []
            for r in sorted_reports:
                data = r.data or {}
                for k in keys:
                    v = data.get(k)
                    if v is not None and _has_content(v):
                        parts.append(str(v))
            return "\n".join(parts)

        complaint_days = days_with(complaint_keys)
        kusum_days = days_with(kusum_doc_keys)
        loan_days = days_with(loan_keys)
        kusum_call_days = days_with(kusum_call_keys)

        # Extract names ONLY from the relevant field's pooled text.
        complaint_sites = _scan_sites(pooled(complaint_keys))
        kusum_projects = _scan_names(pooled(kusum_doc_keys))
        loan_clients = _scan_names(pooled(loan_keys))

        # For the Calls column, prefer an explicit "Calls – N" phrase total
        # when one is present in that field; otherwise fall back to days.
        kusum_call_phrase_total = _sum_phrase_counts(
            pooled(kusum_call_keys),
            re.compile(r"\bcall", re.I),
        )
        kusum_calls = max(kusum_call_phrase_total, kusum_call_days)

        row_subtotal = (
            complaint_days + kusum_days + loan_days + kusum_calls
        )

        user_rows.append({
            "name": full_name,
            "complaints_text": _fmt_with_count(complaint_sites, complaint_days),
            "kusum_text": _fmt_with_count(kusum_projects, kusum_days),
            "loans_text": _fmt_with_count(loan_clients, loan_days),
            "kusum_calls": kusum_calls,
            "subtotal": row_subtotal,
        })
    user_rows.sort(key=lambda r: r["name"].lower())

    row_idx = 5
    for row in user_rows:
        _put(ws, row_idx, 2, row["name"], font=_NAME_FONT, fill=_ROW_FILL, align=body_left)
        for col, key in [
            (3, "complaints_text"), (4, "kusum_text"), (5, "loans_text"),
        ]:
            txt = row[key]
            _put(ws, row_idx, col, txt,
                 font=_BODY_FONT if txt != _DASH else _MUTED_FONT,
                 fill=_ROW_FILL,
                 align=body_left if txt != _DASH else right_align)
        _put(ws, row_idx, 6, row["kusum_calls"],
             font=num_font, fill=_ROW_FILL, align=right_align)
        _put(ws, row_idx, 7, row["subtotal"],
             font=Font(bold=True, size=10, color="1A1A1A"),
             fill=_ROW_FILL, align=right_align)
        row_idx += 1

    if user_rows:
        _put(ws, row_idx, 2, "Total", font=_TOTAL_FONT, fill=_TOTAL_FILL, align=body_left)
        _put(ws, row_idx, 3, "", fill=_TOTAL_FILL)
        _put(ws, row_idx, 4, "", fill=_TOTAL_FILL)
        _put(ws, row_idx, 5, "", fill=_TOTAL_FILL)
        _put(ws, row_idx, 6, sum(r["kusum_calls"] for r in user_rows),
             font=_TOTAL_FONT, fill=_TOTAL_FILL, align=right_align)
        _put(ws, row_idx, 7, sum(r["subtotal"] for r in user_rows),
             font=_TOTAL_FONT, fill=_TOTAL_FILL, align=right_align)

    summary_last_row = row_idx if user_rows else row_idx - 1
    _apply_table_borders(
        ws, header_row=4,
        last_row=summary_last_row,
        first_col=2, last_col=last_col,
    )
    _append_daily_reports_section(ws, group, start_row=summary_last_row)
    ws.freeze_panes = "C5"
    _set_print_landscape_fit(ws)


def _build_hr_detail_sheet(ws, group, *, range_label: str):
    """HR department — 6 aggregated columns matching the daily-report form:
        • Recruitment / Screening          (CV review + interview totals)
        • Induction / Onboarding           (new joiners onboarded)
        • Exit Process                     (employees exited, with names)
        • Attendance, Payroll & Compliance (days with content)
        • Happay & Conveyance              (days with content)
        • Other Work                       (condensed text)
      + Subtotal column.

    Each numeric column is the MAX of:
      (a) days the matching field had real content, AND
      (b) phrase-scanned "X – N" totals inside that field's text.
    So if HR explicitly writes "CVs reviewed – 8", the 8 wins; otherwise the
    day count anchors the value to what's visible in the daily reports.
    """
    last_col = 10  # A margin + B Employee + 7 form cols + Subtotal

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 22
    # Recruit, Induct, Exit, Att/Pay, Happay, MIS, Other, Subtotal
    widths = [40, 32, 40, 22, 22, 22, 50, 12]
    for i, w in enumerate(widths, start=3):
        ws.column_dimensions[get_column_letter(i)].width = w

    _merge_title(ws, 1, last_col, "HR — summary table", _TITLE_FONT)
    ws.row_dimensions[1].height = 26
    n_users = len(group["users"])
    n_reports = sum(len(rs) for rs in group["users"].values())
    subtitle = (
        f"{range_label}  |  "
        f"{n_users} {'employee' if n_users == 1 else 'employees'}  ·  "
        f"{n_reports} reports submitted"
    )
    _merge_title(ws, 2, last_col, subtitle, _SUBTITLE_FONT)

    headers = [
        "Employee",
        "Recruitment / Screening",
        "Induction / Onboarding",
        "Exit Process (with names)",
        "Attendance, Payroll & Compliance",
        "Happay & Conveyance",
        "MIS & Reports",
        "Other Work",
        "Subtotal",
    ]
    for i, h in enumerate(headers, start=2):
        _put(ws, 4, i, h, font=_HEADER_FONT, fill=_HEADER_FILL, align=_CENTER)
    ws.row_dimensions[4].height = 34

    num_font = Font(size=10, color="1A1A1A")
    right_align = Alignment(horizontal="center", vertical="center", wrap_text=False)
    body_left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    subtotal_bold = Font(bold=True, size=10, color="1A1A1A")

    _NULL_TOKENS = {"", "na", "n/a", "n.a.", "n.a", "-", "—", "nil", "none", "on leave"}
    def _has_content(v):
        s = ("" if v is None else str(v)).strip()
        return s.lower() not in _NULL_TOKENS

    # ---- Auto-detect each metric's actual field key ----
    # We pool keys from BOTH the current dept schema (`report_fields`) AND
    # any keys present in actual saved report data — the latter catches the
    # case where the HR form was renamed (e.g. `recruitment` → `recruitmentScreening`)
    # but older reports still carry the previous key.
    dept_fields = list(group["dept"].report_fields) if group["dept"] and group["dept"].report_fields else []
    key_to_label: dict[str, str] = {}
    for f in dept_fields:
        k = f.get("key") or ""
        if k:
            key_to_label[k] = f.get("label") or ""
    # Also harvest keys from actual report data so legacy keys count.
    for _, user_reports in group["users"].items():
        for r in user_reports:
            for k in (r.data or {}).keys():
                if k.startswith("__"):
                    continue
                key_to_label.setdefault(k, "")
    all_known_keys = list(key_to_label.keys())

    def _key_matches(key, pattern):
        text = (key + " " + key_to_label.get(key, "")).lower()
        return bool(re.search(pattern, text))

    recruit_keys: list[str] = []
    induct_keys: list[str] = []
    exit_keys: list[str] = []
    attend_keys: list[str] = []
    happay_keys: list[str] = []
    mis_keys: list[str] = []
    other_keys: list[str] = []
    for k in all_known_keys:
        if _key_matches(k, r"recruit|screen|interview|cv|resume"):
            recruit_keys.append(k)
        elif _key_matches(k, r"induct|onboard|joining|new\s+joiner"):
            induct_keys.append(k)
        elif _key_matches(k, r"exit|resign|relieving|f\s*&\s*f|full\s+and\s+final"):
            exit_keys.append(k)
        elif _key_matches(k, r"happay|conveyance|expense|reimburs"):
            happay_keys.append(k)
        elif _key_matches(k, r"\bmis\b|reporting"):
            # MIS & Reports — must be checked BEFORE attend (which catches
            # "leave") because a MIS field may also contain leave-related text.
            mis_keys.append(k)
        elif _key_matches(k, r"attend|payroll|complian|salary|leave"):
            attend_keys.append(k)
        elif _key_matches(k, r"other"):
            other_keys.append(k)

    # Generic "<header containing keyword>: N" phrase scanner.
    HEADER_NUM_RE = re.compile(r"([^\n.;]{1,120}?)\s*[–\-:]\s*(\d+)\b", re.I)

    # Keyword patterns used to filter phrase headers per metric.
    # Recruitment / Screening is split into TWO sub-counts shown inline:
    #   Recruits   — actual hires / offers / joiners under the recruit field
    #   Screening  — CV / resume / interview / shortlist activity
    RECRUITS_KW  = re.compile(r"\b(?:recruit|hire(?:d)?|offer(?:ed)?|select(?:ed)?|joined|joining)", re.I)
    SCREENING_KW = re.compile(r"\b(?:cv|resume|screen|interview|shortlist|applicant|candidate)", re.I)
    RECRUIT_KW   = re.compile(r"\b(?:cv|resume|candidate|interview|screen|shortlist|applicant|recruit|hire|offer|select)", re.I)
    INDUCT_KW = re.compile(r"\b(?:induct|onboard|joiner|joined|orientation)", re.I)
    EXIT_KW = re.compile(r"\b(?:exit|resign|reliev|f\s*&\s*f|full\s+and\s+final|separat)", re.I)
    ATTEND_KW = re.compile(r"\b(?:attend|payroll|salary|leave|complian)", re.I)
    HAPPAY_KW = re.compile(r"\b(?:happay|conveyance|expense|reimburs)", re.I)

    # Person-name extractor — used for the Exit Process column.  Two or more
    # title-case words, filtered by a blacklist of common non-names.
    NAMEY_RE = re.compile(r"\b([A-Z][a-z]{2,}(?:\s+[A-Z][a-z]{2,}){0,2})\b")
    NAME_BLACKLIST = {
        "Note", "Status", "Date", "Done", "Pending", "Completed", "Yes", "No",
        "Today", "Tomorrow", "Yesterday", "Mon", "Tue", "Wed", "Thu", "Fri",
        "Sat", "Sun", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday",
        "Saturday", "Sunday", "Jan", "Feb", "Mar", "Apr", "May", "Jun",
        "Jul", "Aug", "Sep", "Oct", "Nov", "Dec",
        "Exit", "Process", "Resign", "Resigned", "Relieving", "Leave",
        "Total", "Subtotal", "Other", "Work", "Works", "Reason", "Notice",
        "Period", "Final", "Settlement",
    }

    def _sum_phrase_counts(text, keyword_pattern):
        if not text:
            return 0
        total = 0
        for m in HEADER_NUM_RE.finditer(text):
            header = m.group(1).strip()
            if keyword_pattern.search(header):
                try:
                    total += int(m.group(2))
                except ValueError:
                    pass
        return total

    def _scan_names(text):
        out: dict[str, int] = {}
        if not text:
            return out
        for m in NAMEY_RE.finditer(text):
            name = m.group(1).strip()
            first = name.split()[0]
            if first in NAME_BLACKLIST:
                continue
            if " " not in name:
                continue
            out[name] = out.get(name, 0) + 1
        return out

    user_rows = []
    for _, user_reports in group["users"].items():
        sorted_reports = sorted(user_reports, key=lambda r: r.date or date_type.min)
        u = sorted_reports[0].user
        full_name = ""
        if u:
            full_name = (
                f"{(u.first_name or '').strip()} {(u.last_name or '').strip()}".strip()
                or u.username
            )

        def days_with(keys):
            if not keys:
                return 0
            c = 0
            for r in sorted_reports:
                data = r.data or {}
                if any(_has_content(data.get(k, "")) for k in keys):
                    c += 1
            return c

        def pooled(keys):
            parts = []
            for r in sorted_reports:
                data = r.data or {}
                for k in keys:
                    v = data.get(k)
                    if v is not None and _has_content(v):
                        parts.append(str(v))
            return "\n".join(parts)

        recruit_days = days_with(recruit_keys)
        induct_days  = days_with(induct_keys)
        exit_days    = days_with(exit_keys)
        attend_days  = days_with(attend_keys)
        happay_days  = days_with(happay_keys)
        mis_days     = days_with(mis_keys)

        recruit_text = pooled(recruit_keys)
        induct_text  = pooled(induct_keys)
        exit_text    = pooled(exit_keys)
        attend_text  = pooled(attend_keys)
        happay_text  = pooled(happay_keys)
        mis_text     = pooled(mis_keys)
        other_text   = pooled(other_keys)

        # Recruitment / Screening — three sub-counts that match what Smita
        # actually writes:
        #   Offers / Recruits  = "Offer Released X" / "Offer Releasd X" /
        #                        "Offer release X" mentions (count of NAMES)
        #   Screenings         = sum of "Screening- N" / "Screeening- N" /
        #                        "Screening N" numbers
        #   Interviews         = "Interview" mentions (count) excluding lines
        #                        that are pure screening
        offers_count = len(re.findall(
            r"\boffer\s+release[ds]?\b", recruit_text, re.I,
        ))
        screening_nums = re.findall(
            r"\bscree+ning[s]?\s*[-–:]?\s*(\d{1,3})\b", recruit_text, re.I,
        )
        screenings_count = sum(int(n) for n in screening_nums) if screening_nums else 0
        # If no explicit number but the field has screening mentions, count
        # the mentions themselves so we don't drop to zero.
        if not screenings_count:
            screenings_count = len(re.findall(
                r"\bscree+ning[s]?\b", recruit_text, re.I,
            ))
        interviews_count = len(re.findall(
            r"\binterview[s]?\b", recruit_text, re.I,
        ))
        if (offers_count or screenings_count or interviews_count):
            recruit_disp = (
                f"Offers: {offers_count}\n"
                f"Screenings: {screenings_count}\n"
                f"Interviews: {interviews_count}"
            )
        else:
            recruit_disp = _DASH

        # Induction / Onboarding — ONE total count (no "Days:" — HR asked
        # for just the total).  Counts induction/onboarding events: the
        # bigger of (numbered-list items across all days, day count).
        numbered_items = len(re.findall(r"(?m)^\s*\d+\s*[.)]\s*", induct_text))
        induct_count = max(numbered_items, induct_days)
        induct_disp = (
            f"Inductions: {induct_count}" if induct_count else _DASH
        )

        # Exit Process — extract names ONLY from explicit exit-marker
        # phrases (Experience Letter / Final Payment / Resigned / Reliev),
        # not from random title-cased words.  Avoids the previous bug where
        # "Nandlal Payment Calculation" (activity) showed up as a name.
        EXIT_NAME_RE = re.compile(
            r"(?:experience\s+letter[\s\-:–]*|final\s+payment[\s\-\(:]*for[\s:]*|"
            r"final\s+payment[\s\-:–]*|resign(?:ed|ation)\s+(?:of\s+)?|"
            r"reliev(?:ed|ing)\s+(?:of\s+)?|exit\s+(?:of\s+)?)"
            r"([A-Z][a-zA-Z]+(?:\s+(?:&|and)\s+[A-Z][a-zA-Z]+)?(?:\s+[A-Z][a-zA-Z]+)?)",
            re.I,
        )
        exit_named: list[str] = []
        for m in EXIT_NAME_RE.finditer(exit_text):
            chunk = m.group(1).strip()
            # Split "X & Y" / "X and Y" into individual names.
            for piece in re.split(r"\s+(?:&|and)\s+", chunk, flags=re.I):
                piece = piece.strip().rstrip(",.;")
                if piece and piece not in exit_named:
                    exit_named.append(piece)
        exit_count = max(len(exit_named), exit_days)
        if exit_named:
            top_names = exit_named[:6]
            exit_disp = ", ".join(top_names) + f" (Total: {exit_count})"
        elif exit_count:
            exit_disp = f"Total: {exit_count}"
        else:
            exit_disp = _DASH

        # Attendance / Payroll / Compliance — day count only.
        attend_disp = f"Days: {attend_days}" if attend_days else _DASH

        # Happay & Conveyance — day count only.
        happay_disp = f"Days: {happay_days}" if happay_days else _DASH

        # MIS & Reports — day count only.
        mis_disp = f"Days: {mis_days}" if mis_days else _DASH

        # Other Work — condensed AI-style summary across all dates.  Wider
        # budget so the cell shows representative phrases without being
        # truncated mid-thought.
        other_disp = _condense_summary(other_text, max_chars=220) if other_text else _DASH

        # Subtotal = total DAYS the employee filed any HR activity in this
        # range — consistent unit across all 6 metric columns.
        row_subtotal = (
            recruit_days + induct_days + exit_days
            + attend_days + happay_days + mis_days
        )

        user_rows.append({
            "name": full_name,
            "recruit_disp": recruit_disp,
            "induct_disp": induct_disp,
            "exit_disp": exit_disp,
            "attend_disp": attend_disp,
            "happay_disp": happay_disp,
            "mis_disp": mis_disp,
            "other_disp": other_disp,
            "subtotal": row_subtotal,
        })
    user_rows.sort(key=lambda r: r["name"].lower())

    row_idx = 5
    for row in user_rows:
        _put(ws, row_idx, 2, row["name"],
             font=_NAME_FONT, fill=_ROW_FILL, align=body_left)
        for col, key in [
            (3, "recruit_disp"), (4, "induct_disp"), (5, "exit_disp"),
            (6, "attend_disp"), (7, "happay_disp"), (8, "mis_disp"),
            (9, "other_disp"),
        ]:
            txt = row[key]
            _put(ws, row_idx, col, txt,
                 font=_BODY_FONT if txt != _DASH else _MUTED_FONT,
                 fill=_ROW_FILL,
                 align=body_left if txt != _DASH else right_align)
        _put(ws, row_idx, 10, row["subtotal"],
             font=subtotal_bold, fill=_ROW_FILL, align=right_align)
        row_idx += 1

    if user_rows:
        _put(ws, row_idx, 2, "Total",
             font=_TOTAL_FONT, fill=_TOTAL_FILL, align=body_left)
        # Text cols 3-9 stay blank in the total row.
        for c in range(3, 10):
            _put(ws, row_idx, c, "", fill=_TOTAL_FILL)
        _put(ws, row_idx, 10,
             sum(r["subtotal"] for r in user_rows),
             font=_TOTAL_FONT, fill=_TOTAL_FILL, align=right_align)

    summary_last_row = row_idx if user_rows else row_idx - 1
    _apply_table_borders(
        ws, header_row=4,
        last_row=summary_last_row,
        first_col=2, last_col=last_col,
    )
    _append_daily_reports_section(ws, group, start_row=summary_last_row)
    ws.freeze_panes = "C5"
    _set_print_landscape_fit(ws)


def _build_service_detail_sheet(ws, group, *, range_label: str):
    """Service department — 3 aggregated columns instead of one per
    report_field.  Combines the underlying daily fields so HR sees totals
    for site visits, inverter complaints, and part replacements per
    employee.
    """
    last_col = 6  # A margin + B Employee + C Sites Visited + D Total Site Visit + E Inv Complaint + F Parts Repl

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 55   # Sites Visited (names + counts)
    ws.column_dimensions["D"].width = 16   # Total Site Visit
    ws.column_dimensions["E"].width = 26
    ws.column_dimensions["F"].width = 26

    _merge_title(ws, 1, last_col, "Service — summary table", _TITLE_FONT)
    ws.row_dimensions[1].height = 26
    n_users = len(group["users"])
    n_reports = sum(len(rs) for rs in group["users"].values())
    subtitle = (
        f"{range_label}  |  "
        f"{n_users} {'employee' if n_users == 1 else 'employees'}  ·  "
        f"{n_reports} reports submitted"
    )
    _merge_title(ws, 2, last_col, subtitle, _SUBTITLE_FONT)

    headers = [
        "Employee", "Sites Visited (with count)", "Total Site Visit",
        "Total Inverter Complaint", "Inverter Parts Replacement",
    ]
    for i, h in enumerate(headers, start=2):
        _put(ws, 4, i, h, font=_HEADER_FONT, fill=_HEADER_FILL, align=_CENTER)
    ws.row_dimensions[4].height = 28

    num_font = Font(size=10, color="1A1A1A")
    right_align = Alignment(horizontal="center", vertical="center", wrap_text=False)
    body_left = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Aggregation rules:
    # Each cell shows the COUNT OF DAYS the employee logged meaningful
    # content in that area (a day with "N/A" / "Na" / "-" does NOT count).
    #
    #   Sites Visited            = site names + visit counts parsed out of the
    #                              site-visit text (regardless of key spelling)
    #   Total Site Visit         = days where Solar Install OR Inverter
    #                              Complaint site visit field has real content
    #   Total Inverter Complaint = days where Complaint site visit OR
    #                              Complaint tele/video field has real content
    #   Inverter Parts Replace   = days where Inverter Part Replacement has
    #                              real content
    SITE_VISIT_KEYS = ("solarInstallationSiteVisit", "inverterComplaintSiteVisit")
    INV_COMPLAINT_KEYS = ("inverterComplaintSiteVisit", "inverterComplaintTeleVideo")
    PARTS_REPL_KEYS = ("inverterPartReplacement",)

    _NULL_TOKENS = {"", "na", "n/a", "n.a.", "n.a", "-", "—", "nil", "none", "on leave"}

    def _has_content(v):
        s = ("" if v is None else str(v)).strip()
        return s.lower() not in _NULL_TOKENS

    # Auto-discover site-visit field keys from the dept schema so we're not
    # dependent on the exact spelling of `solarInstallationSiteVisit` etc.
    dept_fields = list(group["dept"].report_fields) if group["dept"] and group["dept"].report_fields else []
    auto_site_keys: list[str] = []
    for f in dept_fields:
        label_l = (f.get("label") or "").lower()
        key_l = (f.get("key") or "").lower()
        if "visit" in label_l or "visit" in key_l or "site" in label_l or "site" in key_l:
            auto_site_keys.append(f["key"])
    if auto_site_keys:
        SITE_VISIT_KEYS = tuple(dict.fromkeys(list(SITE_VISIT_KEYS) + auto_site_keys))

    # Known Indian site / city names — case-insensitive whole-word match.
    SITE_NAMES = [
        "Noida", "Delhi", "Gurugram", "Gurgaon", "Ghaziabad", "Faridabad",
        "Mumbai", "Pune", "Nagpur", "Nashik", "Mandsaur", "Indore",
        "Lucknow", "Agra", "Kanpur", "Varanasi", "Allahabad", "Moradabad",
        "Bulandshahr", "Meerut", "Mathura", "Aligarh", "Bareilly",
        "Jaipur", "Jodhpur", "Udaipur", "Kota", "Ajmer", "Bikaner",
        "Ahmedabad", "Surat", "Vadodara", "Rajkot", "Gandhinagar", "Nadiad",
        "Chennai", "Coimbatore", "Madurai", "Tirupur",
        "Bengaluru", "Bangalore", "Mysuru", "Mysore", "Hubli",
        "Hyderabad", "Vijayawada", "Vizag", "Visakhapatnam",
        "Kolkata", "Howrah", "Durgapur",
        "Chandigarh", "Ludhiana", "Amritsar", "Patiala", "Mohali",
        "Kaithal", "Karnal", "Hisar", "Rohtak", "Panipat", "Sonipat",
        "Bhopal", "Gwalior", "Jabalpur", "Ujjain", "Sagar",
        "Patna", "Ranchi", "Jamshedpur",
        "Goa", "Panaji", "Mangaluru", "Mangalore",
        "Sihari", "Simbhalki", "Lilasons", "Lilason", "Mavikala",
        "Kamala", "Titroda", "Padams", "Tusara", "Kapaseda", "Priyanka",
    ]
    CAPACITY_SITE_RE = re.compile(
        r"\d+(?:\.\d+)?\s*(?:MWp|MW|kWp|kW)\b[\s\-]*([A-Z][a-zA-Z]+(?:\s+[A-Z][a-zA-Z]+)?)",
    )

    user_rows = []
    for _, user_reports in group["users"].items():
        sorted_reports = sorted(user_reports, key=lambda r: r.date or date_type.min)
        u = sorted_reports[0].user
        full_name = ""
        if u:
            full_name = (
                f"{(u.first_name or '').strip()} {(u.last_name or '').strip()}".strip()
                or u.username
            )

        def count_days(keys):
            """Return number of reports where ANY of the given fields has
            meaningful content."""
            total = 0
            for r in sorted_reports:
                data = r.data or {}
                if any(_has_content(data.get(k, "")) for k in keys):
                    total += 1
            return total

        # Pool every site-visit field's text for this employee.
        site_text_parts = []
        for r in sorted_reports:
            data = r.data or {}
            for k in SITE_VISIT_KEYS:
                v = data.get(k)
                if v and _has_content(v):
                    site_text_parts.append(str(v))
        site_blob = " ".join(site_text_parts)

        site_counts: dict[str, int] = {}
        for name in SITE_NAMES:
            c = len(re.findall(rf"\b{re.escape(name)}\b", site_blob, re.I))
            if c > 0:
                site_counts[name] = site_counts.get(name, 0) + c
        for m in CAPACITY_SITE_RE.finditer(site_blob):
            cap = m.group(1).strip()
            if any(cap.lower() in n.lower() or n.lower() in cap.lower() for n in site_counts):
                continue
            site_counts[cap] = site_counts.get(cap, 0) + 1

        sorted_sites = sorted(site_counts.items(), key=lambda x: -x[1])[:8]
        sites_text = ", ".join(f"{n}: {c}" for n, c in sorted_sites) or _DASH

        user_rows.append({
            "name": full_name,
            "sites_text": sites_text,
            "site_visit": count_days(SITE_VISIT_KEYS),
            "inv_complaint": count_days(INV_COMPLAINT_KEYS),
            "parts_repl": count_days(PARTS_REPL_KEYS),
        })
    user_rows.sort(key=lambda r: r["name"].lower())

    def _disp(n):
        return int(n) if float(n).is_integer() else round(n, 2)

    row_idx = 5
    for row in user_rows:
        _put(ws, row_idx, 2, row["name"], font=_NAME_FONT, fill=_ROW_FILL, align=body_left)
        _put(ws, row_idx, 3, row["sites_text"],
             font=_BODY_FONT if row["sites_text"] != _DASH else _MUTED_FONT,
             fill=_ROW_FILL,
             align=body_left if row["sites_text"] != _DASH else right_align)
        for i, key in enumerate(("site_visit", "inv_complaint", "parts_repl")):
            _put(ws, row_idx, 4 + i, _disp(row[key]),
                 font=num_font, fill=_ROW_FILL, align=right_align)
        row_idx += 1

    if user_rows:
        _put(ws, row_idx, 2, "Total", font=_TOTAL_FONT, fill=_TOTAL_FILL, align=body_left)
        _put(ws, row_idx, 3, "", font=_TOTAL_FONT, fill=_TOTAL_FILL, align=body_left)
        for i, key in enumerate(("site_visit", "inv_complaint", "parts_repl")):
            total = sum(r[key] for r in user_rows)
            _put(ws, row_idx, 4 + i, _disp(total),
                 font=_TOTAL_FONT, fill=_TOTAL_FILL, align=right_align)

    summary_last_row = row_idx if user_rows else row_idx - 1
    _apply_table_borders(
        ws, header_row=4,
        last_row=summary_last_row,
        first_col=2, last_col=last_col,
    )
    _append_daily_reports_section(ws, group, start_row=summary_last_row)
    ws.freeze_panes = "C5"
    _set_print_landscape_fit(ws)


def _build_procurement_detail_sheet(ws, group, *, range_label: str):
    """Procurement department — 5 aggregated counter columns.

    Each metric is read two ways and the larger value wins:
      1. Explicit "X Sent – N" / "PO Created – 3" markers in the text.
      2. A pattern-count fallback (e.g. number of PO/OAPL/… IDs found,
         or number of days with content in that field).
    """
    last_col = 7
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 22
    widths = [14, 16, 22, 16, 16]
    for i, w in enumerate(widths, start=3):
        ws.column_dimensions[get_column_letter(i)].width = w

    _merge_title(ws, 1, last_col, "Procurement — summary table", _TITLE_FONT)
    ws.row_dimensions[1].height = 26
    n_users = len(group["users"])
    n_reports = sum(len(rs) for rs in group["users"].values())
    subtitle = (
        f"{range_label}  |  "
        f"{n_users} {'employee' if n_users == 1 else 'employees'}  ·  "
        f"{n_reports} reports submitted"
    )
    _merge_title(ws, 2, last_col, subtitle, _SUBTITLE_FONT)

    headers = [
        "Employee", "Enquiries Done", "Negotiations",
        "Vendor Meetings / New Vendors", "POs Created", "NOPAs Created",
    ]
    for i, h in enumerate(headers, start=2):
        _put(ws, 4, i, h, font=_HEADER_FONT, fill=_HEADER_FILL, align=_CENTER)
    ws.row_dimensions[4].height = 32

    num_font = Font(size=10, color="1A1A1A")
    right_align = Alignment(horizontal="center", vertical="center", wrap_text=False)
    body_left = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Phrase-based scanner: finds every "<some header text> – N" or
    # "<header text>: N" occurrence and lets a per-metric regex decide
    # whether that header counts.  Header is captured up to 100 chars and
    # cannot cross a newline, period, or semicolon — keeps each match scoped
    # to one phrase.
    HEADER_NUM_RE = re.compile(r"([^\n.;]{1,100}?)\s*[–\-:]\s*(\d+)\b", re.I)

    # Per-metric keyword patterns.  Match against the header text — if the
    # keyword appears anywhere in the phrase, the trailing N is counted.
    ENQUIRY_KW = re.compile(r"\b(?:inquir(?:y|ies)|enquir(?:y|ies)|rfq)\b", re.I)
    NEGOT_KW   = re.compile(r"\b(?:negotiat|comparison|rate\s+finalization)", re.I)
    VENDOR_KW  = re.compile(
        r"\b(?:meeting|vendor(?:\s+(?:coordination|registration|onboarding|"
        r"added|registered|visit))?|new\s+vendor)\b",
        re.I,
    )
    PO_KW   = re.compile(
        r"\b(?:po|pos|purchase\s+order)s?\s+(?:created|generated|made|"
        r"raised|prepared)\b",
        re.I,
    )
    NOPA_KW = re.compile(
        r"\b(?:nopa|nopas|nope)\s+(?:created|generated|processed|made|prepared)\b",
        re.I,
    )

    # Distinct-ID patterns (PO/OAPL/2627/0073, NOPA/OAPL/2526/0721/4 etc.).
    PO_ID_RE   = re.compile(r"\bPO/[A-Z]+/\d+/\d+\b", re.I)
    NOPA_ID_RE = re.compile(r"\bNOPA/[A-Z]+/\d+/\d+(?:/\d+)?\b", re.I)

    _NULL_TOKENS = {"", "na", "n/a", "n.a.", "n.a", "-", "—", "nil", "none", "on leave"}
    def _has_content(v):
        s = ("" if v is None else str(v)).strip()
        return s.lower() not in _NULL_TOKENS

    def _sum_phrase_counts(text, keyword_pattern):
        """Sum every '… <keyword phrase> – N' occurrence in the text."""
        if not text:
            return 0
        total = 0
        for m in HEADER_NUM_RE.finditer(text):
            header = m.group(1).strip()
            if keyword_pattern.search(header):
                try:
                    total += int(m.group(2))
                except ValueError:
                    pass
        return total

    user_rows = []
    for _, user_reports in group["users"].items():
        sorted_reports = sorted(user_reports, key=lambda r: r.date or date_type.min)
        u = sorted_reports[0].user
        full_name = ""
        if u:
            full_name = (
                f"{(u.first_name or '').strip()} {(u.last_name or '').strip()}".strip()
                or u.username
            )

        # Pool EVERY field value across all reports.  Robust to unknown field
        # keys — the scan runs on the whole employee corpus.
        all_text_parts = []
        for r in sorted_reports:
            data = r.data or {}
            for k, v in data.items():
                if k.startswith("__"):  # skip __leave__ markers
                    continue
                if v is None:
                    continue
                all_text_parts.append(str(v))
        big_text = "\n".join(all_text_parts)

        enq_count  = _sum_phrase_counts(big_text, ENQUIRY_KW)
        neg_count  = _sum_phrase_counts(big_text, NEGOT_KW)
        ven_count  = _sum_phrase_counts(big_text, VENDOR_KW)
        po_count   = max(
            _sum_phrase_counts(big_text, PO_KW),
            len(set(PO_ID_RE.findall(big_text))),
        )
        nopa_count = max(
            _sum_phrase_counts(big_text, NOPA_KW),
            len(set(NOPA_ID_RE.findall(big_text))),
        )

        user_rows.append({
            "name": full_name,
            "enquiries": enq_count,
            "negotiations": neg_count,
            "vendor_meetings": ven_count,
            "pos": po_count,
            "nopas": nopa_count,
        })
    user_rows.sort(key=lambda r: r["name"].lower())

    NUMERIC_KEYS = ("enquiries", "negotiations", "vendor_meetings", "pos", "nopas")

    row_idx = 5
    for row in user_rows:
        _put(ws, row_idx, 2, row["name"], font=_NAME_FONT, fill=_ROW_FILL, align=body_left)
        for i, key in enumerate(NUMERIC_KEYS, start=3):
            _put(ws, row_idx, i, row[key], font=num_font, fill=_ROW_FILL, align=right_align)
        row_idx += 1

    if user_rows:
        _put(ws, row_idx, 2, "Total", font=_TOTAL_FONT, fill=_TOTAL_FILL, align=body_left)
        for i, key in enumerate(NUMERIC_KEYS, start=3):
            _put(ws, row_idx, i,
                 sum(r[key] for r in user_rows),
                 font=_TOTAL_FONT, fill=_TOTAL_FILL, align=right_align)

    summary_last_row = row_idx if user_rows else row_idx - 1
    _apply_table_borders(
        ws, header_row=4,
        last_row=summary_last_row,
        first_col=2, last_col=last_col,
    )
    _append_daily_reports_section(ws, group, start_row=summary_last_row)
    ws.freeze_panes = "C5"
    _set_print_landscape_fit(ws)


def _build_design_detail_sheet(ws, group, *, range_label: str):
    """Design department — 3 aggregated columns:
      - Projects (names + counts) extracted from the Design field text
      - Designs Made (count of distinct design entries)
      - Site Visits (count of days with real content in Site Survey & Visit)
    """
    last_col = 6
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 22   # Employee
    ws.column_dimensions["C"].width = 70   # Projects (names + counts)
    ws.column_dimensions["D"].width = 14   # Total Projects
    ws.column_dimensions["E"].width = 16   # Designs Made
    ws.column_dimensions["F"].width = 14   # Site Visits

    _merge_title(ws, 1, last_col, "Design — summary table", _TITLE_FONT)
    ws.row_dimensions[1].height = 26
    n_users = len(group["users"])
    n_reports = sum(len(rs) for rs in group["users"].values())
    subtitle = (
        f"{range_label}  |  "
        f"{n_users} {'employee' if n_users == 1 else 'employees'}  ·  "
        f"{n_reports} reports submitted"
    )
    _merge_title(ws, 2, last_col, subtitle, _SUBTITLE_FONT)

    headers = [
        "Employee", "Projects (with mention count)",
        "Total Projects", "Designs Made", "Site Visits",
    ]
    for i, h in enumerate(headers, start=2):
        _put(ws, 4, i, h, font=_HEADER_FONT, fill=_HEADER_FILL, align=_CENTER)
    ws.row_dimensions[4].height = 28

    num_font = Font(size=10, color="1A1A1A")
    right_align = Alignment(horizontal="center", vertical="center", wrap_text=False)
    body_left = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Known project / site names — extracted by case-insensitive match.
    PROJECT_NAMES = [
        "Muzaffarnagar", "Khaithwadi", "Kaithwadi", "Sihari", "Simbhalki",
        "Lilasons", "Lilason", "Vales Function", "Vales", "Tusara",
        "Mavikala", "Mavikalan", "Kamala", "Titroda", "Padams", "Padom",
        "NTPC", "Ledure", "G-Plast", "Vardhmaan", "Vardhman", "Faze",
        "Panipat", "Rooftop", "Goa", "Sumati", "Holy Family", "Inroof",
        "Codal", "Bigwit", "Suryagrid", "Priyanka", "Kapaseda",
    ]
    CAPACITY_SITE_RE = re.compile(
        r"\d+(?:\.\d+)?\s*(?:MWp|MW|kWp|kW)\b[\s\-]*([A-Z][a-zA-Z]+(?:\s+[A-Z][a-zA-Z]+)?)",
    )

    _NULL_TOKENS = {"", "na", "n/a", "n.a.", "n.a", "-", "—", "nil", "none", "on leave"}
    def _has_content(v):
        s = ("" if v is None else str(v)).strip()
        return s.lower() not in _NULL_TOKENS

    # Auto-discover field keys from the dept's report_fields by scanning
    # labels — robust to differences in actual field-key spelling like
    # `siteSurveyVisit` vs `siteVisit` vs `siteSurveyAndVisit`.
    dept_fields = list(group["dept"].report_fields) if group["dept"] and group["dept"].report_fields else []
    design_field_keys: list[str] = []
    site_field_keys: list[str] = []
    for f in dept_fields:
        label_l = (f.get("label") or "").lower()
        key_l = (f.get("key") or "").lower()
        # Site-visit fields: label / key contains "visit" or "survey".
        if "visit" in label_l or "survey" in label_l or "visit" in key_l or "survey" in key_l:
            site_field_keys.append(f["key"])
        # Design fields: label / key contains "design" (excluding "project management").
        elif "design" in label_l or "design" in key_l or "drawing" in label_l or "layout" in label_l:
            design_field_keys.append(f["key"])

    user_rows = []
    for _, user_reports in group["users"].items():
        sorted_reports = sorted(user_reports, key=lambda r: r.date or date_type.min)
        u = sorted_reports[0].user
        full_name = ""
        if u:
            full_name = (
                f"{(u.first_name or '').strip()} {(u.last_name or '').strip()}".strip()
                or u.username
            )

        design_text_pool = []
        designs_count = 0
        site_visit_days = 0
        for r in sorted_reports:
            data = r.data or {}
            # Pool design-field content + count distinct design entries.
            for k in design_field_keys:
                v = data.get(k) or ""
                if _has_content(v):
                    design_text_pool.append(str(v))
                    fragments = [
                        s.strip() for s in re.split(r"[\n.;]+", str(v))
                        if s.strip() and len(s.strip()) > 4
                    ]
                    designs_count += len(fragments)
            # Site-visit day: any of the site fields has real content.
            if site_field_keys and any(_has_content(data.get(k, "")) for k in site_field_keys):
                site_visit_days += 1

        # Project name extraction: scan ALL field text (not just the design
        # column) so projects mentioned in "Site Visit" / "Other Work" /
        # "Project Management" still count.
        all_text_parts = []
        for r in sorted_reports:
            for k, v in (r.data or {}).items():
                if k.startswith("__") or v is None:
                    continue
                all_text_parts.append(str(v))
        big_text = " ".join(all_text_parts)

        project_counts: dict[str, int] = {}
        for name in PROJECT_NAMES:
            c = len(re.findall(rf"\b{re.escape(name)}\b", big_text, re.I))
            if c > 0:
                project_counts[name] = project_counts.get(name, 0) + c
        for m in CAPACITY_SITE_RE.finditer(big_text):
            cap = m.group(1).strip()
            if any(cap.lower() in k.lower() or k.lower() in cap.lower() for k in project_counts):
                continue
            project_counts[cap] = project_counts.get(cap, 0) + 1

        sorted_projects = sorted(project_counts.items(), key=lambda x: -x[1])[:6]
        projects_text = ", ".join(f"{n}: {c}" for n, c in sorted_projects) or _DASH
        total_projects = len(project_counts)

        user_rows.append({
            "name": full_name,
            "projects": projects_text,
            "total_projects": total_projects,
            "designs_made": designs_count,
            "site_visits": site_visit_days,
        })
    user_rows.sort(key=lambda r: r["name"].lower())

    row_idx = 5
    for row in user_rows:
        _put(ws, row_idx, 2, row["name"], font=_NAME_FONT, fill=_ROW_FILL, align=body_left)
        _put(ws, row_idx, 3, row["projects"],
             font=_BODY_FONT if row["projects"] != _DASH else _MUTED_FONT,
             fill=_ROW_FILL, align=body_left)
        _put(ws, row_idx, 4, row["total_projects"], font=num_font, fill=_ROW_FILL, align=right_align)
        _put(ws, row_idx, 5, row["designs_made"], font=num_font, fill=_ROW_FILL, align=right_align)
        _put(ws, row_idx, 6, row["site_visits"], font=num_font, fill=_ROW_FILL, align=right_align)
        row_idx += 1

    if user_rows:
        _put(ws, row_idx, 2, "Total", font=_TOTAL_FONT, fill=_TOTAL_FILL, align=body_left)
        _put(ws, row_idx, 3, "", fill=_TOTAL_FILL)
        # Total Projects in the footer = count of DISTINCT projects across
        # all employees (not the sum of each row, which would double-count
        # shared projects).
        all_projects: set[str] = set()
        for row in user_rows:
            if row["projects"] == _DASH:
                continue
            for chunk in row["projects"].split(","):
                name = chunk.split(":")[0].strip().lower()
                if name:
                    all_projects.add(name)
        _put(ws, row_idx, 4, len(all_projects),
             font=_TOTAL_FONT, fill=_TOTAL_FILL, align=right_align)
        _put(ws, row_idx, 5,
             sum(r["designs_made"] for r in user_rows),
             font=_TOTAL_FONT, fill=_TOTAL_FILL, align=right_align)
        _put(ws, row_idx, 6,
             sum(r["site_visits"] for r in user_rows),
             font=_TOTAL_FONT, fill=_TOTAL_FILL, align=right_align)

    summary_last_row = row_idx if user_rows else row_idx - 1
    _apply_table_borders(
        ws, header_row=4,
        last_row=summary_last_row,
        first_col=2, last_col=last_col,
    )
    _append_daily_reports_section(ws, group, start_row=summary_last_row)
    ws.freeze_panes = "C5"
    _set_print_landscape_fit(ws)


def _build_logistics_detail_sheet(ws, group, *, range_label: str):
    """Logistics department — 9 aggregated columns extracted from the
    free-text daily fields.  Each cell turns paragraphs of activity into
    a useful tally (transporter/warehouse names with hit counts, invoice
    counts, day counts, etc.).
    """
    last_col = 11
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 22                    # Employee
    widths = [44, 44, 14, 18, 12, 14, 14, 14, 14]           # 9 data columns
    for i, w in enumerate(widths, start=3):
        ws.column_dimensions[get_column_letter(i)].width = w

    _merge_title(ws, 1, last_col, "Logistics — summary table", _TITLE_FONT)
    ws.row_dimensions[1].height = 26
    n_users = len(group["users"])
    n_reports = sum(len(rs) for rs in group["users"].values())
    subtitle = (
        f"{range_label}  |  "
        f"{n_users} {'employee' if n_users == 1 else 'employees'}  ·  "
        f"{n_reports} reports submitted"
    )
    _merge_title(ws, 2, last_col, subtitle, _SUBTITLE_FONT)

    headers = [
        "Employee",
        "Transporter Coordination",
        "Warehouse Coordination",
        "Courier Tracking",
        "Stock / Flasher Invoices",
        "Verified Bills",
        "Portal Works (days)",
        "Booked Vehicles",
        "Coordination Calls",
        "Courier Dispatch",
    ]
    for i, h in enumerate(headers, start=2):
        _put(ws, 4, i, h, font=_HEADER_FONT, fill=_HEADER_FILL, align=_CENTER)
    ws.row_dimensions[4].height = 32

    num_font = Font(size=10, color="1A1A1A")
    right_align = Alignment(horizontal="center", vertical="center", wrap_text=False)
    body_left = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Known names — looked up case-insensitively in the pooled free-text logs.
    TRANSPORTERS = [
        "Trackon", "Allcargo", "All Cargo", "Delhivery", "Truxcargo",
        "Trux Cargo", "WheelsEye", "Wheels Eye", "Rivigo", "FEDEX", "FedEx",
        "Aryan", "TCI", "Porter", "Bike Porter", "Blue Dart", "DTDC",
        "Safexpress", "Gati", "VRL", "Mahindra", "ECom Express", "DHL",
        "Shree Maruti", "Shree Anjani", "VTrans", "V-Trans", "VXpress",
        "DRS", "DRS Group", "Spoton", "Maruti Courier",
    ]
    WAREHOUSES = [
        "AQDAS", "Logiwiz", "Rama", "MJ Infa", "Yusen", "Chennai",
        "Pune", "Ludhiana", "Nhava Sheva", "Maharashtra", "Okhla",
        "Bikaner", "Delhi warehouse", "Sonipat", "Bhiwandi", "Gurugram",
        "Noida", "Faridabad", "Mumbai warehouse", "Bangalore warehouse",
        "Hyderabad warehouse",
    ]

    # Invoice / waybill IDs across formats: OS-MAY26-040-DL, DC-MAY26-0009-DL,
    # GR-..., LR-..., AWB.... — be permissive on the prefix and trailing parts.
    INVOICE_RE = re.compile(
        r"\b(?:OS|DC|GR|LR|AWB)[-/][A-Z0-9]+(?:[-/][A-Z0-9]+)+\b",
        re.I,
    )

    # Generic phrase scanner — "<header containing keyword>: N" or "... – N".
    HEADER_NUM_RE = re.compile(r"([^\n.;]{1,120}?)\s*[–\-:]\s*(\d+)\b", re.I)

    # Per-metric keyword patterns (matched against the phrase header).
    VERIFIED_KW = re.compile(r"\b(?:verif|bill\s+verif|verified\s+bill)", re.I)
    BOOKED_KW = re.compile(
        r"\b(?:book(?:ed|ing)?\s+vehicle|vehicle\s+book|truck\s+book|"
        r"vehicle\s+arranged|vehicle\s+placed)",
        re.I,
    )
    CALL_KW = re.compile(
        r"\b(?:coordination\s+call|call(?:ed|s)?|spoke|"
        r"communicated|coordinat(?:ed|ion))",
        re.I,
    )
    DISPATCH_KW = re.compile(
        r"\b(?:courier\s+dispatch|dispatch(?:ed|es)?|sent\s+courier|"
        r"shipment\s+dispatch)",
        re.I,
    )
    PORTAL_KW = re.compile(
        r"\b(?:portal|sap|erp|tally|sunlight|ornate\s+portal)",
        re.I,
    )
    TRACKING_KW = re.compile(
        r"\b(?:tracking|courier\s+track|shipment\s+track|track(?:ing|ed)?)",
        re.I,
    )
    STOCK_INV_KW = re.compile(
        r"\b(?:stock\s+invoice|flasher\s+invoice|stock|flasher|invoice)",
        re.I,
    )

    _NULL_TOKENS = {"", "na", "n/a", "n.a.", "n.a", "-", "—", "nil", "none", "on leave"}
    def _has_content(v):
        s = ("" if v is None else str(v)).strip()
        return s.lower() not in _NULL_TOKENS

    def _count_hits(names, text):
        """Returns dict {display_name: count} for each known name seen in text."""
        if not text:
            return {}
        out = {}
        for n in names:
            c = len(re.findall(rf"\b{re.escape(n)}\b", text, re.I))
            if c > 0:
                out[n] = out.get(n, 0) + c
        return out

    def _format_names_with_counts(counts, top=5):
        if not counts:
            return _DASH
        sorted_items = sorted(counts.items(), key=lambda x: -x[1])[:top]
        return ", ".join(f"{name}: {c}" for name, c in sorted_items)

    def _sum_phrase_counts(text, keyword_pattern):
        """Sum every '<header containing keyword>: N' occurrence."""
        if not text:
            return 0
        total = 0
        for m in HEADER_NUM_RE.finditer(text):
            header = m.group(1).strip()
            if keyword_pattern.search(header):
                try:
                    total += int(m.group(2))
                except ValueError:
                    pass
        return total

    user_rows = []
    for _, user_reports in group["users"].items():
        sorted_reports = sorted(user_reports, key=lambda r: r.date or date_type.min)
        u = sorted_reports[0].user
        full_name = ""
        if u:
            full_name = (
                f"{(u.first_name or '').strip()} {(u.last_name or '').strip()}".strip()
                or u.username
            )

        # Pool EVERY field value across all reports — robust to whatever the
        # actual DB field keys are.
        all_text_parts = []
        for r in sorted_reports:
            data = r.data or {}
            for k, v in data.items():
                if k.startswith("__") or v is None:
                    continue
                all_text_parts.append(str(v))
        big_text = "\n".join(all_text_parts)

        # Name lookups across the whole employee corpus.
        transporter_counts = _count_hits(TRANSPORTERS, big_text)
        warehouse_counts = _count_hits(WAREHOUSES, big_text)

        # Invoice / waybill IDs (distinct).
        invoice_ids = set(INVOICE_RE.findall(big_text))

        # Counter metrics: prefer phrase-scanned numbers; fall back to
        # ID counts / phrase frequencies when no "X – N" line was found.
        courier_tracking = max(
            _sum_phrase_counts(big_text, TRACKING_KW),
            len(invoice_ids),
        )
        stock_invoices = max(
            _sum_phrase_counts(big_text, STOCK_INV_KW),
            len(invoice_ids),
        )
        verified_bills = max(
            _sum_phrase_counts(big_text, VERIFIED_KW),
            len(re.findall(r"\bverif\w+\b", big_text, re.I)),
        )

        # Portal Works = days with any portal-related content.
        portal_days = 0
        for r in sorted_reports:
            day_text_parts = []
            for k, v in (r.data or {}).items():
                if k.startswith("__") or v is None:
                    continue
                day_text_parts.append(str(v))
            day_text = " ".join(day_text_parts)
            if PORTAL_KW.search(day_text):
                portal_days += 1

        booked_vehicles = max(
            _sum_phrase_counts(big_text, BOOKED_KW),
            len(re.findall(
                r"\b(?:vehicle|truck)s?\s+(?:book(?:ed|ing)?|arranged|placed)\b",
                big_text, re.I,
            )),
        )
        coord_calls = max(
            _sum_phrase_counts(big_text, CALL_KW),
            len(re.findall(
                r"\b(?:call(?:ed|s)?|spoke)\b", big_text, re.I,
            )),
        )
        courier_dispatch = max(
            _sum_phrase_counts(big_text, DISPATCH_KW),
            len(invoice_ids),
        )

        user_rows.append({
            "name": full_name,
            "transporter": _format_names_with_counts(transporter_counts),
            "warehouse": _format_names_with_counts(warehouse_counts),
            "courier_tracking": courier_tracking,
            "stock_invoices": stock_invoices,
            "verified_bills": verified_bills,
            "portal": portal_days,
            "booked_vehicles": booked_vehicles,
            "coord_calls": coord_calls,
            "courier_dispatch": courier_dispatch,
        })
    user_rows.sort(key=lambda r: r["name"].lower())

    NUMERIC_KEYS = (
        "courier_tracking", "stock_invoices", "verified_bills",
        "portal", "booked_vehicles", "coord_calls", "courier_dispatch",
    )

    row_idx = 5
    for row in user_rows:
        _put(ws, row_idx, 2, row["name"], font=_NAME_FONT, fill=_ROW_FILL, align=body_left)
        _put(ws, row_idx, 3, row["transporter"],
             font=_BODY_FONT if row["transporter"] != _DASH else _MUTED_FONT,
             fill=_ROW_FILL, align=body_left)
        _put(ws, row_idx, 4, row["warehouse"],
             font=_BODY_FONT if row["warehouse"] != _DASH else _MUTED_FONT,
             fill=_ROW_FILL, align=body_left)
        for i, key in enumerate(NUMERIC_KEYS, start=5):
            _put(ws, row_idx, i, row[key], font=num_font, fill=_ROW_FILL, align=right_align)
        row_idx += 1

    if user_rows:
        _put(ws, row_idx, 2, "Total", font=_TOTAL_FONT, fill=_TOTAL_FILL, align=body_left)
        _put(ws, row_idx, 3, "", fill=_TOTAL_FILL)
        _put(ws, row_idx, 4, "", fill=_TOTAL_FILL)
        for i, key in enumerate(NUMERIC_KEYS, start=5):
            _put(ws, row_idx, i,
                 sum(r[key] for r in user_rows),
                 font=_TOTAL_FONT, fill=_TOTAL_FILL, align=right_align)

    summary_last_row = row_idx if user_rows else row_idx - 1
    _apply_table_borders(
        ws, header_row=4,
        last_row=summary_last_row,
        first_col=2, last_col=last_col,
    )
    _append_daily_reports_section(ws, group, start_row=summary_last_row)
    ws.freeze_panes = "C5"
    _set_print_landscape_fit(ws)


def _build_production_detail_sheet(ws, group, *, range_label: str):
    """Production department — 2 columns:
      - Total Production: per-part breakdown like "M-purlin: 240,
        Bridge clamps: 700, Walkway: 25" — parts named first, quantities
        after, sorted by quantity descending.
      - Other Works Summary: condensed prose from the "Other Works" field.
    """
    last_col = 4  # A margin + B Employee + C Total Production + D Other Works
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 60   # Wider — fits "part: qty, …" list
    ws.column_dimensions["D"].width = 70

    _merge_title(ws, 1, last_col, "Production — summary table", _TITLE_FONT)
    ws.row_dimensions[1].height = 26
    n_users = len(group["users"])
    n_reports = sum(len(rs) for rs in group["users"].values())
    subtitle = (
        f"{range_label}  |  "
        f"{n_users} {'employee' if n_users == 1 else 'employees'}  ·  "
        f"{n_reports} reports submitted"
    )
    _merge_title(ws, 2, last_col, subtitle, _SUBTITLE_FONT)

    headers = ["Employee", "Total Production", "Other Works Summary"]
    for i, h in enumerate(headers, start=2):
        _put(ws, 4, i, h, font=_HEADER_FONT, fill=_HEADER_FILL, align=_CENTER)
    ws.row_dimensions[4].height = 28

    num_font = Font(size=10, color="1A1A1A")
    right_align = Alignment(horizontal="center", vertical="center", wrap_text=False)
    body_left = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Match "<number> nos/pcs/pieces/units" AND "nos/pcs/pieces/units <number>".
    QTY_RE_AFTER = re.compile(
        r"(\d+(?:,\d{3})*)\s*(?:nos\b|pieces?\b|pcs\b|units?\b)", re.I,
    )
    QTY_RE_BEFORE = re.compile(
        r"(?:nos|pieces?|pcs|units?)\s*[-:]?\s*(\d+(?:,\d{3})*)", re.I,
    )
    # Words that get stripped from the start of each sentence when guessing
    # the part name — they're noise, not part identifiers.
    _PART_STOPWORDS = {
        "project", "today", "production", "ready", "final", "complete",
        "completed", "work", "the", "for", "and", "of", "with", "is",
        "are", "ongoing", "ongoing.", "press", "machine", "die",
        "raw", "material", "rm", "cutting", "banding", "assembly",
        "unloading", "loading", "setup", "maintenance", "corner",
        "from", "to", "into", "on", "by", "pear", "amparlin", "perching",
        "ribiting", "carnor", "fitting", "fir", "ka", "for", "and",
    }

    def _normalize_part(name):
        """Lowercase + collapse separators so 'M-purlin' / 'M purlin' /
        'Mparlin' / 'M-Purline' all map to the same bucket."""
        n = name.lower()
        n = re.sub(r"[^a-z0-9]+", "", n)
        # Common variants normalisation.
        n = n.replace("amparlin", "mpurlin").replace("mparlin", "mpurlin")
        n = n.replace("purline", "purlin")
        return n

    def extract_part_qty_pairs(text):
        """Returns list of (display_name, qty) pulled from a production
        log.  Looks at each clause/sentence, finds the first quantity in
        it, and uses the preceding meaningful words as the part name."""
        if not text:
            return []
        pairs = []
        for sentence in re.split(r"[\n.;]+", str(text)):
            s = sentence.strip()
            if not s:
                continue
            s = re.sub(r"^\d{4}-\d{2}-\d{2}\s*[:\-]?\s*", "", s).strip()
            s = re.sub(r"^\d+[.\)]\s*", "", s).strip()
            if not s:
                continue
            m = QTY_RE_AFTER.search(s) or QTY_RE_BEFORE.search(s)
            if not m:
                continue
            try:
                qty = int(m.group(1).replace(",", ""))
            except ValueError:
                continue
            before = s[: m.start()].strip(" -:,")
            words = re.findall(r"[A-Za-z][A-Za-z\-]*", before)
            meaningful = [
                w for w in words
                if w.lower() not in _PART_STOPWORDS and len(w) >= 2
            ]
            if not meaningful:
                continue
            # First 1-2 meaningful words name the part.
            part_name = " ".join(meaningful[:2])
            pairs.append((part_name, qty))
        return pairs

    user_rows = []
    for _, user_reports in group["users"].items():
        sorted_reports = sorted(user_reports, key=lambda r: r.date or date_type.min)
        u = sorted_reports[0].user
        full_name = ""
        if u:
            full_name = (
                f"{(u.first_name or '').strip()} {(u.last_name or '').strip()}".strip()
                or u.username
            )

        # Aggregate (part → total qty) across all reports.
        per_part: dict[str, dict] = {}
        for r in sorted_reports:
            data = r.data or {}
            for name, qty in extract_part_qty_pairs(data.get("todayProduction", "")):
                norm = _normalize_part(name)
                if not norm:
                    continue
                bucket = per_part.setdefault(norm, {"name": name, "qty": 0})
                bucket["qty"] += qty
        # Sort parts by quantity descending and format "Name: qty".
        sorted_parts = sorted(per_part.values(), key=lambda p: -p["qty"])
        production_text = ", ".join(
            f"{p['name']}: {p['qty']:,}" for p in sorted_parts
        ) if sorted_parts else _DASH
        total_qty = sum(p["qty"] for p in sorted_parts)

        other_text_pool = []
        for r in sorted_reports:
            ow = ((r.data or {}).get("otherWorks") or "").strip()
            if ow and ow.lower() not in ("na", "n/a", "n.a.", "-", "—", "nil", "none", "on leave"):
                other_text_pool.append(ow)
        other_summary = _condense_summary("\n".join(other_text_pool), max_chars=180)

        user_rows.append({
            "name": full_name,
            "production_text": production_text,
            "total_qty": total_qty,
            "other_summary": other_summary or _DASH,
        })
    user_rows.sort(key=lambda r: r["name"].lower())

    row_idx = 5
    for row in user_rows:
        _put(ws, row_idx, 2, row["name"], font=_NAME_FONT, fill=_ROW_FILL, align=body_left)
        _put(ws, row_idx, 3, row["production_text"],
             font=_BODY_FONT if row["production_text"] != _DASH else _MUTED_FONT,
             fill=_ROW_FILL, align=body_left)
        _put(ws, row_idx, 4, row["other_summary"],
             font=_BODY_FONT if row["other_summary"] != _DASH else _MUTED_FONT,
             fill=_ROW_FILL, align=body_left)
        row_idx += 1

    if user_rows:
        _put(ws, row_idx, 2, "Total", font=_TOTAL_FONT, fill=_TOTAL_FILL, align=body_left)
        _put(ws, row_idx, 3,
             f"{sum(r['total_qty'] for r in user_rows):,}",
             font=_TOTAL_FONT, fill=_TOTAL_FILL, align=right_align)
        _put(ws, row_idx, 4, "", fill=_TOTAL_FILL)

    summary_last_row = row_idx if user_rows else row_idx - 1
    _apply_table_borders(
        ws, header_row=4,
        last_row=summary_last_row,
        first_col=2, last_col=last_col,
    )
    _append_daily_reports_section(ws, group, start_row=summary_last_row)
    ws.freeze_panes = "C5"
    _set_print_landscape_fit(ws)


def _build_marketing_detail_sheet(ws, group, *, range_label: str):
    """Marketing department — 5 aggregated counter columns.  Each cell
    counts the number of DAYS the employee logged meaningful content in
    that activity area (empty / "N/A" / "Na" / "-" don't count).
    """
    last_col = 7  # A margin + B Employee + 5 counters
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 24
    for i, w in enumerate([22, 22, 18, 12, 16], start=3):
        ws.column_dimensions[get_column_letter(i)].width = w

    _merge_title(ws, 1, last_col, "Marketing — summary table", _TITLE_FONT)
    ws.row_dimensions[1].height = 26
    n_users = len(group["users"])
    n_reports = sum(len(rs) for rs in group["users"].values())
    subtitle = (
        f"{range_label}  |  "
        f"{n_users} {'employee' if n_users == 1 else 'employees'}  ·  "
        f"{n_reports} reports submitted"
    )
    _merge_title(ws, 2, last_col, subtitle, _SUBTITLE_FONT)

    headers = [
        "Employee", "Emails / Videos / Edited",
        "PPT / PDF / Brochures", "Content Writing", "SEO", "Reporting",
    ]
    for i, h in enumerate(headers, start=2):
        _put(ws, 4, i, h, font=_HEADER_FONT, fill=_HEADER_FILL, align=_CENTER)
    ws.row_dimensions[4].height = 28

    num_font = Font(size=10, color="1A1A1A")
    right_align = Alignment(horizontal="center", vertical="center", wrap_text=False)
    body_left = Alignment(horizontal="left", vertical="top", wrap_text=True)

    VIDEO_KEYS    = ("videoEditing",)
    PPTPDF_KEYS   = ("brochurePptPdfEdits", "creatives")
    CONTENT_KEYS  = ("contentWriting",)
    SEO_KEYS      = ("seo",)
    REPORT_KEYS   = ("reporting",)

    _NULL_TOKENS = {"", "na", "n/a", "n.a.", "n.a", "-", "—", "nil", "none", "on leave"}

    def _has_content(v):
        s = ("" if v is None else str(v)).strip()
        return s.lower() not in _NULL_TOKENS

    user_rows = []
    for _, user_reports in group["users"].items():
        sorted_reports = sorted(user_reports, key=lambda r: r.date or date_type.min)
        u = sorted_reports[0].user
        full_name = ""
        if u:
            full_name = (
                f"{(u.first_name or '').strip()} {(u.last_name or '').strip()}".strip()
                or u.username
            )

        def count_days(keys):
            return sum(
                1 for r in sorted_reports
                if any(_has_content((r.data or {}).get(k, "")) for k in keys)
            )

        user_rows.append({
            "name": full_name,
            "video": count_days(VIDEO_KEYS),
            "pptpdf": count_days(PPTPDF_KEYS),
            "content": count_days(CONTENT_KEYS),
            "seo": count_days(SEO_KEYS),
            "report": count_days(REPORT_KEYS),
        })
    user_rows.sort(key=lambda r: r["name"].lower())

    row_idx = 5
    for row in user_rows:
        _put(ws, row_idx, 2, row["name"], font=_NAME_FONT, fill=_ROW_FILL, align=body_left)
        for i, key in enumerate(("video", "pptpdf", "content", "seo", "report")):
            _put(ws, row_idx, 3 + i, row[key],
                 font=num_font, fill=_ROW_FILL, align=right_align)
        row_idx += 1

    if user_rows:
        _put(ws, row_idx, 2, "Total", font=_TOTAL_FONT, fill=_TOTAL_FILL, align=body_left)
        for i, key in enumerate(("video", "pptpdf", "content", "seo", "report")):
            _put(ws, row_idx, 3 + i,
                 sum(r[key] for r in user_rows),
                 font=_TOTAL_FONT, fill=_TOTAL_FILL, align=right_align)

    summary_last_row = row_idx if user_rows else row_idx - 1
    _apply_table_borders(
        ws, header_row=4,
        last_row=summary_last_row,
        first_col=2, last_col=last_col,
    )
    _append_daily_reports_section(ws, group, start_row=summary_last_row)
    ws.freeze_panes = "C5"
    _set_print_landscape_fit(ws)


def _build_project_detail_sheet(ws, group, *, range_label: str):
    """Project department — 2 condensed columns:
      - Total Site Visit  (count of reports the employee submitted)
      - Work on Site      (distinct snippets from Work Done + Work in Progress,
                           short list)
    """
    last_col = 4  # A margin + B Employee + C Total Site Visit + D Work on Site

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 90

    _merge_title(ws, 1, last_col, "Project — summary table", _TITLE_FONT)
    ws.row_dimensions[1].height = 26
    n_users = len(group["users"])
    n_reports = sum(len(rs) for rs in group["users"].values())
    subtitle = (
        f"{range_label}  |  "
        f"{n_users} {'employee' if n_users == 1 else 'employees'}  ·  "
        f"{n_reports} reports submitted"
    )
    _merge_title(ws, 2, last_col, subtitle, _SUBTITLE_FONT)

    headers = ["Employee", "Total Site Visit", "Work on Site"]
    for i, h in enumerate(headers, start=2):
        _put(ws, 4, i, h, font=_HEADER_FONT, fill=_HEADER_FILL, align=_CENTER)
    ws.row_dimensions[4].height = 28

    num_font = Font(size=10, color="1A1A1A")
    right_align = Alignment(horizontal="center", vertical="center", wrap_text=False)
    body_left = Alignment(horizontal="left", vertical="top", wrap_text=True)

    WORK_KEYS = ("workDone", "workInProgress")

    user_rows = []
    for _, user_reports in group["users"].items():
        sorted_reports = sorted(user_reports, key=lambda r: r.date or date_type.min)
        u = sorted_reports[0].user
        full_name = ""
        if u:
            full_name = (
                f"{(u.first_name or '').strip()} {(u.last_name or '').strip()}".strip()
                or u.username
            )

        # Site visit count = number of reports submitted in range.
        visit_count = len(sorted_reports)

        # Work on Site — compact join of distinct headline phrases from
        # Work Done + Work in Progress, packed into ~150 chars.  Trailing
        # "…" hints at overflow without a "(+N more)" tag.
        seen = set()
        distinct = []
        for r in sorted_reports:
            data = r.data or {}
            for k in WORK_KEYS:
                v = (data.get(k) or "").strip()
                if not v or v.lower() == "on leave":
                    continue
                key_lower = v.lower()
                if key_lower in seen:
                    continue
                seen.add(key_lower)
                distinct.append(v)
        work_text = _compact_join(distinct, max_chars=150)

        user_rows.append({
            "name": full_name,
            "visit_count": visit_count,
            "work_text": work_text or _DASH,
        })
    user_rows.sort(key=lambda r: r["name"].lower())

    row_idx = 5
    for row in user_rows:
        _put(ws, row_idx, 2, row["name"], font=_NAME_FONT, fill=_ROW_FILL, align=body_left)
        _put(ws, row_idx, 3, row["visit_count"], font=num_font, fill=_ROW_FILL, align=right_align)
        _put(ws, row_idx, 4,
             row["work_text"],
             font=_BODY_FONT if row["work_text"] != _DASH else _MUTED_FONT,
             fill=_ROW_FILL, align=body_left)
        row_idx += 1

    if user_rows:
        _put(ws, row_idx, 2, "Total", font=_TOTAL_FONT, fill=_TOTAL_FILL, align=body_left)
        _put(ws, row_idx, 3,
             sum(r["visit_count"] for r in user_rows),
             font=_TOTAL_FONT, fill=_TOTAL_FILL, align=right_align)
        _put(ws, row_idx, 4, "", fill=_TOTAL_FILL)

    summary_last_row = row_idx if user_rows else row_idx - 1
    _apply_table_borders(
        ws, header_row=4,
        last_row=summary_last_row,
        first_col=2, last_col=last_col,
    )
    _append_daily_reports_section(ws, group, start_row=summary_last_row)
    ws.freeze_panes = "C5"
    _set_print_landscape_fit(ws)


def _build_combined_summary_sheet(ws, other_groups, *, range_label: str, today_label: str):
    """One-table combined summary for every department except Sales & Inside
    Sales (those get their own detailed tabs).  Columns:
    Department | # | Employee | Focus Area | Key Activities.
    """
    last_col = 6
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 20   # Department
    ws.column_dimensions["C"].width = 5    # #
    ws.column_dimensions["D"].width = 22   # Employee
    ws.column_dimensions["E"].width = 22   # Focus Area (role)
    ws.column_dimensions["F"].width = 80   # Key Activities (wider since it's the tail)

    n_employees = sum(len(g["users"]) for g in other_groups)
    n_reports = sum(len(rs) for g in other_groups for rs in g["users"].values())

    _merge_title(
        ws, 1, last_col,
        "All Department Employee Summary (other departments)",
        _TITLE_FONT,
    )
    ws.row_dimensions[1].height = 28
    subtitle = (
        f"Date range: {range_label}  |  "
        f"{n_employees} employees across {len(other_groups)} departments  |  "
        f"{n_reports} reports  |  Generated: {today_label}  "
        f"(Inside Sales & Sales shown on their own detailed tabs)"
    )
    _merge_title(ws, 2, last_col, subtitle, _SUBTITLE_FONT)

    headers = [
        "Department", "#", "Employee", "Focus Area",
        f"Key Activities ({range_label})",
    ]
    for i, h in enumerate(headers, start=2):
        _put(ws, 4, i, h, font=_HEADER_FONT, fill=_HEADER_FILL, align=_CENTER)
    ws.row_dimensions[4].height = 28

    sorted_groups = sorted(other_groups, key=_dept_sort_key)

    row_idx = 5
    for group in sorted_groups:
        d = group["dept"]
        dept_name = d.name if d else "No Department"
        user_entries = []
        for _, user_reports in group["users"].items():
            sorted_reports = sorted(user_reports, key=lambda r: r.date or date_type.min)
            user_entries.append(_summarise_user_reports(sorted_reports))
        user_entries.sort(key=lambda t: t[0].lower())

        for idx, (name, role, key_acts, _meetings, _revenue, _count) in enumerate(user_entries, start=1):
            # Department badge only on first row of the dept block.
            if idx == 1:
                _put(ws, row_idx, 2, dept_name,
                     font=_BADGE_FONT, fill=_BADGE_FILL, align=_CENTER)
            else:
                _put(ws, row_idx, 2, "", fill=_ROW_FILL)
            _put(ws, row_idx, 3, idx, font=_BODY_FONT, fill=_ROW_FILL, align=_CENTER)
            _put(ws, row_idx, 4, name, font=_NAME_FONT, fill=_ROW_FILL, align=_LEFT_TOP_WRAP)
            _put(ws, row_idx, 5, role or _DASH,
                 font=_BODY_FONT if role else _MUTED_FONT,
                 fill=_ROW_FILL, align=_LEFT_TOP_WRAP)
            _put(ws, row_idx, 6, key_acts or _DASH,
                 font=_BODY_FONT if key_acts else _MUTED_FONT,
                 fill=_ROW_FILL, align=_LEFT_TOP_WRAP)
            row_idx += 1

    summary_last_row = row_idx - 1
    _apply_table_borders(ws, header_row=4, last_row=summary_last_row,
                         first_col=2, last_col=last_col)

    # One "Daily Reports" block per department, stacked beneath the summary.
    cursor = summary_last_row
    for group in sorted_groups:
        d = group["dept"]
        dept_name = d.name if d else "No Department"
        cursor = _append_daily_reports_section(
            ws, group, start_row=cursor,
            dept_label=f"{dept_name} — Daily Reports",
        )

    ws.freeze_panes = "D5"
    _set_print_landscape_fit(ws)


def _empty_xlsx(filename: str) -> Response:
    wb = Workbook()
    ws = wb.active
    ws.title = "No reports"
    ws["B2"] = "No reports match the current filters."
    ws["B2"].font = _SUBTITLE_FONT
    buf = io.BytesIO()
    wb.save(buf)
    data = buf.getvalue()
    return Response(
        content=data,
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Content-Length": str(len(data)),
        },
    )


@router.get("/export.csv")
def export_csv(
    employee: int | None = Query(None, description="Filter by employee id"),
    department: str | None = Query(None, description="Filter by department slug"),
    start: date_type | None = Query(None, description="Inclusive start date"),
    end: date_type | None = Query(None, description="Inclusive end date"),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """Single-file CSV mirror of the styled XLSX summary.

    Columns: Department, Employee, Role, Key Activities, Meetings, Revenue (₹).
    Rows ordered Sales → Inside Sales → other departments alphabetically.
    Key Activities lines are joined with " | " so each row stays on one line
    in spreadsheet apps that don't auto-wrap.  UTF-8 with BOM so Excel
    detects the encoding correctly on Windows.
    """
    q = db.query(DailyReport).join(User, DailyReport.user_id == User.id)
    if employee is not None:
        q = q.filter(DailyReport.user_id == employee)
    if department:
        dept = db.query(Department).filter(Department.slug == department).first()
        if not dept:
            return _empty_csv("daily-reports.csv")
        q = q.filter(User.department_id == dept.id)
    if start:
        q = q.filter(DailyReport.date >= start)
    if end:
        q = q.filter(DailyReport.date <= end)

    rows = (
        q.order_by(DailyReport.date.desc(), DailyReport.user_id)
        .limit(50000)
        .all()
    )

    by_dept: dict[int | None, dict] = {}
    for r in rows:
        if (r.data or {}).get("__leave__") == "1":
            continue
        u = r.user
        d = u.department if u else None
        key = d.id if d else None
        if key not in by_dept:
            by_dept[key] = {"dept": d, "users": {}}
        by_dept[key]["users"].setdefault(u.id if u else 0, []).append(r)

    buf = io.StringIO()
    writer = csv.writer(buf)
    range_label = _format_range_label(start, end)
    writer.writerow([f"All Department Employee Summary Report — {range_label}"])
    writer.writerow([])
    writer.writerow(
        ["Department", "Employee", "Role / Designation", "Key Activities", "Meetings", "Revenue (₹)"]
    )

    for group in sorted(by_dept.values(), key=_dept_sort_key):
        d = group["dept"]
        dept_name = d.name if d else "No Department"
        # Sort users in each dept by name.
        user_entries = []
        for _, user_reports in group["users"].items():
            sorted_reports = sorted(user_reports, key=lambda r: r.date or date_type.min)
            user_entries.append(_summarise_user_reports(sorted_reports))
        user_entries.sort(key=lambda t: t[0].lower())

        for name, role, key_acts, meetings, revenue, _count in user_entries:
            # Collapse multi-line activities into one CSV cell (single line).
            key_acts_line = (key_acts or "").replace("\n", " | ")
            meetings_str = ""
            if meetings is not None:
                meetings_str = (
                    str(int(meetings))
                    if float(meetings).is_integer()
                    else f"{meetings:.2f}"
                )
            revenue_str = _format_revenue(revenue) if revenue is not None else ""
            writer.writerow([
                dept_name,
                name,
                role or "",
                key_acts_line,
                meetings_str,
                revenue_str,
            ])
        # Blank separator row between departments for visual breathing room.
        writer.writerow([])

    # UTF-8 BOM helps Excel for Windows pick the right encoding.
    payload = buf.getvalue().encode("utf-8-sig")

    parts = ["daily-reports"]
    if start:
        parts.append(start.isoformat())
    if end:
        parts.append(end.isoformat())
    if not (start or end):
        parts.append(date_type.today().isoformat())
    filename = "_".join(parts) + ".csv"

    return Response(
        content=payload,
        media_type="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Content-Length": str(len(payload)),
        },
    )


def _empty_csv(filename: str) -> Response:
    payload = "No reports match the current filters.\n".encode("utf-8-sig")
    return Response(
        content=payload,
        media_type="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Content-Length": str(len(payload)),
        },
    )


@router.post("", response_model=ReportOut, status_code=status.HTTP_201_CREATED)
def upsert_report(
    payload: ReportIn,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Create or update a daily report for (user, date).

    Regular employees can only submit/edit their own reports.  HR users may
    submit/edit on behalf of any employee by passing `user_id` in the body —
    this is how the dashboard's "Edit" feature works.
    """
    target_user_id = user.id
    if payload.user_id is not None and payload.user_id != user.id:
        # Confirm the target exists before we evaluate permission — it's the
        # same 404 either way.
        target = db.query(User).filter(User.id == payload.user_id).first()
        if not target:
            raise HTTPException(
                status.HTTP_404_NOT_FOUND,
                f"No user with id={payload.user_id}",
            )
        # Who can submit on behalf of whom:
        #   - HR can submit for anyone.
        #   - A "team head" can submit for any employee in their MANAGED
        #     department.  Managed dept is `team_head_dept_id` when set,
        #     otherwise falls back to the team head's own department.
        #   - Everyone else: forbidden.
        is_team_head = bool(getattr(user, "is_team_head", False))
        managed_dept_id = (
            getattr(user, "team_head_dept_id", None) or user.department_id
        )
        target_in_managed = (
            managed_dept_id is not None
            and target.department_id == managed_dept_id
        )
        if user.role != "hr" and not (is_team_head and target_in_managed):
            raise HTTPException(
                status.HTTP_403_FORBIDDEN,
                "You don't have permission to submit reports for this employee.",
            )
        target_user_id = target.id

    existing = db.query(DailyReport).filter(
        and_(DailyReport.user_id == target_user_id, DailyReport.date == payload.date)
    ).first()

    now = datetime.now(timezone.utc)
    if existing:
        # Employees can now edit their OWN past reports.  HR can edit anyone's
        # (the HR-on-behalf-of override earlier in this function is already
        # gated to HR-only, so non-HR callers can only ever touch their own
        # row).
        existing.data = payload.data
        existing.submitted_at = now
        report = existing
    else:
        report = DailyReport(
            user_id=target_user_id,
            date=payload.date,
            data=payload.data,
            submitted_at=now,
            created_at=now,
        )
        db.add(report)
    db.commit()
    db.refresh(report)
    return report


@router.post("/leave", response_model=list[ReportOut], status_code=status.HTTP_201_CREATED)
def apply_leave(
    payload: LeaveIn,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Mark N consecutive days as leave for the current user (or any user, if HR).

    Each day becomes a DailyReport row where every department field is filled
    with "On Leave" — so all tables, summaries, and detail views render the
    leave state automatically.  Two hidden marker keys (`__leave__`,
    `__leave_reason__`) are also stored so callers can detect leave rows.
    """
    if payload.days < 1 or payload.days > 60:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "days must be between 1 and 60")

    target = user
    if payload.user_id is not None and payload.user_id != user.id:
        if user.role != "hr":
            raise HTTPException(status.HTTP_403_FORBIDDEN, "Only HR can apply leave for others")
        target = db.query(User).filter(User.id == payload.user_id).first()
        if not target:
            raise HTTPException(status.HTTP_404_NOT_FOUND, f"No user with id={payload.user_id}")

    fields = (target.department.report_fields or []) if target.department else []
    leave_payload: dict[str, str] = {f["key"]: "On Leave" for f in fields}
    leave_payload["__leave__"] = "1"
    leave_payload["__leave_reason__"] = payload.reason or ""

    now = datetime.now(timezone.utc)
    created: list[DailyReport] = []
    for i in range(payload.days):
        d = payload.start_date + timedelta(days=i)
        existing = (
            db.query(DailyReport)
            .filter(and_(DailyReport.user_id == target.id, DailyReport.date == d))
            .first()
        )
        if existing:
            existing.data = leave_payload
            existing.submitted_at = now
            created.append(existing)
        else:
            row = DailyReport(
                user_id=target.id,
                date=d,
                data=leave_payload,
                submitted_at=now,
                created_at=now,
            )
            db.add(row)
            created.append(row)
    db.commit()
    for r in created:
        db.refresh(r)
    return created


@router.get("/missing-today", response_model=list[int])
def missing_today(
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """Returns the list of user ids that have NOT submitted a report today.

    Weekends (Sat / Sun) return an empty list — employees aren't expected to
    submit reports on non-working days, so nobody is "missing".
    """
    today = date_type.today()
    if today.weekday() >= 5:  # Mon=0, Fri=4, Sat=5, Sun=6
        return []
    submitted_ids = {r.user_id for r in db.query(DailyReport).filter(DailyReport.date == today).all()}
    all_active = db.query(User.id).filter(User.is_active.is_(True), User.role != "hr").all()
    return [uid for (uid,) in all_active if uid not in submitted_ids]


@router.delete("/{report_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_report(
    report_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    report = db.query(DailyReport).filter(DailyReport.id == report_id).first()
    if not report:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Report not found")
    if report.user_id != user.id and user.role != "hr":
        raise HTTPException(status.HTTP_403_FORBIDDEN, "You can only delete your own reports")
    db.delete(report)
    db.commit()
    return None
