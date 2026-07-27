"""Employee expense claims with approval workflow.

Endpoints:
  POST   /api/expenses              — employee submits a new expense (optionally
                                       with a bill image as multipart upload).
  GET    /api/expenses              — list expenses.  Employees see only their
                                       own; HR + named approvers (TARINI, SMITA)
                                       see ALL.
  GET    /api/expenses/{id}/bill    — stream the bill image bytes.  Same
                                       visibility rules.
  POST   /api/expenses/{id}/decide  — HR / approver approves or rejects.
  DELETE /api/expenses/{id}         — owner can withdraw a PENDING expense;
                                       HR can delete any.

Bills live in MinIO under `expenses/{uuid}.{ext}` in the existing bucket.
"""
from __future__ import annotations

import io
import uuid
from datetime import date as date_type, datetime, timedelta, timezone
from pathlib import Path

from fastapi import (
    APIRouter,
    Body,
    Depends,
    File,
    Form,
    HTTPException,
    Query,
    Response,
    UploadFile,
    status,
)
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import func
from sqlalchemy.orm import Session

import storage
from auth import get_current_user
from database import get_db
from models import Expense, MonthlyExpenseNote, User
from schemas import ExpenseBillOut, ExpenseDecideIn, ExpenseOut, ExpensePatchIn


router = APIRouter(prefix="/api/expenses", tags=["expenses"])

# Hard-coded approver emails — TARINI + SMITA can approve / reject in addition
# to anyone with role=hr.  Stored uppercase for case-insensitive comparison.
APPROVER_EMAILS = {
    "tarini@ornatesolar.com",
    "smita@ornatesolar.com",
}

# Finance-team disbursal — these accounts can mark an approved expense as
# "paid".  Same email-based gate as the approver set so the role doesn't
# need to be remodelled.  Add more entries if Finance brings in cover.
FINANCE_APPROVER_EMAILS = {
    "shivangi@ornatesolar.com",
    "saif@ornatesolar.com",
}

# Allowed expense types — server-side enum so the frontend can't smuggle in
# arbitrary values.  "travel" requires `travel_type` to also be one of the
# travel sub-types below.
ALLOWED_EXPENSE_TYPES = {
    # Legacy values kept for backward-compat with existing rows ("material",
    # "fuel") even though they're no longer in the dropdown.
    "material", "food", "travel", "hotel", "fuel", "others",
    "officereimburse", "sitematerial", "officematerial",
}
ALLOWED_TRAVEL_TYPES = {
    "bus", "cab", "bike", "rapido", "car", "auto", "metro", "rickshaw", "other",
}
ALLOWED_MODES = {"cash", "upi", "card", "bank", "other", ""}

MAX_BILL_BYTES = 5 * 1024 * 1024   # per-file size cap
MAX_BILLS_PER_EXPENSE = 10          # upper bound on attachments per expense
ALLOWED_BILL_EXTENSIONS = {
    ".jpg", ".jpeg", ".png", ".webp", ".heic", ".pdf",
}
# Mime types accepted as the bill's Content-Type header from the browser.
ALLOWED_BILL_MIMETYPES = {
    "image/jpeg", "image/png", "image/webp", "image/heic", "image/heif",
    "application/pdf",
}


def _is_hr(user: User) -> bool:
    return user.role == "hr"


def _is_approver(user: User) -> bool:
    """HR users + the named approvers (TARINI / SMITA) can decide expenses."""
    if _is_hr(user):
        return True
    email = (user.email or "").strip().lower()
    return email in APPROVER_EMAILS


def _is_finance_approver(user: User) -> bool:
    """Finance team (Shivangi, Saif) — the accounts allowed to mark an
    approved expense as paid.  Kept separate from the approval gate; they do
    NOT get to approve/reject, only to disburse."""
    email = (user.email or "").strip().lower()
    return email in FINANCE_APPROVER_EMAILS


# Project coordinators — read-only visibility into a fixed set of team members'
# expenses + advances.  A coordinator is NOT an approver: they can see their
# team's rows but cannot approve / reject / hold / mark-paid.  Map each
# coordinator email to the exact set of team-member emails they may view.
PROJECT_COORDINATORS = {
    "arman@ornatesolar.com": {
        "vijay@ornatesolar.com",      # Vijay Kumar Pal
        "anil@ornatesolar.com",       # Anil Pathak
        "indra@ornatesolar.com",      # Indra Kumar
        "imran@ornatesolar.com",      # Imran Khan
        "biswanath@ornatesolar.com",  # Biswanath Nandi
        "afroj@ornatesolar.com",      # Afroj Alam
        "asim@ornatesolar.com",       # Asim Equbal
        "manoj2@ornatesolar.com",     # Manoj Kumar Yadav
    },
}


def _coordinator_team(user: User) -> set[str] | None:
    """The set of team-member emails this coordinator may view, or None if the
    user is not a project coordinator."""
    email = (user.email or "").strip().lower()
    return PROJECT_COORDINATORS.get(email)


def _coordinator_may_view(user: User, exp: Expense, db: Session) -> bool:
    """True if `user` is a coordinator and `exp` belongs to one of their team."""
    team = _coordinator_team(user)
    if not team:
        return False
    owner = db.get(User, exp.user_id)
    return owner is not None and (owner.email or "").strip().lower() in team


def _full_name(u: User | None) -> str:
    if not u:
        return ""
    first = (u.first_name or "").strip()
    last = (u.last_name or "").strip()
    return (first + " " + last).strip() or (u.username or "")


def _dept_name(u: User | None) -> str:
    if not u or not getattr(u, "department", None):
        return ""
    return (u.department.name or "")


def _to_out(exp: Expense) -> ExpenseOut:
    raw_bills = exp.bills or []
    bills_out = [
        ExpenseBillOut(index=i, filename=(b.get("filename") or f"bill-{i + 1}"))
        for i, b in enumerate(raw_bills)
    ]
    return ExpenseOut(
        id=exp.id,
        user_id=exp.user_id,
        user_name=_full_name(exp.user),
        user_department=_dept_name(exp.user),
        date=exp.date,
        mode=exp.mode or "",
        expense_type=exp.expense_type,
        travel_type=exp.travel_type or "",
        amount=exp.amount or 0,
        advance=exp.advance or 0,
        site_name=exp.site_name or "",
        remarks=exp.remarks or "",
        bills=bills_out,
        status=exp.status,
        decided_by_id=exp.decided_by_id,
        decided_by_name=_full_name(exp.decided_by),
        decided_at=exp.decided_at,
        decision_note=exp.decision_note or "",
        paid_by_id=exp.paid_by_id,
        paid_by_name=_full_name(exp.paid_by),
        paid_at=exp.paid_at,
        payment_ref=exp.payment_ref or "",
        paid_amount=exp.paid_amount,
        approved_by_name=exp.approved_by_name or "",
        created_at=exp.created_at,
    )


@router.post("", response_model=ExpenseOut, status_code=status.HTTP_201_CREATED)
async def create_expense(
    date: date_type = Form(...),
    mode: str = Form(""),
    expense_type: str = Form(...),
    travel_type: str = Form(""),
    amount: int = Form(...),
    advance: int = Form(0),
    site_name: str = Form(""),
    remarks: str = Form(""),
    # Multi-file: the browser sends one `bills` form field per file.  Single
    # files still work (the list arrives with one element).  We also accept
    # the legacy `bill` field name for backward compat if any client still
    # uses it; that file is appended after `bills`.
    bills: list[UploadFile] = File(default=[]),
    bill: UploadFile | None = File(None),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Employee creates a new expense.  Multiple bill files are supported."""
    expense_type = (expense_type or "").strip().lower()
    travel_type = (travel_type or "").strip().lower()
    mode = (mode or "").strip().lower()

    if expense_type not in ALLOWED_EXPENSE_TYPES:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            f"Invalid expense_type. Allowed: {sorted(ALLOWED_EXPENSE_TYPES)}",
        )
    if expense_type == "travel":
        if not travel_type:
            raise HTTPException(
                status.HTTP_400_BAD_REQUEST,
                "travel_type is required when expense_type is 'travel'",
            )
        if travel_type not in ALLOWED_TRAVEL_TYPES:
            raise HTTPException(
                status.HTTP_400_BAD_REQUEST,
                f"Invalid travel_type. Allowed: {sorted(ALLOWED_TRAVEL_TYPES)}",
            )
    else:
        travel_type = ""
    if mode and mode not in ALLOWED_MODES:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            f"Invalid mode. Allowed: {sorted(ALLOWED_MODES - {''})}",
        )
    if amount is None or amount < 0:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST, "amount must be a non-negative integer"
        )

    # Combine new + legacy bill fields, drop any "empty" UploadFile entries
    # (browsers sometimes send an empty file with no filename when the input
    # was left untouched).
    incoming: list[UploadFile] = []
    for f in (bills or []):
        if f and f.filename:
            incoming.append(f)
    if bill and bill.filename:
        incoming.append(bill)
    if len(incoming) > MAX_BILLS_PER_EXPENSE:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            f"At most {MAX_BILLS_PER_EXPENSE} bills can be attached to one expense.",
        )

    stored_bills: list[dict] = []
    for f in incoming:
        suffix = Path(f.filename).suffix.lower()
        if suffix not in ALLOWED_BILL_EXTENSIONS:
            raise HTTPException(
                status.HTTP_400_BAD_REQUEST,
                f"Bill {f.filename!r}: must be one of {sorted(ALLOWED_BILL_EXTENSIONS)}",
            )
        content_type = (f.content_type or "").lower()
        if content_type and content_type not in ALLOWED_BILL_MIMETYPES:
            raise HTTPException(
                status.HTTP_400_BAD_REQUEST,
                f"Bill {f.filename!r}: content-type {content_type!r} not allowed.",
            )
        data = await f.read()
        if len(data) > MAX_BILL_BYTES:
            raise HTTPException(
                status.HTTP_400_BAD_REQUEST,
                f"Bill {f.filename!r}: exceeds {MAX_BILL_BYTES // (1024 * 1024)} MB limit.",
            )
        object_key = f"expenses/{uuid.uuid4().hex}{suffix}"
        try:
            storage.put_object(
                object_key,
                data,
                content_type=content_type or "application/octet-stream",
            )
        except Exception as e:
            # Best-effort cleanup of any earlier bills we already stored so
            # we don't leave orphans on failure mid-upload.
            for prior in stored_bills:
                try:
                    storage.delete_object(prior["object_key"])
                except Exception:
                    pass
            raise HTTPException(
                status.HTTP_500_INTERNAL_SERVER_ERROR,
                f"Could not store bill {f.filename!r}: {e}",
            )
        stored_bills.append({"filename": f.filename, "object_key": object_key})

    exp = Expense(
        user_id=user.id,
        date=date,
        mode=mode,
        expense_type=expense_type,
        travel_type=travel_type,
        amount=amount,
        advance=max(0, int(advance or 0)),
        site_name=(site_name or "").strip()[:255],
        remarks=(remarks or "").strip(),
        bills=stored_bills,
        status="pending",
    )
    db.add(exp)
    db.commit()
    db.refresh(exp)
    return _to_out(exp)


@router.get("", response_model=list[ExpenseOut])
def list_expenses(
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """List expenses visible to the caller.

    - HR + named approvers (Tarini / Smita) see ALL expenses.
    - Finance approvers (Shivangi) see ONLY `approved` + `paid` rows —
      pending / onhold / rejected expenses are filtered out so her view
      is just her disbursal queue + history.
    - Project coordinators (Arman) see ONLY their team members' expenses.
    - Regular employees see only their own expenses.
    """
    q = db.query(Expense).join(User, Expense.user_id == User.id)
    team = _coordinator_team(user)
    if _is_approver(user) or _is_finance_approver(user):
        pass  # see everything — finance approver needs full data for advance table
    elif team is not None:
        # The coordinator's team (read-only) PLUS their own expenses, so they
        # can still file and view their own under "My Expenses".
        q = q.filter(
            (func.lower(User.email).in_(team)) | (Expense.user_id == user.id)
        )
    else:
        q = q.filter(Expense.user_id == user.id)
    rows = q.order_by(Expense.created_at.desc()).limit(2000).all()
    return [_to_out(r) for r in rows]


def _serve_bill(exp: Expense, bill: dict) -> Response:
    """Stream a single bill from MinIO with the right Content-Type."""
    object_key = bill.get("object_key") or ""
    filename = bill.get("filename") or ""
    if not object_key:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Bill is missing its storage key")
    try:
        data = storage.get_object_bytes(object_key)
    except Exception as e:
        raise HTTPException(
            status.HTTP_500_INTERNAL_SERVER_ERROR, f"Could not fetch bill: {e}"
        )
    suffix = Path(filename or object_key).suffix.lower()
    mime = {
        ".jpg":  "image/jpeg",
        ".jpeg": "image/jpeg",
        ".png":  "image/png",
        ".webp": "image/webp",
        ".heic": "image/heic",
        ".pdf":  "application/pdf",
    }.get(suffix, "application/octet-stream")
    headers = {
        "Content-Disposition": (
            f'inline; filename="{filename or ("bill" + suffix)}"'
        ),
    }
    return Response(content=data, media_type=mime, headers=headers)



# ============================================================
# Monthly-summary notes (advance + remark per user × month)
# ============================================================

def _note_to_dict(n: MonthlyExpenseNote) -> dict:
    return {
        "user_id": n.user_id,
        "year": n.year,
        "month": n.month,
        "advance": n.advance or 0,
        "remark": n.remark or "",
        "updated_at": n.updated_at.isoformat() if n.updated_at else None,
    }


@router.get("/monthly-notes")
def list_monthly_notes(
    year: int = Query(...),
    month: int = Query(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """List monthly notes for a (year, month).

    Approvers (HR / Tarini / Smita) see EVERY employee's notes for the
    period.  Regular employees see only their own row — so they can read
    the HR remark for the current month on their own expense view.
    """
    if month < 1 or month > 12 or year < 2000 or year > 2100:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            f"Invalid year/month: {year}/{month}",
        )
    q = db.query(MonthlyExpenseNote).filter(
        MonthlyExpenseNote.year == year,
        MonthlyExpenseNote.month == month,
    )
    if not _is_approver(user):
        q = q.filter(MonthlyExpenseNote.user_id == user.id)
    rows = q.all()
    return {"items": [_note_to_dict(r) for r in rows]}


@router.get("/advances")
def list_advances(
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Approver-only: every advance ever recorded, newest first.

    Combines two sources so HR sees the full picture:
      - `monthly_note` — HR's per-employee per-month advance entered via
        the Monthly Summary modal.
      - `expense` — the `advance` field on an individual expense row
        (the employee reports they already received money against that
        specific claim).  Previously these were invisible on the Advances
        panel, which made employees like Rahul Sharma — who declared a
        ₹21k advance across his 21 expenses — appear with zero advance.

    Returned items share a common shape; `source` tells the frontend
    where the row came from so it can label the period column accordingly.
    """
    if not (_is_approver(user) or _is_finance_approver(user)):
        raise HTTPException(
            status.HTTP_403_FORBIDDEN,
            "Only HR / approvers can view the advances list.",
        )

    note_rows = (
        db.query(MonthlyExpenseNote)
        .join(User, MonthlyExpenseNote.user_id == User.id)
        .filter(MonthlyExpenseNote.advance > 0)
        .all()
    )
    expense_rows = (
        db.query(Expense)
        .join(User, Expense.user_id == User.id)
        .filter(Expense.advance > 0)
        .all()
    )

    items: list[dict] = []
    for n in note_rows:
        items.append({
            "source": "monthly_note",
            "user_id": n.user_id,
            "user_name": _full_name(n.user) or "—",
            "department": _dept_name(n.user) or "",
            "year": n.year,
            "month": n.month,
            "advance": n.advance or 0,
            "updated_at": n.updated_at.isoformat() if n.updated_at else None,
        })
    for e in expense_rows:
        items.append({
            "source": "expense",
            "expense_id": e.id,
            "user_id": e.user_id,
            "user_name": _full_name(e.user) or "—",
            "department": _dept_name(e.user) or "",
            "year": e.date.year if e.date else None,
            "month": e.date.month if e.date else None,
            "advance": e.advance or 0,
            # Use the expense's own date as the "Date of Given" — that's
            # the day the advance was tied to.  Falls back to created_at
            # so the row still sorts correctly when date is missing.
            "updated_at": (
                e.date.isoformat() if e.date
                else (e.created_at.isoformat() if e.created_at else None)
            ),
            "expense_status": e.status,
            "expense_type": e.expense_type or "",
        })

    # Newest first.  Missing timestamps go to the bottom.
    items.sort(key=lambda it: (it.get("updated_at") or ""), reverse=True)
    return {"items": items}


@router.put("/monthly-notes")
def upsert_monthly_note(
    payload: dict = Body(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """HR-only: create or update a (user, year, month) note.

    Body shape:
        { "user_id": 17, "year": 2026, "month": 6,
          "advance": 500, "remark": "On hold — awaiting director approval" }

    Either `advance` or `remark` is optional but at least one should be
    sent (sending both is fine).  Missing means "leave that field as-is".
    """
    if not _is_approver(user):
        raise HTTPException(
            status.HTTP_403_FORBIDDEN,
            "Only HR / approvers can edit monthly notes.",
        )
    try:
        target_user_id = int(payload.get("user_id"))
        year = int(payload.get("year"))
        month = int(payload.get("month"))
    except (TypeError, ValueError):
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            "user_id, year, and month are required integers.",
        )
    if month < 1 or month > 12 or year < 2000 or year > 2100:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            f"Invalid year/month: {year}/{month}",
        )

    # Find or create the note row.
    note = (
        db.query(MonthlyExpenseNote)
        .filter(
            MonthlyExpenseNote.user_id == target_user_id,
            MonthlyExpenseNote.year == year,
            MonthlyExpenseNote.month == month,
        )
        .first()
    )
    if note is None:
        note = MonthlyExpenseNote(
            user_id=target_user_id,
            year=year,
            month=month,
            advance=0,
            remark="",
        )
        db.add(note)

    if "advance" in payload and payload["advance"] is not None:
        try:
            adv = max(0, int(float(payload["advance"])))
        except (TypeError, ValueError):
            adv = 0
        note.advance = adv
    if "remark" in payload and payload["remark"] is not None:
        note.remark = str(payload["remark"])[:2048]
    note.updated_by_id = user.id

    db.commit()
    db.refresh(note)
    return _note_to_dict(note)


# ── Employee self-records an advance received ────────────────────────────────

@router.post("/record-advance", response_model=ExpenseOut, status_code=status.HTTP_201_CREATED)
def record_advance(
    payload: dict = Body(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Employee records an advance they received from HR.

    Creates a synthetic expense row with amount=0 and advance=<value>
    so it appears in the advance column of the expense table.
    Body: { "amount": int, "date": "YYYY-MM-DD" | null, "note": str | null }
    """
    try:
        amount = max(1, int(float(payload.get("amount") or 0)))
    except (TypeError, ValueError):
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "amount must be a positive integer")
    raw_date = payload.get("date") or None
    note = str(payload.get("note") or "")[:500]
    if raw_date:
        try:
            from datetime import date as _date
            exp_date = _date.fromisoformat(raw_date)
        except ValueError:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, "date must be YYYY-MM-DD")
    else:
        from datetime import date as _date
        exp_date = _date.today()
    exp = Expense(
        user_id=user.id,
        expense_type="others",
        travel_type="",
        mode="",
        amount=0,
        advance=amount,
        site_name="Advance received",
        remarks=note,
        date=exp_date,
        status="pending",
    )
    db.add(exp)
    db.commit()
    db.refresh(exp)
    return _to_out(exp)


# ── HR / approver issues an advance to a specific employee ───────────────────

@router.post("/advance-issue", response_model=ExpenseOut, status_code=status.HTTP_201_CREATED)
def advance_issue(
    payload: dict = Body(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """HR issues an advance to a specific employee.

    Body: { "employee_id": int, "amount": int, "date": "YYYY-MM-DD" | null,
            "note": str | null, "approved_by": str | null,
            "paid_date": "YYYY-MM-DD" | null, "paid_amount": int | null }

    Status follows what HR recorded: a paid_date means the money already went
    out ("paid"); otherwise naming an authoriser marks it "approved"; with
    neither it lands in the normal "pending" queue.
    """
    if not _is_approver(user):
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Only HR / approvers can issue advances.")
    try:
        employee_id = int(payload.get("employee_id") or 0)
        amount = max(1, int(float(payload.get("amount") or 0)))
    except (TypeError, ValueError):
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "employee_id and amount are required integers")
    target = db.get(User, employee_id)
    if target is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Employee not found")
    raw_date = payload.get("date") or None
    note = str(payload.get("note") or "")[:500]
    approved_by = str(payload.get("approved_by") or "").strip()[:120]

    from datetime import date as _date

    def _parse_date(raw: str, label: str) -> _date:
        try:
            return _date.fromisoformat(raw)
        except ValueError:
            raise HTTPException(
                status.HTTP_400_BAD_REQUEST, f"{label} must be YYYY-MM-DD"
            )

    exp_date = _parse_date(raw_date, "date") if raw_date else _date.today()

    raw_paid_date = payload.get("paid_date") or None
    paid_on = _parse_date(raw_paid_date, "paid_date") if raw_paid_date else None

    paid_amount = None
    if payload.get("paid_amount") not in (None, ""):
        try:
            paid_amount = int(float(payload.get("paid_amount")))
        except (TypeError, ValueError):
            raise HTTPException(status.HTTP_400_BAD_REQUEST, "paid_amount must be an integer")
        if paid_amount < 0:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, "paid_amount must be non-negative")

    if paid_on is not None:
        # Money is recorded as already disbursed.  Default the paid amount to
        # the full advance when HR didn't type a different figure.
        new_status = "paid"
        if paid_amount is None:
            paid_amount = amount
    elif approved_by:
        new_status = "approved"
    else:
        new_status = "pending"
        paid_amount = None

    now = datetime.now(timezone.utc)
    exp = Expense(
        user_id=employee_id,
        expense_type="others",
        travel_type="",
        mode="",
        amount=0,
        advance=amount,
        site_name="Advance issued",
        remarks=note,
        date=exp_date,
        status=new_status,
        approved_by_name=approved_by,
        paid_amount=paid_amount,
        # Stamp the audit trail so the row matches what a normal
        # approve / mark-paid would have produced.
        decided_by_id=user.id if new_status in ("approved", "paid") else None,
        decided_at=now if new_status in ("approved", "paid") else None,
        paid_by_id=user.id if new_status == "paid" else None,
        paid_at=(
            datetime(paid_on.year, paid_on.month, paid_on.day, tzinfo=timezone.utc)
            if paid_on is not None
            else None
        ),
    )
    db.add(exp)
    db.commit()
    db.refresh(exp)
    return _to_out(exp)


@router.get("/{expense_id}/bill/{index}")
def download_bill_at(
    expense_id: int,
    index: int,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Stream a specific bill (by position in the `bills` list)."""
    exp = db.query(Expense).filter(Expense.id == expense_id).first()
    if not exp:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Expense not found")
    if (
        exp.user_id != user.id
        and not _is_approver(user)
        and not _is_finance_approver(user)
        and not _coordinator_may_view(user, exp, db)
    ):
        raise HTTPException(
            status.HTTP_403_FORBIDDEN, "You can't view this expense's bills"
        )
    raw_bills = exp.bills or []
    if index < 0 or index >= len(raw_bills):
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Bill index out of range")
    return _serve_bill(exp, raw_bills[index])


@router.get("/{expense_id}/bill")
def download_bill_legacy(
    expense_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Backward-compat: serve the FIRST bill when no index is given."""
    exp = db.query(Expense).filter(Expense.id == expense_id).first()
    if not exp:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Expense not found")
    if (
        exp.user_id != user.id
        and not _is_approver(user)
        and not _is_finance_approver(user)
        and not _coordinator_may_view(user, exp, db)
    ):
        raise HTTPException(
            status.HTTP_403_FORBIDDEN, "You can't view this expense's bills"
        )
    raw_bills = exp.bills or []
    if not raw_bills:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "No bill attached")
    return _serve_bill(exp, raw_bills[0])


@router.post("/{expense_id}/decide", response_model=ExpenseOut)
def decide_expense(
    expense_id: int,
    payload: ExpenseDecideIn,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Approve or reject an expense.  HR + named approvers only."""
    if not _is_approver(user):
        raise HTTPException(
            status.HTTP_403_FORBIDDEN,
            "Only HR / approvers (TARINI, SMITA) can decide expenses.",
        )
    decision = (payload.decision or "").strip().lower()
    if decision not in ("approved", "rejected", "onhold"):
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            "decision must be 'approved', 'rejected', or 'onhold'",
        )
    exp = db.query(Expense).filter(Expense.id == expense_id).first()
    if not exp:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Expense not found")
    exp.status = decision
    exp.decided_by_id = user.id
    exp.decided_at = datetime.now(timezone.utc)
    exp.decision_note = (payload.note or "").strip()
    db.commit()
    db.refresh(exp)
    return _to_out(exp)


@router.post("/{expense_id}/mark-paid", response_model=ExpenseOut)
def mark_paid(
    expense_id: int,
    payload: dict = Body(default={}),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Finance-approver-only: mark an approved expense as paid.

    Optional body: { "paid_date": "YYYY-MM-DD", "payment_ref": "UTR123..." }
    """
    if not _is_finance_approver(user):
        raise HTTPException(
            status.HTTP_403_FORBIDDEN,
            "Only the finance approver can mark expenses as paid.",
        )
    exp = db.query(Expense).filter(Expense.id == expense_id).first()
    if not exp:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Expense not found")
    if exp.status != "approved":
        raise HTTPException(
            status.HTTP_409_CONFLICT,
            f"Only approved expenses can be marked paid (this is {exp.status!r}).",
        )
    exp.status = "paid"
    exp.paid_by_id = user.id
    paid_date_str = (payload or {}).get("paid_date")
    if paid_date_str:
        try:
            from datetime import date as date_cls
            d = date_cls.fromisoformat(paid_date_str)
            exp.paid_at = datetime(d.year, d.month, d.day, tzinfo=timezone.utc)
        except (ValueError, TypeError):
            exp.paid_at = datetime.now(timezone.utc)
    else:
        exp.paid_at = datetime.now(timezone.utc)
    exp.payment_ref = (payload or {}).get("payment_ref", "") or ""
    db.commit()
    db.refresh(exp)
    return _to_out(exp)


@router.patch("/{expense_id}", response_model=ExpenseOut)
def update_expense(
    expense_id: int,
    payload: ExpensePatchIn,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Owner edits their own expense.  Allowed while the expense is pending,
    on hold, or rejected — editing a rejected row resubmits it (see below).
    Once approved / paid the row is locked.

    Bills attached to the expense are NOT touched here — bill add/remove
    flows go through the upload + delete endpoints.
    """
    exp = db.query(Expense).filter(Expense.id == expense_id).first()
    if not exp:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Expense not found")
    if exp.user_id != user.id and not _is_hr(user):
        raise HTTPException(
            status.HTTP_403_FORBIDDEN,
            "You can only edit your own expenses (HR can edit any).",
        )
    if exp.status not in ("pending", "onhold", "rejected"):
        raise HTTPException(
            status.HTTP_409_CONFLICT,
            "This expense has already been approved and can no longer be edited.",
        )

    # Apply each provided field with the same validation rules as create.
    if payload.date is not None:
        exp.date = payload.date
    if payload.mode is not None:
        m = (payload.mode or "").strip().lower()
        if m and m not in ALLOWED_MODES:
            raise HTTPException(
                status.HTTP_400_BAD_REQUEST,
                f"Invalid mode. Allowed: {sorted(ALLOWED_MODES - {''})}",
            )
        exp.mode = m
    if payload.expense_type is not None:
        et = (payload.expense_type or "").strip().lower()
        if et not in ALLOWED_EXPENSE_TYPES:
            raise HTTPException(
                status.HTTP_400_BAD_REQUEST,
                f"Invalid expense_type. Allowed: {sorted(ALLOWED_EXPENSE_TYPES)}",
            )
        exp.expense_type = et
    if payload.travel_type is not None:
        tt = (payload.travel_type or "").strip().lower()
        if tt and tt not in ALLOWED_TRAVEL_TYPES:
            raise HTTPException(
                status.HTTP_400_BAD_REQUEST,
                f"Invalid travel_type. Allowed: {sorted(ALLOWED_TRAVEL_TYPES)}",
            )
        exp.travel_type = tt
    if payload.amount is not None:
        if payload.amount < 0:
            raise HTTPException(
                status.HTTP_400_BAD_REQUEST,
                "amount must be a non-negative integer",
            )
        exp.amount = payload.amount
    if payload.advance is not None:
        if payload.advance < 0:
            raise HTTPException(
                status.HTTP_400_BAD_REQUEST,
                "advance must be a non-negative integer",
            )
        exp.advance = payload.advance
    if payload.site_name is not None:
        exp.site_name = (payload.site_name or "").strip()[:255]
    if payload.remarks is not None:
        exp.remarks = (payload.remarks or "").strip()

    # Editing an on-hold or rejected expense resubmits it: back to pending so
    # the approver sees a fresh request to review.  The prior decision (and the
    # rejection note explaining what to fix) is cleared along with it.
    if exp.status in ("onhold", "rejected"):
        exp.status = "pending"
        exp.decided_by_id = None
        exp.decided_at = None
        exp.decision_note = ""

    db.commit()
    db.refresh(exp)
    return _to_out(exp)


@router.delete("/{expense_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_expense(
    expense_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Owner can withdraw a PENDING expense; HR can delete any."""
    exp = db.query(Expense).filter(Expense.id == expense_id).first()
    if not exp:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Expense not found")
    is_owner = exp.user_id == user.id
    if not (_is_hr(user) or (is_owner and exp.status == "pending")):
        raise HTTPException(
            status.HTTP_403_FORBIDDEN,
            "Only the owner (while pending) or HR can delete this expense.",
        )
    # Best-effort cleanup of every attached bill — swallow errors so the
    # row still deletes even if MinIO is briefly unreachable.
    for bill in (exp.bills or []):
        key = bill.get("object_key") or ""
        if key:
            try:
                storage.delete_object(key)
            except Exception:
                pass
    db.delete(exp)
    db.commit()
    return Response(status_code=status.HTTP_204_NO_CONTENT)


def _assert_can_edit_bills(exp: Expense, user: User) -> None:
    """Same gate as PATCH /expenses/{id} — owner or HR, status pending/onhold/
    rejected.  Rejected is allowed so the employee can attach the bill they
    were asked for before resubmitting."""
    if exp.user_id != user.id and not _is_hr(user):
        raise HTTPException(
            status.HTTP_403_FORBIDDEN,
            "You can only manage bills on your own expenses (HR can edit any).",
        )
    if exp.status not in ("pending", "onhold", "rejected"):
        raise HTTPException(
            status.HTTP_409_CONFLICT,
            "This expense has already been decided and its bills are locked.",
        )


@router.post("/{expense_id}/bills", response_model=ExpenseOut)
async def add_bills(
    expense_id: int,
    bills: list[UploadFile] = File(default=[]),
    bill: UploadFile | None = File(None),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Append one or more bill files to an existing pending/onhold expense.

    Mirrors the validation in `create_expense` — same extensions, mime
    types, size cap, and the per-expense max of 10 bills.
    """
    exp = db.query(Expense).filter(Expense.id == expense_id).first()
    if not exp:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Expense not found")
    _assert_can_edit_bills(exp, user)

    incoming: list[UploadFile] = []
    for f in (bills or []):
        if f and f.filename:
            incoming.append(f)
    if bill and bill.filename:
        incoming.append(bill)
    if not incoming:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            "No bill files were uploaded.",
        )

    existing = list(exp.bills or [])
    if len(existing) + len(incoming) > MAX_BILLS_PER_EXPENSE:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            f"At most {MAX_BILLS_PER_EXPENSE} bills can be attached "
            f"({len(existing)} already attached).",
        )

    stored_new: list[dict] = []
    for f in incoming:
        suffix = Path(f.filename).suffix.lower()
        if suffix not in ALLOWED_BILL_EXTENSIONS:
            for prior in stored_new:
                try: storage.delete_object(prior["object_key"])
                except Exception: pass
            raise HTTPException(
                status.HTTP_400_BAD_REQUEST,
                f"Bill {f.filename!r}: must be one of {sorted(ALLOWED_BILL_EXTENSIONS)}",
            )
        content_type = (f.content_type or "").lower()
        if content_type and content_type not in ALLOWED_BILL_MIMETYPES:
            for prior in stored_new:
                try: storage.delete_object(prior["object_key"])
                except Exception: pass
            raise HTTPException(
                status.HTTP_400_BAD_REQUEST,
                f"Bill {f.filename!r}: content-type {content_type!r} not allowed.",
            )
        data = await f.read()
        if len(data) > MAX_BILL_BYTES:
            for prior in stored_new:
                try: storage.delete_object(prior["object_key"])
                except Exception: pass
            raise HTTPException(
                status.HTTP_400_BAD_REQUEST,
                f"Bill {f.filename!r}: exceeds {MAX_BILL_BYTES // (1024 * 1024)} MB limit.",
            )
        object_key = f"expenses/{uuid.uuid4().hex}{suffix}"
        try:
            storage.put_object(
                object_key,
                data,
                content_type=content_type or "application/octet-stream",
            )
        except Exception as e:
            for prior in stored_new:
                try: storage.delete_object(prior["object_key"])
                except Exception: pass
            raise HTTPException(
                status.HTTP_500_INTERNAL_SERVER_ERROR,
                f"Could not store bill {f.filename!r}: {e}",
            )
        stored_new.append({"filename": f.filename, "object_key": object_key})

    # JSONB columns need a NEW list assignment so SQLAlchemy notices the change.
    exp.bills = existing + stored_new
    db.commit()
    db.refresh(exp)
    return _to_out(exp)


@router.delete("/{expense_id}/bill/{index}", response_model=ExpenseOut)
def delete_bill(
    expense_id: int,
    index: int,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Remove one attached bill by its position in the expense's bills list."""
    exp = db.query(Expense).filter(Expense.id == expense_id).first()
    if not exp:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Expense not found")
    _assert_can_edit_bills(exp, user)

    bills = list(exp.bills or [])
    if index < 0 or index >= len(bills):
        raise HTTPException(
            status.HTTP_404_NOT_FOUND, f"No bill at index {index}",
        )
    removed = bills.pop(index)
    # Best-effort MinIO cleanup — the DB write is the source of truth.
    key = removed.get("object_key") or ""
    if key:
        try:
            storage.delete_object(key)
        except Exception:
            pass
    exp.bills = bills
    db.commit()
    db.refresh(exp)
    return _to_out(exp)


# ============================================================
# Monthly summary Excel export
# ============================================================

def _month_bounds(year: int, month: int) -> tuple[date_type, date_type]:
    if month < 1 or month > 12 or year < 2000 or year > 2100:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            f"Invalid year/month: {year}/{month}",
        )
    first = date_type(year, month, 1)
    last = (
        date_type(year + 1, 1, 1) if month == 12 else date_type(year, month + 1, 1)
    ) - timedelta(days=1)
    return first, last


@router.post("/monthly-summary.xlsx")
def monthly_summary_xlsx(
    payload: dict = Body(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """One-tab Excel of monthly expense summary per employee.

    Body shape:
        {
          "year": 2026,
          "month": 6,                  # 1..12
          "advances": { "<user_id>": 500, ... }   # optional advance per emp
        }

    Columns:
        Employee | Total Amount | Status | Advance | Subtotal | Submit Date

    Footer row: Total Amount  ·  Total Advance  ·  Subtotal.

    Restricted to HR + named approvers (Tarini / Smita).
    """
    if not _is_approver(user):
        raise HTTPException(
            status.HTTP_403_FORBIDDEN,
            "Only HR / approvers can download the monthly summary.",
        )
    year = int(payload.get("year") or 0)
    month = int(payload.get("month") or 0)
    advances_in = payload.get("advances") or {}
    if not isinstance(advances_in, dict):
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST, "`advances` must be an object"
        )
    advances: dict[int, int] = {}
    for k, v in advances_in.items():
        try:
            advances[int(k)] = int(float(v or 0))
        except (TypeError, ValueError):
            continue

    first, last = _month_bounds(year, month)

    # All expenses in the month (any status).
    rows = (
        db.query(Expense)
        .join(User, Expense.user_id == User.id)
        .filter(Expense.date >= first, Expense.date <= last)
        .order_by(User.first_name, Expense.created_at)
        .all()
    )

    # Group by user_id.
    by_user: dict[int, dict] = {}
    for r in rows:
        u = r.user
        full_name = _full_name(u) or "—"
        g = by_user.setdefault(r.user_id, {
            "name": full_name,
            "total": 0,
            "pending": 0,
            "approved": 0,
            "rejected": 0,
            "onhold": 0,
            "latest_submit": r.created_at,
        })
        g["total"] += (r.amount or 0)
        if r.status in g:
            g[r.status] += 1
        if r.created_at and (g["latest_submit"] is None or r.created_at > g["latest_submit"]):
            g["latest_submit"] = r.created_at

    # ---- Build the workbook ----
    wb = Workbook()
    ws = wb.active
    month_label = first.strftime("%B %Y")  # e.g. "June 2026"
    ws.title = f"Expenses {first.strftime('%b %Y')}"[:31]

    title_font = Font(bold=True, size=14, color="1B5E8B")
    subtitle_font = Font(size=10, color="666666", italic=True)
    header_font = Font(bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1A2A3A")
    total_font = Font(bold=True, size=10, color="1B5E8B")
    total_fill = PatternFill("solid", fgColor="EAF1F8")
    body_font = Font(size=10, color="1A1A1A")
    border = Border(*[Side(style="thin", color="C9D2DC")] * 4)
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    ws["B1"] = f"Monthly Expense Summary — {month_label}"
    ws["B1"].font = title_font
    ws.merge_cells("B1:G1")
    ws["B2"] = (
        f"{len(by_user)} employee{'s' if len(by_user) != 1 else ''}"
        f"  ·  {len(rows)} expense{'s' if len(rows) != 1 else ''}"
        f"  ·  Generated: {date_type.today().strftime('%d %b %Y')}"
    )
    ws["B2"].font = subtitle_font
    ws.merge_cells("B2:G2")

    headers = ["Employee", "Total Amount", "Status", "Advance", "Subtotal", "Submit Date"]
    widths = [28, 18, 22, 14, 16, 18]
    ws.column_dimensions["A"].width = 3
    for i, (h, w) in enumerate(zip(headers, widths), start=2):
        c = ws.cell(row=4, column=i, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[4].height = 24

    sorted_groups = sorted(by_user.items(), key=lambda kv: kv[1]["name"].lower())

    total_amount_sum = 0
    total_advance_sum = 0
    total_subtotal_sum = 0
    row_idx = 5
    for uid, g in sorted_groups:
        amt = g["total"]
        adv = advances.get(uid, 0)
        sub = amt - adv
        total_amount_sum += amt
        total_advance_sum += adv
        total_subtotal_sum += sub
        status_bits = []
        if g["pending"]:  status_bits.append(f"Pending: {g['pending']}")
        if g["onhold"]:   status_bits.append(f"On Hold: {g['onhold']}")
        if g["approved"]: status_bits.append(f"Approved: {g['approved']}")
        if g["rejected"]: status_bits.append(f"Rejected: {g['rejected']}")
        status_text = ", ".join(status_bits) or "—"
        submit_str = (
            g["latest_submit"].strftime("%d %b %Y")
            if g["latest_submit"] else "—"
        )

        cells = [
            (g["name"], left),
            (f"₹{amt:,.0f}", right),
            (status_text, left),
            (f"₹{adv:,.0f}", right),
            (f"₹{sub:,.0f}", right),
            (submit_str, center),
        ]
        for col_offset, (val, align) in enumerate(cells):
            c = ws.cell(row=row_idx, column=2 + col_offset, value=val)
            c.font = body_font
            c.alignment = align
            c.border = border
        row_idx += 1

    # Footer totals row.
    if sorted_groups:
        ws.cell(row=row_idx, column=2, value="Total")
        ws.cell(row=row_idx, column=3, value=f"₹{total_amount_sum:,.0f}")
        ws.cell(row=row_idx, column=4, value="")
        ws.cell(row=row_idx, column=5, value=f"₹{total_advance_sum:,.0f}")
        ws.cell(row=row_idx, column=6, value=f"₹{total_subtotal_sum:,.0f}")
        ws.cell(row=row_idx, column=7, value="")
        for col in range(2, 8):
            c = ws.cell(row=row_idx, column=col)
            c.font = total_font
            c.fill = total_fill
            c.border = border
            c.alignment = right if col in (3, 5, 6) else left

    # Stream the file out.
    buf = io.BytesIO()
    wb.save(buf)
    payload_bytes = buf.getvalue()
    filename = f"monthly-expenses-{first.strftime('%Y-%m')}.xlsx"
    return Response(
        content=payload_bytes,
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Content-Length": str(len(payload_bytes)),
        },
    )


@router.post("/department-summary.xlsx")
def department_summary_xlsx(
    payload: dict = Body(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Multi-sheet Excel of monthly expenses, with an Overview tab + one
    tab per department.

    Body shape:
        { "year": 2026, "month": 6 }

    Overview sheet:
        - Grand total across the company
        - Per-department breakdown: department, # employees, # expenses, total
        - Per-employee summary: name, department, total, status mix, last submit

    One sheet per department (only departments that have expenses in the
    period):
        - Every individual expense as a row:
          Date | Employee | Type | Mode | Amount | Status | Submit Date | Remarks

    Restricted to HR + named approvers (Tarini / Smita).
    """
    if not _is_approver(user):
        raise HTTPException(
            status.HTTP_403_FORBIDDEN,
            "Only HR / approvers can download the department summary.",
        )
    year = int(payload.get("year") or 0)
    month = int(payload.get("month") or 0)
    first, last = _month_bounds(year, month)

    rows = (
        db.query(Expense)
        .join(User, Expense.user_id == User.id)
        .filter(Expense.date >= first, Expense.date <= last)
        .order_by(User.first_name, Expense.date)
        .all()
    )

    UNASSIGNED = "Unassigned"
    by_dept: dict[str, list[Expense]] = {}
    by_user: dict[int, list[Expense]] = {}
    employees_per_dept: dict[str, set[int]] = {}
    employee_totals: dict[int, dict] = {}
    grand_total = 0

    for r in rows:
        dept = _dept_name(r.user) or UNASSIGNED
        by_dept.setdefault(dept, []).append(r)
        by_user.setdefault(r.user_id, []).append(r)
        employees_per_dept.setdefault(dept, set()).add(r.user_id)
        grand_total += (r.amount or 0)

        g = employee_totals.setdefault(r.user_id, {
            "name": _full_name(r.user) or "—",
            "dept": dept,
            "total": 0,
            "pending": 0,
            "approved": 0,
            "rejected": 0,
            "onhold": 0,
            "latest_submit": r.created_at,
        })
        g["total"] += (r.amount or 0)
        if r.status in g:
            g[r.status] += 1
        if r.created_at and (g["latest_submit"] is None or r.created_at > g["latest_submit"]):
            g["latest_submit"] = r.created_at

    # ---- Shared styles ----
    title_font = Font(bold=True, size=14, color="1B5E8B")
    subtitle_font = Font(size=10, color="666666", italic=True)
    header_font = Font(bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1A2A3A")
    total_font = Font(bold=True, size=10, color="1B5E8B")
    total_fill = PatternFill("solid", fgColor="EAF1F8")
    section_font = Font(bold=True, size=11, color="1A1A1A")
    body_font = Font(size=10, color="1A1A1A")
    border = Border(*[Side(style="thin", color="C9D2DC")] * 4)
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    wb = Workbook()
    overview = wb.active
    month_label = first.strftime("%B %Y")
    overview.title = "Overview"

    overview["B1"] = f"Expense Summary — {month_label}"
    overview["B1"].font = title_font
    overview.merge_cells("B1:F1")
    overview["B2"] = (
        f"Total Expense: ₹{grand_total:,.0f}"
        f"  ·  {len(employee_totals)} employee{'s' if len(employee_totals) != 1 else ''}"
        f"  ·  {len(by_dept)} department{'s' if len(by_dept) != 1 else ''}"
        f"  ·  {len(rows)} expense{'s' if len(rows) != 1 else ''}"
        f"  ·  Generated: {date_type.today().strftime('%d %b %Y')}"
    )
    overview["B2"].font = subtitle_font
    overview.merge_cells("B2:F2")
    overview.column_dimensions["A"].width = 3

    # Section 1 — per-department breakdown.
    overview["B4"] = "By Department"
    overview["B4"].font = section_font

    dept_headers = ["Department", "Employees", "Expenses", "Total Amount"]
    dept_widths = [28, 14, 14, 18]
    for i, (h, w) in enumerate(zip(dept_headers, dept_widths), start=2):
        c = overview.cell(row=5, column=i, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border
        overview.column_dimensions[get_column_letter(i)].width = w
    overview.row_dimensions[5].height = 22

    dept_sorted = sorted(by_dept.keys(), key=lambda n: n.lower())
    row_idx = 6
    dept_total_sum = 0
    for dept in dept_sorted:
        exps = by_dept[dept]
        dept_total = sum((e.amount or 0) for e in exps)
        dept_total_sum += dept_total
        cells = [
            (dept, left),
            (len(employees_per_dept[dept]), center),
            (len(exps), center),
            (f"₹{dept_total:,.0f}", right),
        ]
        for col_offset, (val, align) in enumerate(cells):
            c = overview.cell(row=row_idx, column=2 + col_offset, value=val)
            c.font = body_font
            c.alignment = align
            c.border = border
        row_idx += 1

    if dept_sorted:
        overview.cell(row=row_idx, column=2, value="Total")
        overview.cell(row=row_idx, column=3, value=len(employee_totals))
        overview.cell(row=row_idx, column=4, value=len(rows))
        overview.cell(row=row_idx, column=5, value=f"₹{dept_total_sum:,.0f}")
        for col in range(2, 6):
            c = overview.cell(row=row_idx, column=col)
            c.font = total_font
            c.fill = total_fill
            c.border = border
            c.alignment = right if col in (3, 4, 5) else left
        row_idx += 2

    # Section 2 — per-employee summary (all employees, all departments).
    overview.cell(row=row_idx, column=2, value="By Employee").font = section_font
    row_idx += 1

    emp_headers = ["Employee", "Department", "Total Amount", "Last Submit"]
    emp_widths = [24, 22, 16, 18]
    for i, (h, w) in enumerate(zip(emp_headers, emp_widths), start=2):
        c = overview.cell(row=row_idx, column=i, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border
        # Don't shrink columns already wider from the dept table above.
        cur_w = overview.column_dimensions[get_column_letter(i)].width or 0
        if w > cur_w:
            overview.column_dimensions[get_column_letter(i)].width = w
    overview.row_dimensions[row_idx].height = 22
    row_idx += 1

    emp_sorted = sorted(
        employee_totals.items(),
        key=lambda kv: (kv[1]["dept"].lower(), kv[1]["name"].lower()),
    )
    for _uid, g in emp_sorted:
        cells = [
            (g["name"], left),
            (g["dept"], left),
            (f"₹{g['total']:,.0f}", right),
            (g["latest_submit"].strftime("%d %b %Y") if g["latest_submit"] else "—", center),
        ]
        for col_offset, (val, align) in enumerate(cells):
            c = overview.cell(row=row_idx, column=2 + col_offset, value=val)
            c.font = body_font
            c.alignment = align
            c.border = border
        row_idx += 1

    # ---- Per-employee sheets (one tab per employee with expenses). ----
    # Sort by (department, name) so the tab order mirrors the Overview
    # By-Employee table — easier to navigate when there are many tabs.
    user_sorted = sorted(
        by_user.keys(),
        key=lambda uid: (
            (employee_totals.get(uid, {}).get("dept") or "").lower(),
            (employee_totals.get(uid, {}).get("name") or "").lower(),
        ),
    )
    used_titles: set[str] = {"Overview"}
    for uid in user_sorted:
        info = employee_totals.get(uid, {})
        emp_name = info.get("name") or "—"
        emp_dept = info.get("dept") or ""

        # Sheet titles capped at 31 chars by Excel; sanitize illegal chars.
        clean = "".join(ch for ch in emp_name if ch not in r":\/?*[]") or "Employee"
        title = clean[:31] or "Employee"
        base = title
        suffix = 2
        while title in used_titles:
            tail = f"-{suffix}"
            title = (base[: 31 - len(tail)]) + tail
            suffix += 1
        used_titles.add(title)

        ws = wb.create_sheet(title=title)
        emp_total = sum((e.amount or 0) for e in by_user[uid])
        ws["B1"] = f"{emp_name} — {month_label}"
        ws["B1"].font = title_font
        ws.merge_cells("B1:I1")
        ws["B2"] = (
            f"Total: ₹{emp_total:,.0f}"
            f"  ·  {emp_dept or 'Unassigned'}"
            f"  ·  {len(by_user[uid])} expense{'s' if len(by_user[uid]) != 1 else ''}"
        )
        ws["B2"].font = subtitle_font
        ws.merge_cells("B2:I2")
        ws.column_dimensions["A"].width = 3

        # Employee column is dropped — the whole tab is one employee, so
        # repeating the name on every row was redundant.
        headers = [
            "Date", "Site", "Type", "Mode",
            "Amount", "Status", "Submit Date", "Remarks",
        ]
        widths = [12, 22, 16, 10, 14, 12, 14, 40]
        for i, (h, w) in enumerate(zip(headers, widths), start=2):
            c = ws.cell(row=4, column=i, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = center
            c.border = border
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.row_dimensions[4].height = 24

        sorted_exps = sorted(
            by_user[uid],
            key=lambda e: (e.date or date_type.min),
        )
        r_idx = 5
        for e in sorted_exps:
            etype = e.expense_type or ""
            if e.travel_type:
                etype = f"{etype} ({e.travel_type})" if etype else e.travel_type
            cells = [
                (e.date.strftime("%d %b %Y") if e.date else "—", center),
                (e.site_name or "—", left),
                (etype or "—", left),
                ((e.mode or "—"), left),
                (f"₹{(e.amount or 0):,.0f}", right),
                ((e.status or "").capitalize() or "—", left),
                (e.created_at.strftime("%d %b %Y") if e.created_at else "—", center),
                (e.remarks or "", left),
            ]
            for col_offset, (val, align) in enumerate(cells):
                c = ws.cell(row=r_idx, column=2 + col_offset, value=val)
                c.font = body_font
                c.alignment = align
                c.border = border
            r_idx += 1

        # Footer total row for the employee.  Amount column sits at col 6
        # after dropping the Employee column.
        if sorted_exps:
            ws.cell(row=r_idx, column=2, value="Total")
            ws.cell(row=r_idx, column=6, value=f"₹{emp_total:,.0f}")
            for col in range(2, 10):
                c = ws.cell(row=r_idx, column=col)
                c.font = total_font
                c.fill = total_fill
                c.border = border
                c.alignment = right if col == 6 else left

    buf = io.BytesIO()
    wb.save(buf)
    payload_bytes = buf.getvalue()
    filename = f"expense-departments-{first.strftime('%Y-%m')}.xlsx"
    return Response(
        content=payload_bytes,
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Content-Length": str(len(payload_bytes)),
        },
    )


