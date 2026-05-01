"""
Replace dummy seed data with real employees from employees.xlsx.

Two-step usage from the container:

    docker compose exec django python manage.py shell -c \
        "exec(open('/app/import_employees.py').read())"

By default DRY_RUN = True — first run prints what *would* happen without
touching the database. Once you're happy, set DRY_RUN = False below and
run the same command again.
"""
import re
from collections import OrderedDict

from django.contrib.auth import get_user_model
from django.db import transaction

from departments.models import Department
from reports.models import DailyReport

# ============================================================================
# CONFIG — flip this to False when you're ready to actually write
# ============================================================================
DRY_RUN = False
EXCEL_PATH = "/app/employees.xlsx"

# Default 5-field report template every newly-created department gets.
# You can refine each department's fields later in Django admin.
GENERIC_FIELDS = [
    {"key": "workDone", "label": "Work Done"},
    {"key": "workInProgress", "label": "Work in Progress"},
    {"key": "upcomingPriorities", "label": "Upcoming Priorities"},
    {"key": "challenges", "label": "Challenges Faced / Support Needed"},
    {"key": "otherUpdate", "label": "Other Update"},
]

# Excel-name -> (slug, display_name, color).  "R & D" and "R& D" map to the
# SAME slug (typo unification).  "R & D-BESS" stays separate per the user.
DEPT_MAP = {
    "BESS-Sales":  ("bessSales",  "BESS Sales",  "amber"),
    "Design":      ("design",     "Design",      "indigo"),
    "Finance":     ("finance",    "Finance",     "emerald"),
    "HR":          ("hrDept",     "HR",          "rose"),
    "Logistics":   ("logistics",  "Logistics",   "sky"),
    "Marketing":   ("marketing",  "Marketing",   "indigo"),
    "O & M":       ("om",         "O & M",       "zinc"),
    "Procurement": ("procurement","Procurement", "emerald"),
    "Production":  ("production", "Production",  "amber"),
    "Project":     ("project",    "Project",     "rose"),
    "R & D":       ("rd",         "R & D",       "indigo"),
    "R& D":        ("rd",         "R & D",       "indigo"),  # typo alias
    "R & D-BESS":  ("rdBess",     "R & D-BESS",  "sky"),
    "Sales":       ("sales",      "Sales",       "rose"),
    "Service":     ("service",    "Service",     "emerald"),
    "Support":     ("support",    "Support",     "amber"),
    "Web Dev":     ("webDev",     "Web Dev",     "indigo"),
}

# Dummy employees we seeded earlier — these get deleted (and their reports).
DUMMY_USERNAMES = {
    "divya", "arjun", "pooja", "tarini", "naveen", "ishita", "kabir", "riya",
    "vivek", "anjali", "mohit",
}

User = get_user_model()


def step(label):
    bar = "=" * 70
    print()
    print(bar)
    print(label)
    print(bar)


def _slugify_local(s):
    """Lowercase + strip non-alphanumeric for the email local-part."""
    return re.sub(r"[^a-z0-9]", "", (s or "").lower())


def _split_name(full_name):
    parts = (full_name or "").strip().split()
    if not parts:
        return "", ""
    if len(parts) == 1:
        return parts[0], ""
    return parts[0], " ".join(parts[1:])


def _read_excel():
    import openpyxl  # imported lazily so missing pkg gives a clean message
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    rows = []  # list of dicts so call sites read the fields by name

    def _to_date(v):
        if v is None:
            return None
        if hasattr(v, "date"):  # datetime
            return v.date()
        return v  # already a date

    # Sheet 1: 'Ornate Ram Ware'  cols: Organisation, S.No, Name, Dept, RM, DOJ
    ws1 = wb["Ornate Ram Ware"]
    for row in ws1.iter_rows(min_row=2, values_only=True):
        org, sn, name, dept, rm, doj = row
        if not name:
            continue
        rows.append({
            "sheet": "Ornate Ram Ware",
            "name": str(name).strip(),
            "dept_excel": (dept or "").strip(),
            "title": "",
            "organisation": (org or "").strip(),
            "reporting_manager": (rm or "").strip(),
            "date_of_joining": _to_date(doj),
        })

    # Sheet 2: 'SG Ornate'  cols: NAME, EMP CODE, DESIGNATION, DEPT, RM, DOJ
    ws2 = wb["SG Ornate"]
    for row in ws2.iter_rows(min_row=2, values_only=True):
        name, code, des, dept, rm, doj = row
        if not name:
            continue
        rows.append({
            "sheet": "SG Ornate",
            "name": str(name).strip(),
            "dept_excel": (dept or "").strip(),
            "title": (des or "").strip(),
            "organisation": "SG Ornate",  # sheet name doubles as the org for SG sheet
            "reporting_manager": (rm or "").strip(),
            "date_of_joining": _to_date(doj),
        })

    return rows


# ---------------------------------------------------------------------------
# Action
# ---------------------------------------------------------------------------

def main():
    print(f"\nDRY_RUN = {DRY_RUN}    (set DRY_RUN = False inside the script to actually write)")

    # ------------- 1. Read Excel -------------
    step("1. Reading Excel")
    excel_rows = _read_excel()
    print(f"  {len(excel_rows)} rows with a name")

    # Validate every row's department is mapped
    unknown_depts = sorted({r["dept_excel"] for r in excel_rows if r["dept_excel"] not in DEPT_MAP})
    if unknown_depts:
        print(f"  ERROR: these dept names from the Excel are not in DEPT_MAP:")
        for d in unknown_depts:
            print(f"    - {d!r}")
        print("  Aborting.")
        return

    # ------------- 2. Plan deletions -------------
    step("2. Plan: delete dummy data")
    dummy_users = User.objects.filter(username__in=DUMMY_USERNAMES)
    dummy_user_count = dummy_users.count()
    dummy_report_count = DailyReport.objects.filter(user__in=dummy_users).count()
    inside_sales_dept = Department.objects.filter(slug="insideSales").first()

    print(f"  Dummy employees to delete:  {dummy_user_count}")
    print(f"  Their seeded reports:       {dummy_report_count}")
    if inside_sales_dept:
        in_use = User.objects.filter(department=inside_sales_dept).exclude(username__in=DUMMY_USERNAMES).count()
        print(f"  Department 'insideSales':   delete (currently {in_use} non-dummy users)")
    else:
        print(f"  Department 'insideSales':   already gone")

    # ------------- 3. Plan dept creates -------------
    step("3. Plan: ensure all departments exist")
    # Distinct (slug, name, color) tuples we need
    needed = OrderedDict()
    for excel_name, (slug, display, color) in DEPT_MAP.items():
        needed[slug] = (display, color)
    existing = {d.slug for d in Department.objects.all()}
    for slug, (display, color) in needed.items():
        if slug in existing:
            print(f"  exists      {slug!r:<14} -> {display}")
        else:
            print(f"  WILL CREATE {slug!r:<14} -> {display}  (color: {color})")

    # ------------- 4. Plan user creates -------------
    step("4. Plan: import employees (with email + password)")
    taken_emails = set()
    # Pre-seed taken_emails with all currently-existing emails (e.g. admin)
    for e in User.objects.exclude(username__in=DUMMY_USERNAMES).values_list("email", flat=True):
        if e:
            taken_emails.add(e.lower())

    # Pre-load existing imported users so re-runs become updates (matched by full name within an organisation).
    # We use first_name + last_name as the natural key since email was generated by us and could change
    # if a collision was resolved differently last time.
    existing_by_fullname = {}
    for u in User.objects.exclude(role="hr").exclude(username__in=DUMMY_USERNAMES):
        key = (u.first_name.strip().lower(), u.last_name.strip().lower())
        existing_by_fullname[key] = u

    plan = []  # list of dicts
    for r in excel_rows:
        first, last = _split_name(r["name"])
        local = _slugify_local(first)
        if not local:
            print(f"  SKIP (no usable first name): {r['name']!r}")
            continue
        existing = existing_by_fullname.get((first.lower(), last.lower()))
        if existing:
            email = existing.email
        else:
            # Find a non-colliding email
            candidate = f"{local}@ornatesolar.com"
            n = 2
            while candidate in taken_emails:
                candidate = f"{local}{n}@ornatesolar.com"
                n += 1
            taken_emails.add(candidate)
            email = candidate
        password = f"{local}@ornate"
        slug = DEPT_MAP[r["dept_excel"]][0]
        plan.append({
            "sheet": r["sheet"],
            "name": r["name"],
            "first": first,
            "last": last,
            "email": email,
            "password": password,
            "dept_slug": slug,
            "title": r["title"],
            "organisation": r["organisation"],
            "reporting_manager": r["reporting_manager"],
            "date_of_joining": r["date_of_joining"],
            "existing": existing,
        })

    creates = sum(1 for p in plan if not p["existing"])
    updates = sum(1 for p in plan if p["existing"])
    print(f"  Plan: {creates} new, {updates} update existing  (total {len(plan)})")
    print(f"  First 5 mappings:")
    for p in plan[:5]:
        marker = "UPD" if p["existing"] else "NEW"
        doj = p["date_of_joining"].isoformat() if p["date_of_joining"] else "—"
        print(f"    [{marker}] {p['name']:<28} {p['email']:<28} dept={p['dept_slug']:<11} org={p['organisation']!r:<18} rm={p['reporting_manager']!r:<22} doj={doj}")
    if len(plan) > 5:
        print(f"    ... +{len(plan) - 5} more")

    # ------------- 5. Execute (only if not dry-run) -------------
    if DRY_RUN:
        step("DRY RUN COMPLETE — nothing was written")
        print("Set DRY_RUN = False at top of this file and re-run to apply.")
        return

    step("5. Executing")
    with transaction.atomic():
        # 5a — delete dummy employees + their reports (cascade handles reports)
        deleted_users, _ = dummy_users.delete()
        print(f"  Deleted dummy employees + reports: {deleted_users} rows")

        # 5b — delete insideSales (if no real users in it)
        if inside_sales_dept and not User.objects.filter(department=inside_sales_dept).exists():
            inside_sales_dept.delete()
            print("  Deleted department 'insideSales'")

        # 5c — create / update departments
        slug_to_dept = {}
        for slug, (display, color) in needed.items():
            # On create: seed with the generic 5-field template.
            # On update: keep whatever report_fields HR has customised
            # (only sync name + color from the import config).
            dept, created = Department.objects.get_or_create(
                slug=slug,
                defaults={
                    "name": display,
                    "color": color,
                    "report_fields": GENERIC_FIELDS,
                },
            )
            if not created:
                dept.name = display
                dept.color = color
                dept.save()
            slug_to_dept[slug] = dept
            print(f"  {'created' if created else 'updated'} dept: {slug}")

        # 5d — create or update employees
        created_count = 0
        updated_count = 0
        for p in plan:
            user = p["existing"]
            is_new = user is None
            if is_new:
                user = User(
                    username=p["email"].split("@")[0],
                    email=p["email"],
                    is_active=True,
                    is_staff=False,
                    role="employee",
                )
                user.set_password(p["password"])
            user.first_name = p["first"]
            user.last_name = p["last"]
            user.title = p["title"] or ""
            user.department = slug_to_dept[p["dept_slug"]]
            user.organisation = p["organisation"] or ""
            user.reporting_manager = p["reporting_manager"] or ""
            user.date_of_joining = p["date_of_joining"]
            user.save()
            if is_new:
                created_count += 1
            else:
                updated_count += 1
        print(f"  Employees created: {created_count}, updated: {updated_count}")

    step("DONE")
    print(f"  Total employees now: {User.objects.filter(role='employee').count()}")
    print(f"  Total departments:   {Department.objects.count()}")
    print(f"  Total reports:       {DailyReport.objects.count()}")


main()
